#!/usr/bin/env python3
"""
cm_data_collector.py  —  Continuous SNMP + Kafka polling collector for LLD test sessions.

  vCMTS  →  Kafka (DS + US latency/throughput) + SNMP (modem-side US stats)
  iCMTS  →  SNMP only (modem US + iCMTS DS)

Wraps the existing snmp_collector.py (SSH SNMP) and kafka_collector.py (Kafka)
into continuous poll threads that write time-series CSVs compatible with
metrics_pdf_report.py.

Output structure:
    Results/
    └── <MAC>_<cmts_type>/
        └── <YYYYMMDD_HHMMSS>/
            ├── raw_data/
            │   ├── snmp_us_<MAC>_<ts>.csv
            │   ├── snmp_ds_<MAC>_<ts>.csv   (iCMTS only)
            │   ├── kafka_<MAC>_<ts>.csv     (vCMTS only)
            │   ├── Kafka_Raw_Messages_<MAC>_<ts>.txt
            │   └── snmp_polls/

Usage:
    python cm_data_collector.py
    python cm_data_collector.py --debug
"""
import re
import csv
import os
import sys
import time
import threading
import logging
from datetime import datetime, timezone

from config_loader import config
from snmp_collector import ssh_snmp_collector, parse_modem_info
from kafka_collector import CmtsCollector, METRIC_NAMES, NUM_BINS, _PROM_RE, _parse_labels

HERE = os.path.dirname(os.path.abspath(__file__))
DEFAULT_RESULTS_DIR = os.path.join(HERE, 'Results')

# ---------------------------------------------------------------------------
# Config defaults
# ---------------------------------------------------------------------------
DEFAULT_SNMP_JUMPSERVER    = config.snmp_jumpserver    or ''
DEFAULT_SNMP_USERNAME      = config.snmp_username      or ''
DEFAULT_MODEM_COMMUNITY    = config.snmp_community     or 'open'
DEFAULT_ICMTS_COMMUNITY    = config.icmts_community    or 'NMISread'
DEFAULT_ICMTS_TARGET_IP    = config.icmts_ip           or ''
DEFAULT_CMTS_HOST          = (config.get('cmts_hosts', default=[{}]) or [{}])[0].get('name', '')
DEFAULT_TACACS_PASSWORD    = config.vcmts_password     or ''
DEFAULT_VCMTS_HOST         = config.get('vcmts', 'host', default='')
DEFAULT_KAFKA_BROKER       = config.get('kafka', 'broker', default='')
DEFAULT_KAFKA_TOPIC        = config.get('kafka', 'topic', default='')
DEFAULT_SNMP_TIMEOUT       = config.snmp_timeout
DEFAULT_SNMP_RETRIES       = config.snmp_retries
DEFAULT_SNMP_POLL_INTERVAL = int(os.environ.get('SNMP_POLL_INTERVAL', '15'))

# ---------------------------------------------------------------------------
# CSV field schemas  (same as CM_Collector — metrics_pdf_report.py reads these)
# ---------------------------------------------------------------------------
SNMP_CSV_FIELDS = [
    'captured_utc', 'poll_index', 'target_ip', 'target_label', 'cmts_type', 'sfid',
    'sf_direction', 'sf_primary', 'sf_agg_sfid', 'sf_buffer_size',
    'ps_scn', 'ps_priority', 'ps_max_rate', 'ps_max_rate_64', 'ps_max_burst',
    'ps_max_concat_burst', 'ps_aqm_latency_target',
    'ps_min_buffer', 'ps_target_buffer', 'ps_max_buffer',
    'flow_pkts', 'flow_octets', 'flow_policed_drop', 'flow_policed_delay', 'flow_aqm_drop',
    'lat_bin_scn', 'lat_aqm_target',
    'lat_edge_bin1',  'lat_edge_bin2',  'lat_edge_bin3',  'lat_edge_bin4',
    'lat_edge_bin5',  'lat_edge_bin6',  'lat_edge_bin7',  'lat_edge_bin8',
    'lat_edge_bin9',  'lat_edge_bin10', 'lat_edge_bin11', 'lat_edge_bin12',
    'lat_edge_bin13', 'lat_edge_bin14', 'lat_edge_bin15',
    'lat_max_usec', 'lat_updates',
    'lat_bin1',  'lat_bin2',  'lat_bin3',  'lat_bin4',
    'lat_bin5',  'lat_bin6',  'lat_bin7',  'lat_bin8',
    'lat_bin9',  'lat_bin10', 'lat_bin11', 'lat_bin12',
    'lat_bin13', 'lat_bin14', 'lat_bin15', 'lat_bin16',
    'cong_aqm_drop', 'cong_scn_marked', 'cong_ce_marked', 'cong_sanctioned',
    'cong_ect0', 'cong_ect1', 'cong_ce_ect1', 'cong_arrived_ce',
]

KAFKA_CSV_FIELDS = [
    'captured_utc', 'kafka_timestamp_ms', 'dir', 'sfIndex', 'sfid', 'scn',
    'mdName', 'node', 'pod', 'cluster',
    'delta_octets', 'delta_pkts', 'delta_pkts_dropped',
    'total_octets', 'total_pkts',
    'lat_avg_usec', 'lat_max_usec',
    'aqm_drop_pkts', 'aqm_marked_pkts', 'sanctioned_pkts',
    'lat_bin01', 'lat_bin02', 'lat_bin03', 'lat_bin04',
    'lat_bin05', 'lat_bin06', 'lat_bin07', 'lat_bin08',
    'lat_bin09', 'lat_bin10', 'lat_bin11', 'lat_bin12',
    'lat_bin13', 'lat_bin14', 'lat_bin15', 'lat_bin16',
    'lat_bin_edge01', 'lat_bin_edge02', 'lat_bin_edge03', 'lat_bin_edge04',
    'lat_bin_edge05', 'lat_bin_edge06', 'lat_bin_edge07', 'lat_bin_edge08',
    'lat_bin_edge09', 'lat_bin_edge10', 'lat_bin_edge11', 'lat_bin_edge12',
    'lat_bin_edge13', 'lat_bin_edge14', 'lat_bin_edge15', 'lat_bin_edge16',
    'max_rate_bps', 'aqm_target_msecs',
]

# ---------------------------------------------------------------------------
# OID → CSV column maps  (docsIf3CmService .21.1.*)
# ---------------------------------------------------------------------------
_OID_COL_MAP = {
    '3.1.7':  'sf_direction',   '3.1.8':  'sf_primary',
    '3.1.19': 'sf_agg_sfid',    '3.1.17': 'sf_buffer_size',
    '4.1.1':  'flow_pkts',      '4.1.2':  'flow_octets',
    '4.1.6':  'flow_policed_drop', '4.1.7': 'flow_policed_delay',
    '4.1.8':  'flow_aqm_drop',
    '29.1.1.2':  'lat_bin_scn',
    '29.1.1.3':  'lat_edge_bin1',  '29.1.1.4':  'lat_edge_bin2',
    '29.1.1.5':  'lat_edge_bin3',  '29.1.1.6':  'lat_edge_bin4',
    '29.1.1.7':  'lat_edge_bin5',  '29.1.1.8':  'lat_edge_bin6',
    '29.1.1.9':  'lat_edge_bin7',  '29.1.1.10': 'lat_edge_bin8',
    '29.1.1.11': 'lat_edge_bin9',  '29.1.1.12': 'lat_edge_bin10',
    '29.1.1.13': 'lat_edge_bin11', '29.1.1.14': 'lat_edge_bin12',
    '29.1.1.15': 'lat_edge_bin13', '29.1.1.16': 'lat_edge_bin14',
    '29.1.1.17': 'lat_edge_bin15', '29.1.1.18': 'lat_aqm_target',
    '29.2.1.1':  'lat_max_usec',   '29.2.1.2':  'lat_updates',
    '29.2.1.3':  'lat_bin1',  '29.2.1.4':  'lat_bin2',
    '29.2.1.5':  'lat_bin3',  '29.2.1.6':  'lat_bin4',
    '29.2.1.7':  'lat_bin5',  '29.2.1.8':  'lat_bin6',
    '29.2.1.9':  'lat_bin7',  '29.2.1.10': 'lat_bin8',
    '29.2.1.11': 'lat_bin9',  '29.2.1.12': 'lat_bin10',
    '29.2.1.13': 'lat_bin11', '29.2.1.14': 'lat_bin12',
    '29.2.1.15': 'lat_bin13', '29.2.1.16': 'lat_bin14',
    '29.2.1.17': 'lat_bin15', '29.2.1.18': 'lat_bin16',
    '30.1.1': 'cong_aqm_drop',   '30.1.2': 'cong_scn_marked',
    '30.1.3': 'cong_ce_marked',  '30.1.4': 'cong_sanctioned',
    '30.1.5': 'cong_ect0',       '30.1.6': 'cong_ect1',
    '30.1.7': 'cong_ce_ect1',    '30.1.8': 'cong_arrived_ce',
}

_OID_PARAM_MAP = {
    '2.1.4':  'ps_scn',          '2.1.5':  'ps_priority',
    '2.1.6':  'ps_max_rate',     '2.1.7':  'ps_max_burst',
    '2.1.8':  'ps_max_concat_burst',
    '2.1.39': 'ps_min_buffer',   '2.1.40': 'ps_target_buffer',
    '2.1.41': 'ps_max_buffer',   '2.1.43': 'ps_aqm_latency_target',
    '2.1.44': 'ps_max_rate_64',
}

# Human-readable labels for OID suffixes in .txt poll files
_OID_LABEL_MAP = {
    # SF Table (.21.1.3)
    '3.1.6':  'sf_agg_sfid',         '3.1.7':  'sf_direction',
    '3.1.8':  'sf_primary',          '3.1.9':  'sf_sid_cluster',
    '3.1.10': 'sf_active',           '3.1.11': 'sf_admitted',
    '3.1.12': 'sf_pkts_dropped',     '3.1.13': 'sf_buffer_size',
    '3.1.14': 'sf_max_burst',        '3.1.15': 'sf_max_rate',
    '3.1.16': 'sf_aqm_target',       '3.1.17': 'sf_buffer_size_bytes',
    '3.1.18': 'sf_scn',              '3.1.19': 'sf_agg_sfid_2',
    '3.1.20': 'sf_active_2',         '3.1.21': 'sf_admitted_2',
    # Param Set (.21.1.2)
    '2.1.4':  'ps_scn',              '2.1.5':  'ps_priority',
    '2.1.6':  'ps_max_rate_bps',     '2.1.7':  'ps_max_burst',
    '2.1.8':  'ps_max_concat_burst', '2.1.9':  'ps_min_reserved_rate',
    '2.1.10': 'ps_min_reserved_pkt', '2.1.11': 'ps_active_timeout',
    '2.1.12': 'ps_admitted_timeout', '2.1.13': 'ps_direction',
    '2.1.14': 'ps_tos_overwrite',    '2.1.20': 'ps_aqm_flags',
    '2.1.21': 'ps_aqm_reserved',     '2.1.22': 'ps_aqm_drop_policy',
    '2.1.24': 'ps_aqm_target_delay', '2.1.25': 'ps_aqm_config',
    '2.1.27': 'ps_aqm_interval',     '2.1.28': 'ps_aqm_burst',
    '2.1.29': 'ps_aqm_ecn',          '2.1.30': 'ps_aqm_drop_rate',
    '2.1.31': 'ps_aqm_num_bins',     '2.1.32': 'ps_aqm_bin_width',
    '2.1.37': 'ps_aqm_reserved2',    '2.1.38': 'ps_aqm_enable',
    '2.1.39': 'ps_min_buffer',       '2.1.40': 'ps_target_buffer',
    '2.1.41': 'ps_max_buffer',       '2.1.42': 'ps_aqm_mode',
    '2.1.43': 'ps_aqm_latency_target_usec', '2.1.44': 'ps_max_rate_64',
    '2.1.47': 'ps_aqm_type',         '2.1.48': 'ps_aqm_reserved3',
    '2.1.49': 'ps_aqm_reserved4',    '2.1.50': 'ps_aqm_reserved5',
    '2.1.51': 'ps_aqm_interval2',    '2.1.52': 'ps_aqm_bin_count',
    '2.1.53': 'ps_aqm_reserved6',    '2.1.54': 'ps_aqm_ecn_mode',
    # Flow Stats (.21.1.4)
    '4.1.1':  'flow_pkts',           '4.1.2':  'flow_octets',
    '4.1.3':  'flow_elapsed_time',   '4.1.4':  'flow_time_active',
    '4.1.6':  'flow_policed_drop',   '4.1.7':  'flow_policed_delay',
    '4.1.8':  'flow_aqm_drop',
    # Lat Edges (.21.1.29.1)
    '29.1.1.1':  'lat_sfid_type',    '29.1.1.2':  'lat_bin_scn',
    '29.1.1.3':  'lat_edge_bin1_usec',  '29.1.1.4':  'lat_edge_bin2_usec',
    '29.1.1.5':  'lat_edge_bin3_usec',  '29.1.1.6':  'lat_edge_bin4_usec',
    '29.1.1.7':  'lat_edge_bin5_usec',  '29.1.1.8':  'lat_edge_bin6_usec',
    '29.1.1.9':  'lat_edge_bin7_usec',  '29.1.1.10': 'lat_edge_bin8_usec',
    '29.1.1.11': 'lat_edge_bin9_usec',  '29.1.1.12': 'lat_edge_bin10_usec',
    '29.1.1.13': 'lat_edge_bin11_usec', '29.1.1.14': 'lat_edge_bin12_usec',
    '29.1.1.15': 'lat_edge_bin13_usec', '29.1.1.16': 'lat_edge_bin14_usec',
    '29.1.1.17': 'lat_edge_bin15_usec', '29.1.1.18': 'lat_aqm_target_usec',
    # Lat Stats (.21.1.29.2)
    '29.2.1.1':  'lat_max_usec',     '29.2.1.2':  'lat_updates',
    '29.2.1.3':  'lat_bin1_pkts',    '29.2.1.4':  'lat_bin2_pkts',
    '29.2.1.5':  'lat_bin3_pkts',    '29.2.1.6':  'lat_bin4_pkts',
    '29.2.1.7':  'lat_bin5_pkts',    '29.2.1.8':  'lat_bin6_pkts',
    '29.2.1.9':  'lat_bin7_pkts',    '29.2.1.10': 'lat_bin8_pkts',
    '29.2.1.11': 'lat_bin9_pkts',    '29.2.1.12': 'lat_bin10_pkts',
    '29.2.1.13': 'lat_bin11_pkts',   '29.2.1.14': 'lat_bin12_pkts',
    '29.2.1.15': 'lat_bin13_pkts',   '29.2.1.16': 'lat_bin14_pkts',
    '29.2.1.17': 'lat_bin15_pkts',   '29.2.1.18': 'lat_bin16_pkts',
    # Congestion (.21.1.30)
    '30.1.1': 'cong_aqm_drop',       '30.1.2': 'cong_scn_marked',
    '30.1.3': 'cong_ce_marked',      '30.1.4': 'cong_sanctioned',
    '30.1.5': 'cong_ect0',           '30.1.6': 'cong_ect1',
    '30.1.7': 'cong_ce_ect1',        '30.1.8': 'cong_arrived_ce',
    # DS SF Index (.21.1.11)
    '11.1.1': 'ds_sf_ifindex',       '11.1.2': 'ds_sf_direction',
    '11.1.3': 'ds_sf_primary',
}

_RE_OID_TXT = re.compile(
    r'SNMPv2-SMI::enterprises\.4491\.2\.1\.21\.1\.'
    r'(\d+(?:\.\d+)*)\s*=\s*(\S+):\s*(.*)'
)

# Used by _pivot_snmp_output to extract OID suffix and value for CSV pivot
_RE_OID_SFID = re.compile(
    r'SNMPv2-SMI::enterprises\.4491\.2\.1\.21\.1\.'
    r'(\d+(?:\.\d+)*)\s*=\s*\S+:\s*(.*)'
)


def _label_snmp_line(line):
    """Replace raw OID prefix with human-readable field name, keep sfid suffix and value."""
    m = _RE_OID_TXT.match(line.strip())
    if not m:
        return line
    oid_suffix, vtype, value = m.group(1), m.group(2), m.group(3).strip()
    parts = oid_suffix.split('.')

    # Try progressively shorter prefixes to find a label
    label = None
    sfid  = None
    for cut in range(len(parts), 0, -1):
        key = '.'.join(parts[:cut])
        if key in _OID_LABEL_MAP:
            label = _OID_LABEL_MAP[key]
            sfid  = '.'.join(parts[cut:])
            break

    if label is None:
        return line
    sfid_str = f'  sfid={parts[-1]}' if parts else ''
    return f'  {label}{sfid_str} = {value}'


# Kafka metric name → KAFKA_CSV_FIELDS column
_KAFKA_METRIC_MAP = {
    'K_Samis1_DeltaOctetsPassed':        'delta_octets',
    'K_Samis1_DeltaPacketsPassed':       'delta_pkts',
    'K_Samis1_DeltaPacketsDropped':      'delta_pkts_dropped',
    'snmp_docsQosServiceFlowOctets':     'total_octets',
    'snmp_docsQosServiceFlowPackets':    'total_pkts',
    'dp_flow_QueueLatencyAvgUsec':       'lat_avg_usec',
    'dp_flow_QueueLatencyMaxUsec':       'lat_max_usec',
    'dp_flow_AqmDroppedPackets':         'aqm_drop_pkts',
    'dp_flow_AqmMarkedCongestedPackets': 'aqm_marked_pkts',
    'dp_flow_SanctionedPackets':         'sanctioned_pkts',
}

# ---------------------------------------------------------------------------
# MAC / path helpers
# ---------------------------------------------------------------------------

def _norm_mac(mac):
    return mac.strip().replace(':', '').replace('.', '').replace('-', '').lower()

def _mac_colon(mac_norm):
    return ':'.join(mac_norm[i:i+2] for i in range(0, 12, 2))

def _mac_to_decimal(mac_norm):
    return '.'.join(str(int(mac_norm[i:i+2], 16)) for i in range(0, 12, 2))

def _norm_mac_dotted(mac_norm):
    return f'{mac_norm[0:4]}.{mac_norm[4:8]}.{mac_norm[8:12]}'

def _make_session_dir(mac_norm, cmts_type, ts_str):
    d = os.path.join(DEFAULT_RESULTS_DIR, f'{mac_norm}_{cmts_type}', ts_str)
    os.makedirs(d, exist_ok=True)
    return d

def _prompt(label, default=None):
    suffix = f' [{default}]' if default not in (None, '') else ''
    val = input(f'  {label}{suffix}: ').strip()
    return val if val else (default or '')

def _prompt_choice(label, choices, default=None):
    opts = '/'.join(choices)
    while True:
        val = input(f'  {label} ({opts}): ').strip().lower()
        if not val and default:
            return default.lower()
        if val in [c.lower() for c in choices]:
            return val
        print(f'    Please enter one of: {opts}')

def _write_csv_comment(f, mac_colon, cmts_type):
    ts = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
    f.write(f'# CM Collector — {mac_colon}  {cmts_type.upper()}  collected {ts} UTC\n')

# ---------------------------------------------------------------------------
# SNMP result pivot  (raw ssh_snmp_collector output → per-SFID row dicts)
# ---------------------------------------------------------------------------

def _build_snmp_commands(modem_ip, icmts_ip, modem_community, icmts_community, timeout, retries):
    """Return (us_cmds, us_lbls, ds_cmds, ds_lbls) for the given targets."""
    t, r = timeout, retries
    base = f'snmpwalk -v 2c -c {modem_community} -t {t} -r {r} {modem_ip}'
    # Walk only the specific param-set columns needed (type 2 = active),
    # instead of the entire .21.1.2 subtree which includes all 3 types × 54 columns.
    param_oids = [
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.4.2',   # ps_scn
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.5.2',   # ps_priority
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.6.2',   # ps_max_rate
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.7.2',   # ps_max_burst
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.8.2',   # ps_max_concat_burst
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.39.2',  # ps_min_buffer
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.40.2',  # ps_target_buffer
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.41.2',  # ps_max_buffer
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.43.2',  # ps_aqm_latency_target
        '1.3.6.1.4.1.4491.2.1.21.1.2.1.44.2',  # ps_max_rate_64
    ]
    us_cmds = [
        f"{base} 1.3.6.1.4.1.4491.2.1.21.1.3",
        ' && '.join(f"{base} {o}" for o in param_oids),
        f"{base} 1.3.6.1.4.1.4491.2.1.21.1.4",
        f"{base} 1.3.6.1.4.1.4491.2.1.21.1.29.1",
        f"{base} 1.3.6.1.4.1.4491.2.1.21.1.29.2",
        f"{base} 1.3.6.1.4.1.4491.2.1.21.1.30",
    ]
    us_lbls = ['US SF Table', 'US Param Set', 'US Flow Stats',
               'US Lat Edges', 'US Lat Stats', 'US Congestion']

    ds_cmds, ds_lbls = [], []
    if icmts_ip:
        ds_cmds = [
            f"snmpwalk -v 2c -c {icmts_community} -t {t} -r {r} {icmts_ip} 1.3.6.1.4.1.4491.2.1.21.1.11.1",
            f"snmpwalk -v 2c -c {icmts_community} -t {t} -r {r} {icmts_ip} 1.3.6.1.4.1.4491.2.1.21.1.4",
            f"snmpwalk -v 2c -c {icmts_community} -t {t} -r {r} {icmts_ip} 1.3.6.1.4.1.4491.2.1.21.1.29",
            f"snmpwalk -v 2c -c {icmts_community} -t {t} -r {r} {icmts_ip} 1.3.6.1.4.1.4491.2.1.21.1.30",
        ]
        ds_lbls = ['DS SF Index', 'DS Flow Stats', 'DS Lat Stats', 'DS Congestion']

    return us_cmds, us_lbls, ds_cmds, ds_lbls


def _pivot_snmp_output(raw_output_pairs, ts, poll_idx, target_ip, target_label, cmts_type):
    """Parse list of (label, raw_output) from ssh_snmp_collector into per-SFID row dicts."""
    sfid_rows = {}

    def _row(sfid):
        if sfid not in sfid_rows:
            sfid_rows[sfid] = {
                'captured_utc': ts, 'poll_index': poll_idx,
                'target_ip': target_ip, 'target_label': target_label,
                'cmts_type': cmts_type, 'sfid': sfid,
            }
        return sfid_rows[sfid]

    for _label, output in raw_output_pairs:
        for line in output.splitlines():
            m = _RE_OID_SFID.search(line)
            if not m:
                continue
            parts = m.group(1).split('.')
            val   = m.group(2).strip()

            # .2 param set: targeted walks return col.2.ifindex.sfid or col.2.sfid
            # The .2 (active type) is already in the OID prefix, so parts[0]=='2' with
            # len>=3 means: [type=2, ifindex, sfid] or [type=2, sfid]
            if parts[0] == '2' and len(parts) >= 3:
                # Full form: 2.1.col.2.ifindex.sfid → handled by _OID_PARAM_MAP key '2.1.col'
                # Targeted form: col.2.ifindex.sfid → parts[-1] is sfid
                field = _OID_PARAM_MAP.get('.'.join(parts[:-3])) or \
                        _OID_PARAM_MAP.get('.'.join(parts[:-2]))
                if field:
                    _row(parts[-1])[field] = val
                continue

            if len(parts) < 3:
                continue
            field = _OID_COL_MAP.get('.'.join(parts[:-2]))
            if field:
                _row(parts[-1])[field] = val

    return list(sfid_rows.values())


def _ssh_connect(jumpserver, username):
    """Open and return a new Paramiko SSH client connection."""
    import paramiko
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    key_path = config.ssh_key_path
    if os.path.exists(key_path):
        ssh.connect(jumpserver, username=username, key_filename=key_path,
                    timeout=config.ssh_connect_timeout)
    else:
        ssh.connect(jumpserver, username=username,
                    password=DEFAULT_TACACS_PASSWORD or None,
                    timeout=config.ssh_connect_timeout)
    return ssh


def _run_snmp_via_ssh(jumpserver, username, cmds, lbls):
    """Run SNMP commands using 2 SSH connections — commands split evenly across them."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    # Split commands into 2 batches; each batch runs sequentially on its own connection
    mid = (len(cmds) + 1) // 2
    batches = [(cmds[:mid], lbls[:mid]), (cmds[mid:], lbls[mid:])]

    def _run_batch(batch_cmds, batch_lbls):
        ssh = _ssh_connect(jumpserver, username)
        results = []
        try:
            for label, cmd in zip(batch_lbls, batch_cmds):
                _, stdout, stderr = ssh.exec_command(cmd)
                stdout.channel.recv_exit_status()
                out = stdout.read().decode(errors='replace')
                err = stderr.read().decode(errors='replace').strip()
                if err:
                    print(f'  [SNMP] [{label}] stderr: {err[:200]}')
                if not out.strip():
                    print(f'  [SNMP] [{label}] WARNING — empty output')
                results.append((label, out))
        finally:
            ssh.close()
        return results

    results_map = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = [ex.submit(_run_batch, bc, bl) for bc, bl in batches if bc]
        for fut in as_completed(futures):
            try:
                for label, out in fut.result():
                    results_map[label] = out
            except Exception as e:
                print(f'  [SNMP] ✗ batch failed: {e}')
    return [(lbl, results_map.get(lbl, '')) for lbl in lbls]


def _run_snmp_local(cmds, lbls):
    """Run SNMP commands locally in parallel (2 workers) via subprocess."""
    import subprocess
    from concurrent.futures import ThreadPoolExecutor, as_completed

    mid = (len(cmds) + 1) // 2
    batches = [(cmds[:mid], lbls[:mid]), (cmds[mid:], lbls[mid:])]

    def _run_batch(batch_cmds, batch_lbls):
        results = []
        for label, cmd in zip(batch_lbls, batch_cmds):
            try:
                proc = subprocess.run(cmd, shell=True, capture_output=True, timeout=120)
                results.append((label, proc.stdout.decode(errors='replace')))
            except Exception as e:
                print(f'  [SNMP] ✗ [{label}] {e}')
                results.append((label, ''))
        return results

    results_map = {}
    with ThreadPoolExecutor(max_workers=2) as ex:
        futures = [ex.submit(_run_batch, bc, bl) for bc, bl in batches if bc]
        for fut in as_completed(futures):
            for label, out in fut.result():
                results_map[label] = out
    return [(lbl, results_map.get(lbl, '')) for lbl in lbls]


def _write_snmp_txt(session_dir, ts, poll_idx, modem_ip, us_raw, ds_raw, cmts_type, label=None):
    """Write one raw SNMP .txt file per poll into session_dir/raw_data/snmp_polls/."""
    polls_dir = os.path.join(session_dir, 'raw_data', 'snmp_polls')
    os.makedirs(polls_dir, exist_ok=True)
    ts_file   = datetime.now().strftime('%Y%m%d_%H%M%S')
    prefix    = f'{label}_' if label else ''
    fname     = f'{prefix}SNMP_poll_{poll_idx:04d}_{ts_file}.txt'
    path      = os.path.join(polls_dir, fname)
    lines   = [
        f'SNMP Collection - {ts}',
        f'Target IP: {modem_ip}',
        '=' * 60,
    ]
    for lbl, output in us_raw:
        lines.append(f'\n{lbl}')
        lines.append('=' * 50)
        if output.strip():
            lines.extend(_label_snmp_line(l) for l in output.splitlines() if l.strip())
        else:
            lines.append('(no data)')
    if ds_raw:
        for lbl, output in ds_raw:
            lines.append(f'\n{lbl}')
            lines.append('=' * 50)
            if output.strip():
                lines.extend(_label_snmp_line(l) for l in output.splitlines() if l.strip())
            else:
                lines.append('(no data)')
    with open(path, 'w') as f:
        f.write('\n'.join(lines) + '\n')


# ---------------------------------------------------------------------------
# SNMP collector thread
# ---------------------------------------------------------------------------

def snmp_collector_thread(cfg, stop_event, csv_paths, poll_index_ref):
    """Continuous SNMP poll loop — writes snmp_us_*.csv and snmp_ds_*.csv."""
    jumpserver   = cfg['snmp_jumpserver']
    modem_ip     = cfg.get('target_ip', '')
    icmts_ip     = cfg.get('icmts_target', '')
    cmts_type    = cfg['cmts_type']
    # session_dir_ref is a mutable [path] list so the orchestrator can redirect
    # poll .txt files into the active scenario folder between scenarios.
    session_dir_ref  = cfg.get('session_dir_ref', [cfg.get('session_dir', '.')])
    scenario_ref     = cfg.get('scenario_ref', [''])

    if not jumpserver:
        print('[SNMP] No jumpserver configured — skipping')
        return
    if not modem_ip:
        print('[SNMP] No modem IP — skipping')
        return

    # Open CSV writers
    file_handles, writers = {}, {}
    for key, path in csv_paths.items():
        fh = open(path, 'w', newline='')
        _write_csv_comment(fh, cfg['mac_colon'], cmts_type)
        w  = csv.DictWriter(fh, fieldnames=SNMP_CSV_FIELDS)
        w.writeheader()
        file_handles[key] = fh
        writers[key]      = w

    print(f'[SNMP] Continuous polling started — interval {cfg["snmp_poll_interval"]}s  modem {modem_ip}')

    us_cmds, us_lbls, ds_cmds, ds_lbls = _build_snmp_commands(
        modem_ip,
        icmts_ip if cmts_type == 'icmts' else '',
        cfg['modem_community'], cfg['icmts_community'],
        cfg['snmp_timeout'], cfg['snmp_retries'],
    )

    interval = cfg['snmp_poll_interval']
    try:
        next_deadline = time.monotonic()
        while not stop_event.is_set():
            poll_idx          = poll_index_ref[0]
            poll_index_ref[0] += 1
            ts = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

            # US — always via SSH jump (modem IPv6 only reachable from jump server)
            us_raw  = _run_snmp_via_ssh(jumpserver, cfg['snmp_username'], us_cmds, us_lbls)
            us_rows = _pivot_snmp_output(us_raw, ts, poll_idx,
                                          modem_ip, 'modem_us', cmts_type)
            if not us_rows:
                print(f'[SNMP] Poll #{poll_idx} ERROR — no US data returned (SSH or walk failed)')
            for row in us_rows:
                writers['us'].writerow(row)
            file_handles['us'].flush()

            # DS — iCMTS only, run locally (CMTS reachable from this host)
            if cmts_type == 'icmts' and ds_cmds and 'ds' in writers:
                ds_raw  = _run_snmp_local(ds_cmds, ds_lbls)
                ds_rows = _pivot_snmp_output(ds_raw, ts, poll_idx,
                                              icmts_ip, 'icmts_ds', cmts_type)
                if not ds_rows:
                    print(f'[SNMP] Poll #{poll_idx} ERROR — no DS data returned')
                for row in ds_rows:
                    writers['ds'].writerow(row)
                file_handles['ds'].flush()
                _write_snmp_txt(session_dir_ref[0], ts, poll_idx, modem_ip, us_raw, ds_raw, cmts_type, label=scenario_ref[0] or None)
            else:
                _write_snmp_txt(session_dir_ref[0], ts, poll_idx, modem_ip, us_raw, None, cmts_type, label=scenario_ref[0] or None)

            next_deadline += interval
            sleep_s = next_deadline - time.monotonic()
            if sleep_s < 0:
                skipped = int(-sleep_s // interval)
                print(f'[SNMP] Poll #{poll_idx} took {interval - sleep_s:.1f}s — '
                      f'overran by {-sleep_s:.1f}s, skipping {skipped} deadline(s)')
                next_deadline += skipped * interval
                sleep_s = next_deadline - time.monotonic()
            stop_event.wait(timeout=max(0.0, sleep_s))
    finally:
        for fh in file_handles.values():
            fh.close()
    print('[SNMP] Done')

# ---------------------------------------------------------------------------
# Kafka collector thread  (vCMTS only)
# ---------------------------------------------------------------------------

def kafka_collector_thread(cfg, stop_event, csv_path):
    """
    Listens to Kafka using CmtsCollector's consumer loop internals.
    Writes one CSV row per (kafka_timestamp, dir, sfIndex) batch flush,
    producing the same kafka_*.csv format as CM_Collector.
    """
    try:
        from kafka import KafkaConsumer
    except ImportError:
        print('[Kafka] kafka-python not installed — pip install kafka-python')
        return

    mac_colon   = cfg['mac_colon']
    mac_b_colon = mac_colon.encode('ascii')
    mac_b_norm  = cfg['mac_norm'].encode('ascii')
    ready_event = cfg.get('kafka_ready_event')

    try:
        consumer = KafkaConsumer(
            cfg['kafka_topic'],
            bootstrap_servers=cfg['kafka_broker'],
            group_id=f'lld_cm_data_collector_{int(time.time())}',
            auto_offset_reset='latest',
            enable_auto_commit=True,
        )
    except Exception as e:
        print(f'[Kafka] Connect failed: {e}')
        return

    print(f'[Kafka] Connected  broker={cfg["kafka_broker"]}  '
          f'topic={cfg["kafka_topic"]}  mac={mac_colon}')

    # pending rows keyed by (kafka_ts_ms, dir, sfIndex)
    pending      = {}
    sfid_map     = {}   # (kafka_ts, dir, sfIndex) → sfid string
    params_map   = {}   # (dir, sfIndex) → {scn, max_rate_bps, aqm_target_msecs}
    written      = set()
    count        = 0
    raw_messages = []

    # Raw messages .txt path mirrors CmtsCollector naming convention:
    # Kafka_Raw_Messages_<test_name>_<timestamp>.txt
    # We derive it from the csv_path: kafka_<mac>_<ts>.csv → Kafka_Raw_Messages_<mac>_<ts>.txt
    csv_base  = os.path.splitext(os.path.basename(csv_path))[0]  # kafka_<mac>_<ts>
    raw_name  = 'Kafka_Raw_Messages_' + csv_base[len('kafka_'):] + '.txt'
    raw_path  = os.path.join(os.path.dirname(csv_path), raw_name)

    with open(csv_path, 'w', newline='') as f:
        _write_csv_comment(f, mac_colon, cfg['cmts_type'])
        writer = csv.DictWriter(f, fieldnames=KAFKA_CSV_FIELDS, extrasaction='ignore')
        writer.writeheader()

        def _flush(current_ts=None):
            nonlocal count
            done = [k for k in pending
                    if current_ts is None or k[0] != current_ts]
            for key in done:
                if key in written:
                    del pending[key]
                    continue
                row = pending.pop(key)
                row['sfid'] = sfid_map.get(key, '')
                p = params_map.get((key[1], key[2]), {})
                row.setdefault('scn',              p.get('scn', ''))
                row.setdefault('max_rate_bps',     p.get('max_rate_bps', ''))
                row.setdefault('aqm_target_msecs', p.get('aqm_target_msecs', ''))
                writer.writerow(row)
                written.add(key)
                count += 1
            if done:
                f.flush()

        while not stop_event.is_set():
            batch = consumer.poll(timeout_ms=2000)
            if not batch:
                continue
            current_kafka_ts = None

            for tp, messages in batch.items():
                for message in messages:
                    if stop_event.is_set():
                        break
                    raw = message.value
                    # Quick MAC filter before decode
                    if mac_b_norm not in raw and mac_b_colon not in raw:
                        continue
                    line = raw.decode('utf-8', errors='replace').strip()

                    # Use CmtsCollector's compiled regex
                    m = _PROM_RE.match(line)
                    if not m:
                        continue
                    metric, label_str, value_str, ts_str = m.groups()
                    if metric not in METRIC_NAMES and metric not in _KAFKA_METRIC_MAP:
                        continue

                    # Signal ready on first valid message for this MAC
                    if ready_event and not ready_event.is_set():
                        ready_event.set()
                        print(f'[Kafka] First poll received — starting test')

                    labels = _parse_labels(label_str)

                    # MAC filter via label
                    if labels.get('cmMacAddr', '').lower() != mac_colon:
                        continue

                    dir_   = labels.get('dir', '')
                    sfidx  = labels.get('sfIndex', '')
                    key    = (ts_str, dir_, sfidx)
                    current_kafka_ts = ts_str

                    # K_DocsQos_Params — params only, no row
                    if metric == 'K_DocsQos_Params':
                        p = params_map.setdefault((dir_, sfidx), {})
                        if labels.get('scn'):            p['scn']              = labels['scn']
                        if labels.get('maxRateBps'):     p['max_rate_bps']     = labels['maxRateBps']
                        if labels.get('aqmTargetMsecs'): p['aqm_target_msecs'] = labels['aqmTargetMsecs']
                        continue

                    if not dir_ or not sfidx:
                        continue

                    if key not in pending:
                        pending[key] = {
                            'captured_utc':       datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S'),
                            'kafka_timestamp_ms': ts_str,
                            'dir':                dir_,
                            'sfIndex':            sfidx,
                            'mdName':             labels.get('mdName', ''),
                            'node':               labels.get('node', ''),
                            'pod':                labels.get('pod', ''),
                            'cluster':            labels.get('cluster', ''),
                        }

                    row = pending[key]

                    # Store raw line for debug .txt
                    raw_messages.append(line)

                    # SFID + SCN from K_Samis1_Sfid
                    if metric == 'K_Samis1_Sfid':
                        sfid = str(int(float(value_str)))
                        row['sfid'] = sfid
                        sfid_map[key] = sfid
                        scn = labels.get('scn', '')
                        if scn:
                            row['scn'] = scn
                            params_map.setdefault((dir_, sfidx), {})['scn'] = scn
                        continue

                    # Latency bins
                    if metric == 'dp_flow_QueueLatencyBinPktCount':
                        bin_n = labels.get('bin', '')
                        if bin_n:
                            row[f'lat_bin{bin_n}'] = value_str
                            upper = labels.get('edgeUpperMsec', '')
                            if upper:
                                row[f'lat_bin_edge{bin_n}'] = upper
                        continue

                    col = _KAFKA_METRIC_MAP.get(metric)
                    if col:
                        row[col] = value_str

            _flush(current_kafka_ts)

        _flush()

    consumer.close()
    print(f'[Kafka] Done — {count} rows → {os.path.basename(csv_path)}')

    # Write raw messages .txt  (same format as CmtsCollector._write_raw_messages)
    if raw_messages:
        ts_str = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')
        with open(raw_path, 'w') as rf:
            rf.write(f'# Raw Kafka messages for MAC: {mac_colon} direction: both\n')
            rf.write(f'# Collected: {ts_str}\n')
            rf.write(f'# Total messages: {len(raw_messages)}\n')
            rf.write(f'# {"="*80}\n')
            for msg in raw_messages:
                rf.write(msg + '\n')
        print(f'[Kafka] Raw messages → {os.path.basename(raw_path)} ({len(raw_messages)} lines)')

# ---------------------------------------------------------------------------
# Excel time-series report  (written after collection stops)
# Keeps: Modem Info, TimeSeries, Throughput intervals, per-SF bin sheets,
#        Bin_Edges.  Calculations (percentiles, weighted avg) go in the PDF.
# ---------------------------------------------------------------------------

def generate_excel_timeseries(session_dir, mac_colon, cmts_type, kafka_csv=None, snmp_us_csv=None):
    """
    Build Excel workbook from collected CSVs.
    - TimeSeries  : every poll row from snmp_us CSV + kafka CSV
    - Throughput  : per-poll octet deltas from kafka (or snmp flow_octets)
    - Bin_Edges   : bin edge values per SFID from kafka lat_bin columns
    - SFID_<n>    : per-SFID latency bin counts across all polls
    - Modem Info  : MAC, type, session timestamps
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import pandas as pd
    except ImportError:
        print('[Excel] openpyxl or pandas not available — skipping Excel report')
        return None

    HEADER_FONT  = Font(bold=True, size=11, color='FFFFFF')
    HEADER_FILL  = PatternFill('solid', fgColor='4472C4')
    CALC_FILL    = PatternFill('solid', fgColor='D9E2F3')
    RESULT_FILL  = PatternFill('solid', fgColor='C6EFCE')
    INFO_FILL    = PatternFill('solid', fgColor='E2EFDA')
    BOLD         = Font(bold=True, size=11)
    THIN         = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
    CENTER       = Alignment(horizontal='center', vertical='center')

    def _c(ws, row, col, value, font=None, fill=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = CENTER
        cell.border = THIN
        if font: cell.font = font
        if fill: cell.fill = fill
        if fmt:  cell.number_format = fmt
        return cell

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ts_now = datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')

    # ── Modem Info ──────────────────────────────────────────────────────────
    ws_info = wb.create_sheet(title='Modem Info', index=0)
    ws_info.sheet_properties.tabColor = '70AD47'
    ws_info.merge_cells('A1:B1')
    ws_info['A1'] = 'MODEM INFORMATION'
    ws_info['A1'].font = Font(bold=True, size=14)
    ws_info['A1'].alignment = CENTER
    for i, (lbl, val) in enumerate([
        ('CM MAC',     mac_colon),
        ('CMTS Type',  cmts_type.upper()),
        ('Session Dir', session_dir),
        ('Generated',  ts_now),
    ], start=3):
        _c(ws_info, i, 1, lbl, font=BOLD, fill=HEADER_FILL)
        _c(ws_info, i, 2, val, fill=INFO_FILL)
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 55

    # ── Load CSVs ────────────────────────────────────────────────────────────
    snmp_df  = None
    kafka_df = None

    if snmp_us_csv and os.path.exists(snmp_us_csv):
        try:
            snmp_df = pd.read_csv(snmp_us_csv, comment='#')
        except Exception as e:
            print(f'[Excel] Could not read SNMP CSV: {e}')

    if kafka_csv and os.path.exists(kafka_csv):
        try:
            kafka_df = pd.read_csv(kafka_csv, comment='#')
        except Exception as e:
            print(f'[Excel] Could not read Kafka CSV: {e}')

    # ── TimeSeries sheet ─────────────────────────────────────────────────────
    ws_ts = wb.create_sheet(title='TimeSeries')
    ws_ts.sheet_properties.tabColor = 'FF6600'

    ts_headers = ['Timestamp (UTC)', 'Poll / Kafka TS', 'Source', 'SFID / sfIndex',
                  'Direction', 'Metric', 'Value']
    for col, h in enumerate(ts_headers, 1):
        _c(ws_ts, 1, col, h, font=HEADER_FONT, fill=HEADER_FILL)

    row = 2
    # SNMP rows — key metrics only
    if snmp_df is not None:
        snmp_metrics = ['flow_octets', 'flow_pkts', 'flow_aqm_drop',
                        'lat_max_usec', 'cong_aqm_drop', 'cong_ce_marked']
        for _, r in snmp_df.iterrows():
            for metric in snmp_metrics:
                if metric in snmp_df.columns and pd.notna(r.get(metric)):
                    _c(ws_ts, row, 1, str(r.get('captured_utc', '')))
                    _c(ws_ts, row, 2, str(r.get('poll_index', '')))
                    _c(ws_ts, row, 3, 'SNMP')
                    _c(ws_ts, row, 4, str(r.get('sfid', '')))
                    _c(ws_ts, row, 5, 'upstream')
                    _c(ws_ts, row, 6, metric)
                    _c(ws_ts, row, 7, r[metric])
                    row += 1

    # Kafka rows — key metrics only
    if kafka_df is not None:
        kafka_metrics = ['delta_octets', 'delta_pkts', 'total_octets',
                         'lat_avg_usec', 'lat_max_usec', 'aqm_drop_pkts', 'aqm_marked_pkts']
        for _, r in kafka_df.iterrows():
            for metric in kafka_metrics:
                if metric in kafka_df.columns and pd.notna(r.get(metric)):
                    _c(ws_ts, row, 1, str(r.get('captured_utc', '')))
                    _c(ws_ts, row, 2, str(r.get('kafka_timestamp_ms', '')))
                    _c(ws_ts, row, 3, 'Kafka')
                    _c(ws_ts, row, 4, str(r.get('sfid', r.get('sfIndex', ''))))
                    _c(ws_ts, row, 5, str(r.get('dir', '')))
                    _c(ws_ts, row, 6, metric)
                    _c(ws_ts, row, 7, r[metric])
                    row += 1

    for i, w in enumerate([22, 18, 8, 14, 12, 30, 18], 1):
        ws_ts.column_dimensions[get_column_letter(i)].width = w

    # ── Throughput sheet ─────────────────────────────────────────────────────
    ws_tp = wb.create_sheet(title='Throughput')
    ws_tp.sheet_properties.tabColor = '00B050'
    ws_tp.merge_cells('A1:G1')
    ws_tp['A1'] = f'QOS SERVICE FLOW OCTETS — {mac_colon}'
    ws_tp['A1'].font = Font(bold=True, size=14)
    ws_tp['A1'].alignment = CENTER

    tp_headers = ['Timestamp (UTC)', 'Source', 'SFID', 'Direction',
                  'Delta Octets', 'Delta Pkts', 'Rate (Mbps)']
    for col, h in enumerate(tp_headers, 1):
        _c(ws_tp, 3, col, h, font=HEADER_FONT, fill=HEADER_FILL)

    row = 4
    # Kafka delta_octets per poll
    if kafka_df is not None and 'delta_octets' in kafka_df.columns:
        for _, r in kafka_df.iterrows():
            delta_oct = pd.to_numeric(r.get('delta_octets', None), errors='coerce')
            if pd.isna(delta_oct):
                continue
            rate = delta_oct * 8 / 15 / 1_000_000  # 15s Kafka interval
            _c(ws_tp, row, 1, str(r.get('captured_utc', '')))
            _c(ws_tp, row, 2, 'Kafka')
            _c(ws_tp, row, 3, str(r.get('sfid', r.get('sfIndex', ''))))
            _c(ws_tp, row, 4, str(r.get('dir', '')))
            _c(ws_tp, row, 5, int(delta_oct), fill=CALC_FILL)
            _c(ws_tp, row, 6, int(pd.to_numeric(r.get('delta_pkts', 0), errors='coerce') or 0))
            _c(ws_tp, row, 7, round(rate, 4), fill=RESULT_FILL, fmt='0.0000')
            row += 1
    # SNMP flow_octets (cumulative — show raw value, delta computed in PDF)
    elif snmp_df is not None and 'flow_octets' in snmp_df.columns:
        for _, r in snmp_df.iterrows():
            val = pd.to_numeric(r.get('flow_octets', None), errors='coerce')
            if pd.isna(val):
                continue
            _c(ws_tp, row, 1, str(r.get('captured_utc', '')))
            _c(ws_tp, row, 2, 'SNMP')
            _c(ws_tp, row, 3, str(r.get('sfid', '')))
            _c(ws_tp, row, 4, 'upstream')
            _c(ws_tp, row, 5, int(val), fill=CALC_FILL)
            _c(ws_tp, row, 6, int(pd.to_numeric(r.get('flow_pkts', 0), errors='coerce') or 0))
            _c(ws_tp, row, 7, '')
            row += 1

    for i, w in enumerate([22, 8, 12, 12, 18, 14, 14], 1):
        ws_tp.column_dimensions[get_column_letter(i)].width = w

    # ── Per-SFID latency bin sheets ──────────────────────────────────────────
    if kafka_df is not None:
        bin_cols = [f'lat_bin{str(i).zfill(2)}' for i in range(1, 17)]
        present  = [c for c in bin_cols if c in kafka_df.columns]
        if present:
            sfids = kafka_df['sfid'].dropna().unique() if 'sfid' in kafka_df.columns else []
            for sfid in sorted(sfids, key=lambda x: int(x) if str(x).isdigit() else 0):
                sf_df = kafka_df[kafka_df['sfid'] == sfid]
                if sf_df.empty:
                    continue

                # Bin edges from first/last bin columns
                bin01_lower = pd.to_numeric(sf_df['bin01_lower_msec'].dropna(), errors='coerce').iloc[0] \
                              if 'bin01_lower_msec' in sf_df.columns and not sf_df['bin01_lower_msec'].dropna().empty else 0.0
                bin16_upper = pd.to_numeric(sf_df['bin16_upper_msec'].dropna(), errors='coerce').iloc[0] \
                              if 'bin16_upper_msec' in sf_df.columns and not sf_df['bin16_upper_msec'].dropna().empty else 500.0

                ws_sf = wb.create_sheet(title=f'SFID_{sfid}'[:31])
                ws_sf.merge_cells('A1:F1')
                ws_sf['A1'] = f'LATENCY BINS — SFID {sfid}  ({mac_colon})'
                ws_sf['A1'].font = Font(bold=True, size=13)
                ws_sf['A1'].alignment = CENTER

                # Bin edges header
                _c(ws_sf, 3, 1, 'Bin #',       font=HEADER_FONT, fill=HEADER_FILL)
                _c(ws_sf, 3, 2, 'Bin Col',      font=HEADER_FONT, fill=HEADER_FILL)
                _c(ws_sf, 3, 3, 'Poll Time',    font=HEADER_FONT, fill=HEADER_FILL)
                _c(ws_sf, 3, 4, 'Direction',    font=HEADER_FONT, fill=HEADER_FILL)
                _c(ws_sf, 3, 5, 'Pkt Count',    font=HEADER_FONT, fill=HEADER_FILL)

                row = 4
                for _, r in sf_df.iterrows():
                    for i, bc in enumerate(present, 1):
                        val = pd.to_numeric(r.get(bc, None), errors='coerce')
                        if pd.isna(val):
                            continue
                        _c(ws_sf, row, 1, i)
                        _c(ws_sf, row, 2, bc)
                        _c(ws_sf, row, 3, str(r.get('captured_utc', '')))
                        _c(ws_sf, row, 4, str(r.get('dir', '')))
                        _c(ws_sf, row, 5, int(val), fill=CALC_FILL)
                        row += 1

                # Bin_Edges sub-section
                row += 1
                ws_sf.merge_cells(f'A{row}:F{row}')
                ws_sf.cell(row=row, column=1, value='BIN EDGES (from Kafka labels)').font = BOLD
                row += 1
                _c(ws_sf, row, 1, 'bin01_lower_msec', font=BOLD)
                _c(ws_sf, row, 2, bin01_lower, fill=INFO_FILL, fmt='0.000')
                _c(ws_sf, row, 3, 'bin16_upper_msec', font=BOLD)
                _c(ws_sf, row, 4, bin16_upper, fill=INFO_FILL, fmt='0.000')

                for i, w in enumerate([8, 14, 22, 12, 14], 1):
                    ws_sf.column_dimensions[get_column_letter(i)].width = w

    # ── Save ─────────────────────────────────────────────────────────────────
    mac_clean = mac_colon.replace(':', '')
    ts_file   = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_path  = os.path.join(session_dir, f'TimeSeries_{cmts_type}_{mac_clean}_{ts_file}.xlsx')
    wb.save(out_path)
    print(f'[Excel] TimeSeries report → {os.path.basename(out_path)}')
    return out_path
