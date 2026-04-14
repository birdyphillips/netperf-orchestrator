#!/usr/bin/env python3
import paramiko
import sys
import os
import re
import glob
import logging
from datetime import datetime
from config_loader import config

# Suppress paramiko's verbose logging (including SSH banners)
logging.getLogger("paramiko").setLevel(logging.WARNING)

# Excel export functionality
try:
    import pandas as pd
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---------------------------------------------------------------------------
# Latency bin constants
# ---------------------------------------------------------------------------
BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 100.00, 150.00, 200.00, 500.00,
]
NUM_BINS = len(BIN_EDGES_MS) - 1

def parse_snmp_data(file_path):
    """Parse SNMP data from text file and return structured data"""
    with open(file_path, 'r') as f:
        content = f.read()
    
    # Extract metadata
    timestamp_match = re.search(r'SNMP Collection - (.+)', content)
    target_ip_match = re.search(r'Target IP: (.+)', content)
    
    timestamp = timestamp_match.group(1) if timestamp_match else "Unknown"
    target_ip = target_ip_match.group(1) if target_ip_match else "Unknown"
    
    # Split content into sections
    sections = re.split(r'\n([A-Z][^\n]+)\n=+\n', content)
    
    data = []
    current_section = "Unknown"
    
    for i in range(len(sections)):
        # Section headers are at odd indices after split
        if i > 0 and i % 2 == 1:
            current_section = sections[i].strip()
        elif i % 2 == 0 and i > 0:
            # Parse SNMP entries in this section
            snmp_pattern = r'SNMPv2-SMI::(.+?) = (.+?): (.+)'
            matches = re.findall(snmp_pattern, sections[i])
            
            for oid, data_type, value in matches:
                data.append({
                    'SNMP_Name': current_section,
                    'OID': oid,
                    'Type': data_type,
                    'Value': value,
                    'Timestamp': timestamp,
                    'Target_IP': target_ip
                })
    
    return data

def export_to_excel(test_name, phase, timestamp, output_dir):
    """DEPRECATED: Combined SNMP Excel export. Replaced by latency_calculator.py
    which generates Latency_Bin_Report.xlsx with per-service-flow analysis.
    Kept for backward compatibility — call directly if needed."""
    if not EXCEL_AVAILABLE:
        return None
    
    # Create Excel filename with same naming convention
    excel_filename = os.path.join(output_dir, f"{test_name}_SNMP_Combined_{timestamp}.xlsx")
    
    # Check if Excel file already exists to avoid duplicates
    if os.path.exists(excel_filename):
        print(f"Excel file already exists: {excel_filename}")
        return excel_filename
    
    # Find all SNMP files for this test
    before_file = None
    after_file = None
    
    for file in os.listdir(output_dir):
        if file.startswith(f"{test_name}_SNMP_before_"):
            before_file = os.path.join(output_dir, file)
        elif file.startswith(f"{test_name}_SNMP_after_"):
            after_file = os.path.join(output_dir, file)
    
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Process before data
        if before_file and os.path.exists(before_file):
            before_data = parse_snmp_data(before_file)
            if before_data:
                df_before = pd.DataFrame(before_data)
                df_before.to_excel(writer, sheet_name='Before_Test', index=False)
        
        # Process after data
        if after_file and os.path.exists(after_file):
            after_data = parse_snmp_data(after_file)
            if after_data:
                df_after = pd.DataFrame(after_data)
                df_after.to_excel(writer, sheet_name='After_Test', index=False)
        
        # Create comparison sheet if both exist
        if before_file and after_file and os.path.exists(before_file) and os.path.exists(after_file):
            before_data = parse_snmp_data(before_file)
            after_data = parse_snmp_data(after_file)
            
            if before_data and after_data:
                # Create comparison dataframe
                df_before = pd.DataFrame(before_data)
                df_after = pd.DataFrame(after_data)
                
                # Merge on OID for comparison, include SNMP_Name
                comparison = pd.merge(df_before[['SNMP_Name', 'OID', 'Value']], 
                                    df_after[['OID', 'Value']], 
                                    on='OID', 
                                    suffixes=('_Before', '_After'))
                
                # Calculate differences for numeric values
                def safe_numeric_diff(before_val, after_val):
                    try:
                        before_num = float(re.sub(r'[^\d.]', '', str(before_val)))
                        after_num = float(re.sub(r'[^\d.]', '', str(after_val)))
                        return after_num - before_num
                    except:
                        return "N/A"
                
                comparison['Difference'] = comparison.apply(
                    lambda row: safe_numeric_diff(row['Value_Before'], row['Value_After']), 
                    axis=1
                )
                
                comparison.to_excel(writer, sheet_name='Comparison', index=False)
    
    print(f"Excel file created: {excel_filename}")
    return excel_filename

def ssh_snmp_collector(username, jumpserver, target_ip, output_file=None, snmp_community=None, snmp_timeout=None, snmp_retries=None):
    """SSH into jump server and execute SNMP commands"""
    
    # Use config values if not provided
    if snmp_community is None:
        snmp_community = config.snmp_community
    if snmp_timeout is None:
        snmp_timeout = config.snmp_timeout
    if snmp_retries is None:
        snmp_retries = config.snmp_retries
    
    # Modem information command
    modem_info_cmd = f"snmpwalk -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} sysDescr"
    
    # SNMP commands from your notes
    commands = [
        f"snmpwalk -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} 1.3.6.1.4.1.4491.2.1.21.1.4",
        f"snmpwalk -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} 1.3.6.1.4.1.4491.2.1.21.1.27", 
        f"snmpwalk -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} 1.3.6.1.4.1.4491.2.1.21.1.29.2",
        f"snmpwalk -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} 1.3.6.1.4.1.4491.2.1.21.1.30",
        f"snmpbulkget -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} .1.3.6.1.4.1.4998.1.1.15.10.2",
        f"snmpbulkget -v 2c -c {snmp_community} -t {snmp_timeout} -r {snmp_retries} {target_ip} .1.3.6.1.4.1.4998.1.1.15.10.8"
    ]
    
    labels = [
        "Flow Stats Table (Entry Qos Service Flow Octets)",
        "Aggregate Service Flow Stats Table", 
        "Latency Stats Table",
        "Congestion Stats Table",
        "Cadant Map Stats Mib",
        "Map Stats Pages Flows"
    ]
    
    try:
        # Create SSH client
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        # Try SSH key first
        key_path = config.ssh_key_path
        connected = False
        connection_method = None
        
        try:
            if os.path.exists(key_path):
                print(f"Attempting SSH connection to {jumpserver} with key: {key_path}")
                ssh.connect(jumpserver, username=username, key_filename=key_path, timeout=10)
                connected = True
                connection_method = f"SSH key: {key_path}"
        except Exception as e:
            print(f"Failed to connect with {key_path}: {e}")
        
        if not connected:
            try:
                print(f"Attempting SSH connection to {jumpserver} with default keys")
                ssh.connect(jumpserver, username=username, timeout=10)
                connected = True
                connection_method = "Default SSH keys"
            except Exception as e:
                print(f"Failed to connect with default keys: {e}")
        
        if not connected:
            raise Exception(f"Failed to connect to {jumpserver} as {username}. Tried: {key_path} and default keys. No authentication methods available.")
        
        print(f"Successfully connected to {jumpserver} using {connection_method}")
        
        results = {}
        output_lines = []
        
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        output_lines.append(f"SNMP Collection - {timestamp}")
        output_lines.append(f"Target IP: {target_ip}")
        output_lines.append("="*60)
        
        # Get modem information first
        print(f"Executing: snmpwalk -v 2c -c open {target_ip} sysDescr")
        stdin, stdout, stderr = ssh.exec_command(modem_info_cmd)
        modem_output = stdout.read().decode()
        modem_error = stderr.read().decode()
        
        output_lines.append(f"\nCurrent Modem Information")
        output_lines.append("="*50)
        if modem_output:
            output_lines.append(modem_output)
            print(modem_output)  # Display in terminal
        if modem_error:
            output_lines.append(f"ERROR: {modem_error}")
            print(f"ERROR: {modem_error}")
        
        # Execute each SNMP command
        for i, cmd in enumerate(commands):
            print(f"Executing: {labels[i]}")
            stdin, stdout, stderr = ssh.exec_command(cmd)
            
            output = stdout.read().decode()
            error = stderr.read().decode()
            
            results[labels[i]] = {
                'output': output,
                'error': error if error else None
            }
            
            # Add to output lines
            output_lines.append(f"\n{labels[i]}")
            output_lines.append("="*50)
            if output:
                output_lines.append(output)
            if error:
                output_lines.append(f"ERROR: {error}")
                print(f"Error in {labels[i]}: {error}")
        
        ssh.close()
        
        # Write to file if specified
        if output_file:
            with open(output_file, 'w') as f:
                f.write('\n'.join(output_lines))
            print(f"Results saved to: {output_file}")
        
        return results
        
    except Exception as e:
        error_msg = f"SSH connection failed: {e}"
        print(error_msg)
        if output_file:
            with open(output_file, 'w') as f:
                f.write(f"ERROR: {error_msg}\n")
        return None

def collect_snmp_data(target_ip, test_name, phase, output_dir):
    """Collect SNMP data for a specific test and phase"""
    username = config.snmp_username
    jumpserver = config.snmp_jumpserver
    
    # Use provided output directory
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Create filename in output folder with SCN naming convention
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(output_dir, f"{test_name}_SNMP_{phase}_{timestamp}.txt")
    
    print(f"Collecting SNMP data - {phase} {test_name}")
    result = ssh_snmp_collector(username, jumpserver, target_ip, filename)
    
    # DEPRECATED: Combined SNMP Excel export — replaced by latency_calculator.py
    # Latency bin reports are now generated via generate_latency_report() in the
    # orchestrator after all iterations complete.
    # To re-enable, uncomment the block below:
    # if phase == "after":
    #     try:
    #         export_to_excel(test_name, phase, timestamp, output_dir)
    #     except Exception as e:
    #         print(f"Excel export failed: {e}")
    
    return result


# ---------------------------------------------------------------------------
# Latency bin parsing & calculation (moved from latency_calculator.py)
# ---------------------------------------------------------------------------

def parse_snmp_timestamp(filepath):
    """Extract collection timestamp from SNMP file header."""
    with open(filepath, "r") as f:
        for line in f:
            m = re.match(r"SNMP Collection - (\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})", line)
            if m:
                return datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S")
    return None


def parse_flow_stats(filepath):
    """Parse Flow Stats Table for packets, octets, and dropped packets per SFID."""
    with open(filepath, "r") as f:
        content = f.read()
    match = re.search(r"Flow Stats Table.*?\n=+\n(.*?)(?:\n\n|\Z)", content, re.DOTALL)
    if not match:
        return {}
    section = match.group(1)
    stats = {}
    sub_map = {"1": "packets", "2": "octets", "8": "dropped"}
    for line in section.splitlines():
        m = re.search(r"\.4\.1\.(\d+)\.2\.(\d+)\s*=\s*(?:Counter64|Counter32):\s*(\d+)", line)
        if m:
            sub_oid, sfid_str, val_str = m.group(1), m.group(2), m.group(3)
            if sub_oid in sub_map:
                sfid = int(sfid_str)
                stats.setdefault(sfid, {"packets": 0, "octets": 0, "dropped": 0})
                stats[sfid][sub_map[sub_oid]] = int(val_str)
    return stats


DEFAULT_DURATION = {"byteblower": 60, "iperf3": 30}


def _infer_duration(filepath):
    base = os.path.basename(filepath).lower()
    for key, dur in DEFAULT_DURATION.items():
        if key in base:
            return dur
    return None


def compute_throughput_and_loss(before_file, after_file, duration_s=None):
    """Compute per-SFID throughput (Mbps) and packet loss from flow stats deltas."""
    fs_before = parse_flow_stats(before_file)
    fs_after = parse_flow_stats(after_file)
    if not fs_before or not fs_after:
        return {}
    if duration_s is None:
        duration_s = _infer_duration(before_file)
    if duration_s is None:
        ts_before = parse_snmp_timestamp(before_file)
        ts_after = parse_snmp_timestamp(after_file)
        if ts_before and ts_after:
            duration_s = (ts_after - ts_before).total_seconds()
    if not duration_s or duration_s <= 0:
        return {}
    results = {}
    for sfid in sorted(set(fs_before) & set(fs_after)):
        d_octets = max(fs_after[sfid]["octets"] - fs_before[sfid]["octets"], 0)
        d_packets = max(fs_after[sfid]["packets"] - fs_before[sfid]["packets"], 0)
        d_dropped = max(fs_after[sfid]["dropped"] - fs_before[sfid]["dropped"], 0)
        if d_packets == 0 and d_octets == 0:
            continue
        total_pkts = d_packets + d_dropped
        results[sfid] = {
            "throughput_mbps": (d_octets * 8) / (duration_s * 1_000_000),
            "lost_packets": d_dropped,
            "total_packets": total_pkts,
            "loss_pct": (d_dropped / total_pkts * 100) if total_pkts > 0 else 0.0,
        }
    return results


def parse_latency_bins(filepath):
    """Parse Latency Stats Table from an SNMP text file."""
    with open(filepath, "r") as f:
        content = f.read()
    match = re.search(r"Latency Stats Table\n=+\n(.*?)(?:\n\n|\Z)", content, re.DOTALL)
    if not match:
        return {}
    section = match.group(1)
    bins = {}
    for line in section.splitlines():
        m = re.search(r"\.29\.2\.1\.(\d+)\.2\.(\d+)\s*=\s*(?:Counter64|Gauge32):\s*(\d+)", line)
        if m:
            sub_oid = int(m.group(1))
            sfid = int(m.group(2))
            value = int(m.group(3))
            bins.setdefault(sfid, {})[sub_oid] = value
    return bins


def compute_deltas(before_bins, after_bins):
    """Compute per-service-flow bin deltas (after - before)."""
    results = {}
    all_sfids = set(before_bins) & set(after_bins)
    for sfid in sorted(all_sfids):
        before_vals, after_vals, deltas = [], [], []
        for sub in range(2, 2 + NUM_BINS):
            bv = before_bins[sfid].get(sub, 0)
            av = after_bins[sfid].get(sub, 0)
            before_vals.append(bv)
            after_vals.append(av)
            deltas.append(max(av - bv, 0))
        if sum(deltas) > 0:
            results[sfid] = {"before": before_vals, "after": after_vals, "deltas": deltas}
    return results


# ---------------------------------------------------------------------------
# Percentile & average calculations
# ---------------------------------------------------------------------------

def calc_percentile(deltas, percentile):
    """Linear interpolation percentile from bin deltas."""
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            prev_cum = cumulative - count
            bin_low = BIN_EDGES_MS[i]
            bin_high = BIN_EDGES_MS[i + 1]
            denom = count if count > 0 else 1
            return bin_low + ((target - prev_cum) / denom) * (bin_high - bin_low)
    return BIN_EDGES_MS[-1]


def calc_weighted_avg(deltas):
    """Weighted average latency: sum(delta_i x bin_avg_i) / total."""
    total = sum(deltas)
    if total == 0:
        return 0.0
    weighted = sum(
        deltas[i] * (BIN_EDGES_MS[i] + BIN_EDGES_MS[i + 1]) / 2
        for i in range(NUM_BINS)
    )
    return weighted / total


def calc_percentile_avg(deltas, percentile):
    """AVG method: return the bin midpoint of the first bin where
    cumulative count >= percentile target."""
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            return (BIN_EDGES_MS[i] + BIN_EDGES_MS[i + 1]) / 2
    return (BIN_EDGES_MS[-2] + BIN_EDGES_MS[-1]) / 2


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------

if OPENPYXL_AVAILABLE:
    _HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    _HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    _CALC_FILL = PatternFill("solid", fgColor="D9E2F3")
    _RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
    _INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
    _THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _CENTER = Alignment(horizontal="center", vertical="center")
    _BOLD = Font(bold=True, size=11)


def _styled_cell(ws, row, col, value, font=None, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _CENTER
    cell.border = _THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


def write_sf_sheet(wb, sheet_name, sf_data, tp_data=None):
    """Write one worksheet for a service flow's latency bin data."""
    deltas = sf_data["deltas"]
    before = sf_data["before"]
    after = sf_data["after"]
    ws = wb.create_sheet(title=sheet_name)

    ws.merge_cells("A1:J1")
    ws["A1"] = f"CMTS DS LATENCY \u2014 {sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    if tp_data:
        _styled_cell(ws, 2, 1, "Throughput (Mbps):", font=_BOLD)
        _styled_cell(ws, 2, 2, round(tp_data["throughput_mbps"], 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, 2, 4, "Packet Loss:", font=_BOLD)
        _styled_cell(ws, 2, 5, tp_data["lost_packets"], fill=_RESULT_FILL)
        _styled_cell(ws, 2, 6, f"{tp_data['loss_pct']:.4f}%", fill=_RESULT_FILL)

    headers = [
        "BIN", "LOWER (ms)", "UPPER (ms)", "AVG (ms)",
        "START COUNT", "END COUNT", "DELTA",
        "CUMULATIVE", "CUMULATIVE %", "BIN %",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    total = sum(deltas)
    cumulative = 0

    for i in range(NUM_BINS):
        row = 4 + i
        low = BIN_EDGES_MS[i]
        high = BIN_EDGES_MS[i + 1]
        avg = (low + high) / 2
        delta = deltas[i]
        cumulative += delta
        cum_pct = (cumulative / total * 100) if total else 0
        bin_pct = (delta / total * 100) if total else 0

        _styled_cell(ws, row, 1, i + 1)
        _styled_cell(ws, row, 2, low, fmt="0.00")
        _styled_cell(ws, row, 3, high if i < 15 else "200.00+", fmt="0.00")
        _styled_cell(ws, row, 4, avg, fill=_CALC_FILL, fmt="0.0000")
        _styled_cell(ws, row, 5, before[i], fill=_INPUT_FILL)
        _styled_cell(ws, row, 6, after[i], fill=_INPUT_FILL)
        _styled_cell(ws, row, 7, delta, fill=_CALC_FILL)
        _styled_cell(ws, row, 8, cumulative, fill=_CALC_FILL)
        _styled_cell(ws, row, 9, cum_pct, fill=_CALC_FILL, fmt="0.00")
        _styled_cell(ws, row, 10, bin_pct, fill=_CALC_FILL, fmt="0.00")

    total_row = 4 + NUM_BINS + 1
    _styled_cell(ws, total_row, 1, "TOTAL", font=_BOLD)
    _styled_cell(ws, total_row, 5, sum(before), font=_BOLD, fill=_INPUT_FILL)
    _styled_cell(ws, total_row, 6, sum(after), font=_BOLD, fill=_INPUT_FILL)
    _styled_cell(ws, total_row, 7, total, font=_BOLD, fill=_CALC_FILL)
    _styled_cell(ws, total_row, 9, 100.00 if total else 0, font=_BOLD, fill=_CALC_FILL, fmt="0.00")

    pct_row = total_row + 2
    ws.merge_cells(f"A{pct_row}:J{pct_row}")
    ws.cell(row=pct_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

    for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
        pct_row += 1
        _styled_cell(ws, pct_row, 1, f"{label} TARGET", font=_BOLD)
        _styled_cell(ws, pct_row, 2, round(total * pct, 2), fill=_RESULT_FILL, fmt="0.00")
        _styled_cell(ws, pct_row, 3, f"{label} (ms)", font=_BOLD)
        _styled_cell(ws, pct_row, 4, round(calc_percentile(deltas, pct), 4), fill=_RESULT_FILL, fmt="0.0000")

    legend_row = pct_row + 2
    ws.cell(row=legend_row, column=1, value="FORMULA:").font = _BOLD
    ws.merge_cells(f"B{legend_row}:J{legend_row}")
    ws.cell(row=legend_row, column=2, value="P = BIN_LOW + ((TARGET \u2212 PREV_CUMULATIVE) / BIN_COUNT) \u00d7 (BIN_HIGH \u2212 BIN_LOW)").font = Font(italic=True, size=10)

    avg_row = legend_row + 2
    ws.merge_cells(f"A{avg_row}:J{avg_row}")
    ws.cell(row=avg_row, column=1, value="PERCENTILE RESULTS (AVG METHOD)").font = Font(bold=True, size=12)

    for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
        avg_row += 1
        _styled_cell(ws, avg_row, 1, f"{label} TARGET", font=_BOLD)
        _styled_cell(ws, avg_row, 2, round(total * pct, 2), fill=_RESULT_FILL, fmt="0.00")
        _styled_cell(ws, avg_row, 3, f"{label} AVG (ms)", font=_BOLD)
        _styled_cell(ws, avg_row, 4, round(calc_percentile_avg(deltas, pct), 4), fill=_RESULT_FILL, fmt="0.0000")

    legend_row2 = avg_row + 2
    ws.cell(row=legend_row2, column=1, value="FORMULA:").font = _BOLD
    ws.merge_cells(f"B{legend_row2}:J{legend_row2}")
    ws.cell(row=legend_row2, column=2, value="P = AVG (col D) of first bin where cumulative count >= percentile target").font = Font(italic=True, size=10)

    for i, w in enumerate([8, 14, 14, 14, 16, 16, 14, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def parse_congestion_stats(filepath):
    """Parse Congestion Stats Table for AQM drops, congestion marked, sanctioned per SFID.
    OID .30.1.1 = AQM dropped, .30.1.3 = congestion marked, .30.1.4 = sanctioned."""
    with open(filepath, "r") as f:
        content = f.read()
    match = re.search(r"Congestion Stats Table\n=+\n(.*?)(?:\n\n|\Z)", content, re.DOTALL)
    if not match:
        return {}
    section = match.group(1)
    stats = {}
    sub_map = {"1": "aqm_drops", "3": "congestion_marked", "4": "sanctioned"}
    for line in section.splitlines():
        m = re.search(r"\.30\.1\.(\d+)\.2\.(\d+)\s*=\s*Counter64:\s*(\d+)", line)
        if m:
            sub_oid, sfid_str, val_str = m.group(1), m.group(2), m.group(3)
            if sub_oid in sub_map:
                sfid = int(sfid_str)
                stats.setdefault(sfid, {"aqm_drops": 0, "congestion_marked": 0, "sanctioned": 0})
                stats[sfid][sub_map[sub_oid]] = int(val_str)
    return stats


def write_timeseries_sheet(wb, before_file, after_file, fs_before, fs_after, cong_before, cong_after, before_bins, after_bins):
    """Write raw SNMP before/after values into a TimeSeries tab."""
    ws = wb.create_sheet(title="TimeSeries")
    ws.sheet_properties.tabColor = "FF6600"

    headers = ["Phase", "Timestamp", "SFID", "Metric", "Value"]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 1, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    ts_before = parse_snmp_timestamp(before_file) or "unknown"
    ts_after = parse_snmp_timestamp(after_file) or "unknown"
    ts_b_str = ts_before.strftime("%Y-%m-%d %H:%M:%S") if hasattr(ts_before, 'strftime') else str(ts_before)
    ts_a_str = ts_after.strftime("%Y-%m-%d %H:%M:%S") if hasattr(ts_after, 'strftime') else str(ts_after)

    row = 2
    for phase, ts_str, fs, cong, bins in [
        ("BEFORE", ts_b_str, fs_before, cong_before, before_bins),
        ("AFTER", ts_a_str, fs_after, cong_after, after_bins),
    ]:
        for sfid in sorted(fs.keys()):
            for metric, val in [("packets", fs[sfid]["packets"]), ("octets", fs[sfid]["octets"]), ("dropped", fs[sfid]["dropped"])]:
                _styled_cell(ws, row, 1, phase)
                _styled_cell(ws, row, 2, ts_str)
                _styled_cell(ws, row, 3, sfid)
                _styled_cell(ws, row, 4, metric)
                _styled_cell(ws, row, 5, val)
                row += 1
        for sfid in sorted(cong.keys()):
            for metric, val in [("aqm_drops", cong[sfid]["aqm_drops"]), ("congestion_marked", cong[sfid]["congestion_marked"]), ("sanctioned", cong[sfid]["sanctioned"])]:
                _styled_cell(ws, row, 1, phase)
                _styled_cell(ws, row, 2, ts_str)
                _styled_cell(ws, row, 3, sfid)
                _styled_cell(ws, row, 4, metric)
                _styled_cell(ws, row, 5, val)
                row += 1
        for sfid in sorted(bins.keys()):
            for sub_oid in sorted(bins[sfid].keys()):
                _styled_cell(ws, row, 1, phase)
                _styled_cell(ws, row, 2, ts_str)
                _styled_cell(ws, row, 3, sfid)
                _styled_cell(ws, row, 4, f"latency_bin_{sub_oid}")
                _styled_cell(ws, row, 5, bins[sfid][sub_oid])
                row += 1

    for i, w in enumerate([10, 22, 10, 22, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_throughput_sheet(wb, fs_before, fs_after, before_file, after_file):
    """Write Throughput tab showing octet/packet deltas per SFID."""
    ws = wb.create_sheet(title="Throughput")
    ws.sheet_properties.tabColor = "00B050"

    ws.merge_cells("A1:G1")
    ws["A1"] = "QOS SERVICE FLOW OCTETS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    ts_before = parse_snmp_timestamp(before_file)
    ts_after = parse_snmp_timestamp(after_file)
    ts_b_str = ts_before.strftime("%Y-%m-%d %H:%M:%S") if ts_before else "unknown"
    ts_a_str = ts_after.strftime("%Y-%m-%d %H:%M:%S") if ts_after else "unknown"
    duration_s = (ts_after - ts_before).total_seconds() if ts_before and ts_after else 0

    headers = ["SFID", "Poll Before", "Poll After", "Before Octets", "After Octets",
               "Delta Octets", "Rate (Mbps)"]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    row = 4
    grand_total_delta = 0
    all_sfids = sorted(set(fs_before) & set(fs_after))
    for sfid in all_sfids:
        d_octets = max(fs_after[sfid]["octets"] - fs_before[sfid]["octets"], 0)
        rate = (d_octets * 8) / (duration_s * 1_000_000) if duration_s > 0 else 0
        fill = _RESULT_FILL if d_octets > 0 else None
        _styled_cell(ws, row, 1, f"SFID {sfid}")
        _styled_cell(ws, row, 2, ts_b_str, fill=fill)
        _styled_cell(ws, row, 3, ts_a_str, fill=fill)
        _styled_cell(ws, row, 4, fs_before[sfid]["octets"], fill=fill)
        _styled_cell(ws, row, 5, fs_after[sfid]["octets"], fill=fill)
        _styled_cell(ws, row, 6, d_octets, fill=fill)
        _styled_cell(ws, row, 7, round(rate, 4), fill=fill, fmt="0.0000")
        grand_total_delta += d_octets
        row += 1

    grand_rate = (grand_total_delta * 8) / (duration_s * 1_000_000) if duration_s > 0 else 0
    _styled_cell(ws, row, 1, "TOTAL", font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 6, grand_total_delta, font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 7, round(grand_rate, 4), font=_BOLD, fill=_RESULT_FILL, fmt="0.0000")

    for i, w in enumerate([12, 22, 22, 18, 18, 18, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, all_results, tp_stats=None, fs_before=None, fs_after=None,
                        cong_before=None, cong_after=None, before_file=None, after_file=None):
    """Write a summary sheet matching cmts_collector format."""
    ws = wb.create_sheet(title="Summary")
    ws.sheet_properties.tabColor = "4472C4"

    ws.merge_cells("A1:R1")
    ws["A1"] = "CMTS UPSTREAM LATENCY SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    headers = [
        "Service Flow", "Total Bins Pkts",
        "Weighted Avg (ms)", "P50 (ms)", "P99 (ms)", "P99.9 (ms)",
        "P50 AVG (ms)", "P99 AVG (ms)", "P99.9 AVG (ms)",
        "AQM Drops", "Congestion Marked", "Sanctioned Pkts",
        "Throughput (Mbps)", "Pkt Loss %",
        "Total Pkt Delta", "Total Octet Delta",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    if not tp_stats:
        tp_stats = {}
    if not fs_before:
        fs_before = {}
    if not fs_after:
        fs_after = {}
    if not cong_before:
        cong_before = {}
    if not cong_after:
        cong_after = {}

    ts_before = parse_snmp_timestamp(before_file) if before_file else None
    ts_after = parse_snmp_timestamp(after_file) if after_file else None
    duration_s = (ts_after - ts_before).total_seconds() if ts_before and ts_after else 0

    row = 4
    sum_throughput = 0
    sum_pkt_delta = 0
    sum_octet_delta = 0

    for sfid, sf_data in sorted(all_results.items()):
        deltas = sf_data["deltas"]
        total = sum(deltas)
        w_avg = calc_weighted_avg(deltas)
        p50 = calc_percentile(deltas, 0.50)
        p99 = calc_percentile(deltas, 0.99)
        p999 = calc_percentile(deltas, 0.999)
        p50a = calc_percentile_avg(deltas, 0.50)
        p99a = calc_percentile_avg(deltas, 0.99)
        p999a = calc_percentile_avg(deltas, 0.999)

        cb = cong_before.get(sfid, {"aqm_drops": 0, "congestion_marked": 0, "sanctioned": 0})
        ca = cong_after.get(sfid, {"aqm_drops": 0, "congestion_marked": 0, "sanctioned": 0})
        aqm = max(ca["aqm_drops"] - cb["aqm_drops"], 0)
        cong = max(ca["congestion_marked"] - cb["congestion_marked"], 0)
        sanc = max(ca["sanctioned"] - cb["sanctioned"], 0)

        d_octets = 0
        d_pkts = 0
        throughput = 0
        loss_pct = 0
        if sfid in fs_before and sfid in fs_after:
            d_octets = max(fs_after[sfid]["octets"] - fs_before[sfid]["octets"], 0)
            d_pkts = max(fs_after[sfid]["packets"] - fs_before[sfid]["packets"], 0)
            d_dropped = max(fs_after[sfid]["dropped"] - fs_before[sfid]["dropped"], 0)
            throughput = (d_octets * 8) / (duration_s * 1_000_000) if duration_s > 0 else 0
            total_pkts = d_pkts + d_dropped
            loss_pct = (d_dropped / total_pkts * 100) if total_pkts > 0 else 0

        _styled_cell(ws, row, 1, f"SFID {sfid}")
        _styled_cell(ws, row, 2, total, fill=_CALC_FILL)
        _styled_cell(ws, row, 3, round(w_avg, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 4, round(p50, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 5, round(p99, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 6, round(p999, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 7, round(p50a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 8, round(p99a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 9, round(p999a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 10, aqm, fill=_CALC_FILL)
        _styled_cell(ws, row, 11, cong, fill=_CALC_FILL)
        _styled_cell(ws, row, 12, sanc, fill=_CALC_FILL)
        _styled_cell(ws, row, 13, round(throughput, 4), fill=_CALC_FILL, fmt="0.0000")
        _styled_cell(ws, row, 14, round(loss_pct, 4), fill=_CALC_FILL, fmt="0.0000")
        _styled_cell(ws, row, 15, d_pkts, fill=_CALC_FILL)
        _styled_cell(ws, row, 16, d_octets, fill=_CALC_FILL)
        sum_throughput += throughput
        sum_pkt_delta += d_pkts
        sum_octet_delta += d_octets
        row += 1

    _styled_cell(ws, row, 1, "TOTAL", font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 13, round(sum_throughput, 4), font=_BOLD, fill=_RESULT_FILL, fmt="0.0000")
    _styled_cell(ws, row, 15, sum_pkt_delta, font=_BOLD, fill=_RESULT_FILL)
    _styled_cell(ws, row, 16, sum_octet_delta, font=_BOLD, fill=_RESULT_FILL)

    for i, w in enumerate([16, 16, 18, 14, 14, 14, 16, 16, 16, 14, 18, 16, 18, 14, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_latency_report(before_file, after_file, output_file=None):
    """Main entry: parse SNMP files, compute deltas, write Excel report."""
    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not available \u2014 skipping latency report")
        return None

    before_bins = parse_latency_bins(before_file)
    after_bins = parse_latency_bins(after_file)

    if not before_bins or not after_bins:
        print("WARNING: No latency stats found in one or both SNMP files.")
        return None

    all_deltas = compute_deltas(before_bins, after_bins)

    if not all_deltas:
        print("WARNING: All latency bin deltas are zero \u2014 no traffic detected.")
        return None

    tp_stats = compute_throughput_and_loss(before_file, after_file)
    fs_before = parse_flow_stats(before_file)
    fs_after = parse_flow_stats(after_file)
    cong_before = parse_congestion_stats(before_file)
    cong_after = parse_congestion_stats(after_file)

    if output_file is None:
        output_dir = os.path.dirname(after_file) or "."
        dir_name = os.path.basename(os.path.abspath(output_dir))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"Latency_Bin_Report_{dir_name}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    write_timeseries_sheet(wb, before_file, after_file, fs_before, fs_after,
                           cong_before, cong_after, before_bins, after_bins)
    write_summary_sheet(wb, all_deltas, tp_stats, fs_before, fs_after,
                        cong_before, cong_after, before_file, after_file)
    write_throughput_sheet(wb, fs_before, fs_after, before_file, after_file)

    for sfid, sf_data in sorted(all_deltas.items()):
        write_sf_sheet(wb, f"SFID_{sfid}", sf_data, tp_stats.get(sfid))

    wb.save(output_file)
    print(f"Latency report saved: {output_file}")

    print("\n--- Latency Summary ---")
    for sfid, sf_data in sorted(all_deltas.items()):
        deltas = sf_data["deltas"]
        total = sum(deltas)
        avg = calc_weighted_avg(deltas)
        p50 = calc_percentile(deltas, 0.50)
        p99 = calc_percentile(deltas, 0.99)
        p999 = calc_percentile(deltas, 0.999)
        tp = tp_stats.get(sfid)
        tp_str = f"  Throughput={tp['throughput_mbps']:.4f}Mbps  Loss={tp['loss_pct']:.4f}%" if tp else ""
        print(f"  SFID {sfid}: {total} pkts | AVG={avg:.4f}ms  P50={p50:.4f}ms  P99={p99:.4f}ms  P99.9={p999:.4f}ms{tp_str}")

    return output_file


def find_snmp_files(results_dir):
    """Auto-discover before/after SNMP .txt files in a results directory."""
    before_files = sorted(glob.glob(os.path.join(results_dir, "**", "*SNMP_before_*.txt"), recursive=True))
    after_files = sorted(glob.glob(os.path.join(results_dir, "**", "*SNMP_after_*.txt"), recursive=True))
    if not before_files or not after_files:
        return None, None
    return before_files[0], after_files[0]


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python snmp_collector.py <target_ip> [test_name] [phase]")
        print("Example: python snmp_collector.py <target_ip> ByteBlower_Test before")
        sys.exit(1)
    
    target_ip = sys.argv[1]
    test_name = sys.argv[2] if len(sys.argv) > 2 else "manual_test"
    phase = sys.argv[3] if len(sys.argv) > 3 else "standalone"
    
    if len(sys.argv) > 2:
        collect_snmp_data(target_ip, test_name, phase, "Results")
    else:
        # Original behavior for backward compatibility
        username = config.snmp_username
        jumpserver = config.snmp_jumpserver
        
        results = ssh_snmp_collector(username, jumpserver, target_ip)
        
        if results:
            for label, data in results.items():
                print(f"\n{'='*50}")
                print(f"{label}")
                print(f"{'='*50}")
                if data['output']:
                    print(data['output'])
                if data['error']:
                    print(f"ERROR: {data['error']}")