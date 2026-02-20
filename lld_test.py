#!/home/aphillips/Projects/LLD_TEST_CLT_v1.4_20260115_160000_linux_compatible/venv/bin/python3

import argparse
import sys
import subprocess
import os
import glob
from datetime import datetime

import csv

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from packetstorm_logic import PacketStormLogic
from byteblower_logic import ByteBlowerLogic
from iperf3_logic import IPerf3Logic
from speedtest_logic import SpeedTestLogic
from logger import Logger
from snmp_collector import collect_snmp_data

class ByteBlowerCLI:
    def __init__(self):
        self.logger = Logger("ByteBlowerCLI")
        # Prompt for modem IPv6 address
        user_input = input("Enter modem IPv6 address: ")
        while not user_input.strip():
            user_input = input("Modem IPv6 address is required. Please enter: ")
        self.target_ip = user_input.strip()
        self.logger.info(f"Using modem IPv6: {self.target_ip}")
        self.output_dir = None
    
    def run_snmp_collection(self, test_name, phase, output_dir=None, rtt_suffix=""):
        try:
            if output_dir is None:
                output_dir = "Results"
            snmp_name = f"{test_name}{rtt_suffix}"
            self.logger.info(f"Running SNMP collection - {phase} {snmp_name}")
            collect_snmp_data(self.target_ip, snmp_name, phase, output_dir)
            return True
        except Exception as e:
            self.logger.error(f"SNMP collection failed: {e}")
            return False
    
    def execute(self, bbp_file, rtt_files, iterations, scenarios, test_group_name=None, client_ip=None, output_format="json", byteblower_only=False, packetstorm_only=False, iperf3_only=False, iperf3_darwin=False, speedtest_only=False, speedtest_clients=None, report_formats="html pdf csv xls xlsx json docx"):
        """Execute workflow based on selected modes"""
        try:
            scenario_list = [s.strip() for s in scenarios.split(',')] if scenarios else ['default']
            rtt_list = [r.strip() for r in rtt_files.split(',')] if rtt_files else ['default.json']
            
            # Determine traffic type
            if byteblower_only or (not iperf3_only and not iperf3_darwin and not speedtest_only):
                traffic_type = "ByteBlower"
            elif iperf3_darwin:
                traffic_type = "iPerf3_macOS"
            elif iperf3_only:
                traffic_type = "iPerf3_Linux"
            elif speedtest_only:
                traffic_type = "SpeedTest"
            else:
                traffic_type = "PacketStorm"
            
            # Extract RTT suffix from first RTT file
            rtt_suffix = ""
            if rtt_list[0] and rtt_list[0] != 'default.json':
                import re
                rtt_match = re.search(r'(\d+)ms', rtt_list[0])
                if rtt_match:
                    rtt_suffix = f"_RTT_{rtt_match.group(1)}ms"
            
            # Create parent output directory with test group name, traffic type, RTT and timestamp
            parent_output_dir = f"Results/{test_group_name}_{traffic_type}{rtt_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            os.makedirs(parent_output_dir, exist_ok=True)
            
            # Store traffic type and rtt_suffix for subdirectory naming
            self.traffic_type = traffic_type
            self.rtt_suffix = rtt_suffix
            
            if len(scenario_list) > 1 or len(rtt_list) > 1:
                self.logger.info(f"\n{'='*60}")
                self.logger.info(f"Multi-scenario/RTT run: {len(scenario_list)} scenarios x {len(rtt_list)} RTT configs")
                self.logger.info(f"Output: {parent_output_dir}")
                self.logger.info(f"{'='*60}\n")
            
            all_success = True
            all_snmp_files = []
            
            # Run combinations
            for scenario in scenario_list:
                for rtt_file in rtt_list:
                    success, snmp_files = self._run_single_test(bbp_file, rtt_file, iterations, scenario, test_group_name, client_ip, output_format, byteblower_only, packetstorm_only, iperf3_only, iperf3_darwin, speedtest_only, speedtest_clients, report_formats, parent_output_dir, rtt_list)
                    all_success = all_success and success
                    all_snmp_files.extend(snmp_files)
            
            if all_success:
                self.logger.info("All workflows completed successfully")
                return 0
            else:
                self.logger.error("Some workflows completed with errors")
                return 1
                
        except Exception as e:
            self.logger.error(f"Workflow failed: {e}")
            return 1
    
    def _run_single_test(self, bbp_file, rtt_file, iterations, scenario_name, test_group_name, client_ip, output_format, byteblower_only, packetstorm_only, iperf3_only, iperf3_darwin, speedtest_only, speedtest_clients, report_formats, parent_output_dir, rtt_list):
        """Run a single test scenario"""
        try:
            success = True
            snmp_files = []
            
            # Extract RTT value from filename for naming
            rtt_suffix = ""
            rtt_dir_suffix = ""
            if rtt_file and rtt_file != 'default.json':
                import re
                rtt_match = re.search(r'(\d+)ms', rtt_file)
                if rtt_match:
                    rtt_suffix = f"_RTT_{rtt_match.group(1)}ms"
                    rtt_dir_suffix = f"RTT_{rtt_match.group(1)}ms"
            
            # Create subdirectory only if multiple RTT values
            test_output_dir = parent_output_dir
            if len(rtt_list) > 1 and rtt_dir_suffix:
                traffic_subdir = f"{self.traffic_type}_{rtt_dir_suffix}"
                test_output_dir = os.path.join(parent_output_dir, traffic_subdir)
                os.makedirs(test_output_dir, exist_ok=True)
            
            # Determine test name for SNMP collection
            test_name = scenario_name
            if byteblower_only:
                test_name = f"ByteBlower_{scenario_name}"
            elif iperf3_only:
                test_name = f"iPerf3_Linux_{scenario_name}"
            elif iperf3_darwin:
                test_name = f"iPerf3_macOS_{scenario_name}"
            elif packetstorm_only:
                test_name = f"PacketStorm_{scenario_name}"
            else:
                if client_ip:
                    test_name = f"iPerf3_Linux_{scenario_name}"
                else:
                    test_name = f"ByteBlower_{scenario_name}"
            
            if speedtest_only:
                self.logger.info(f"SpeedTest mode - clients: {speedtest_clients}")
                st = SpeedTestLogic(speedtest_clients, test_group_name)
                self.output_dir = test_output_dir
                success = st.run_iterations(iterations)
            elif packetstorm_only:
                self.logger.info(f"PacketStorm only mode - config: {rtt_file}")
                ps = PacketStormLogic(rtt_file)
                success = ps.start_config() and ps.stop_config()
            elif byteblower_only:
                self.logger.info(f"ByteBlower only mode - file: {bbp_file}, scenario: {scenario_name}")
                bb = ByteBlowerLogic(bbp_file, scenario_name, scenario_name, test_group_name, rtt_suffix, report_formats)
                self.output_dir = test_output_dir
                success = True
                snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
                os.makedirs(snmp_dir, exist_ok=True)
                
                for i in range(iterations):
                    self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "before", snmp_dir, "")
                    if not bb.run_scenario(i, iterations, test_output_dir):
                        success = False
                    self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "after", snmp_dir, "")
            elif iperf3_only or iperf3_darwin:
                platform_override = 'macos' if iperf3_darwin else None
                platform_suffix = "_macOS" if iperf3_darwin else "_Linux"
                self.logger.info(f"iPerf3 only mode - client: {client_ip}, scenario: {scenario_name}")
                iperf3 = IPerf3Logic(client_ip, scenario_name, test_group_name, rtt_suffix, output_format, platform_override, test_output_dir)
                self.output_dir = test_output_dir
                
                # Setup SSH and servers once
                if not iperf3.setup_ssh_keys():
                    self.logger.error("SSH key setup failed")
                    return False, []
                if not iperf3.setup_iperf3_servers():
                    self.logger.error("iPerf3 server setup failed")
                    return False, []
                
                success = True
                snmp_dir = os.path.join(test_output_dir, scenario_name + platform_suffix + rtt_suffix)
                os.makedirs(snmp_dir, exist_ok=True)
                
                for i in range(iterations):
                    self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "before", snmp_dir, "")
                    if not iperf3.run_scenario(i, iterations):
                        success = False
                    self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "after", snmp_dir, "")
                
                iperf3.stop_iperf3_servers()
            else:
                ps = PacketStormLogic(rtt_file)
                
                if client_ip:
                    iperf3 = IPerf3Logic(client_ip, scenario_name, test_group_name, rtt_suffix, output_format, None, test_output_dir)
                    self.output_dir = test_output_dir
                    
                    success = ps.start_config()
                    if success:
                        # Setup SSH and servers once
                        if not iperf3.setup_ssh_keys():
                            self.logger.error("SSH key setup failed")
                            ps.stop_config()
                            return False, []
                        if not iperf3.setup_iperf3_servers():
                            self.logger.error("iPerf3 server setup failed")
                            ps.stop_config()
                            return False, []
                        
                        snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
                        os.makedirs(snmp_dir, exist_ok=True)
                        
                        for i in range(iterations):
                            self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "before", snmp_dir, "")
                            if not iperf3.run_scenario(i, iterations):
                                success = False
                            self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "after", snmp_dir, "")
                        
                        iperf3.stop_iperf3_servers()
                        success = success and ps.stop_config()
                else:
                    bb = ByteBlowerLogic(bbp_file, scenario_name, scenario_name, test_group_name, rtt_suffix, report_formats)
                    self.output_dir = test_output_dir
                    
                    success = ps.start_config()
                    if success:
                        snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
                        os.makedirs(snmp_dir, exist_ok=True)
                        
                        for i in range(iterations):
                            self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "before", snmp_dir, "")
                            if not bb.run_scenario(i, iterations, test_output_dir):
                                success = False
                            self.run_snmp_collection(f"{test_name}_iteration_{i+1}", "after", snmp_dir, "")
                        
                        success = success and ps.stop_config()
            
            # Collect SNMP files from scenario directory
            if iperf3_only or iperf3_darwin:
                platform_suffix = "_macOS" if iperf3_darwin else "_Linux"
                scenario_snmp_dir = os.path.join(test_output_dir, scenario_name + platform_suffix + rtt_suffix)
            else:
                scenario_snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
            scenario_snmp_files = glob.glob(os.path.join(scenario_snmp_dir, "*_SNMP_*.txt"))
            
            snmp_files = scenario_snmp_files if scenario_snmp_files else glob.glob(os.path.join(test_output_dir, "*_SNMP_*.txt"))
            
            return success, snmp_files
                
        except Exception as e:
            self.logger.error(f"Single test failed: {e}")
            return False, []
    
    def _consolidate_snmp_to_excel(self, snmp_files, parent_dir, test_group_name):
        """Consolidate all SNMP files into Excel with separate sheets"""
        try:
            if not snmp_files:
                return
            
            # Create Excel with separate sheets if pandas available
            if HAS_PANDAS:
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    
                    for snmp_file in sorted(set(snmp_files)):
                        sheet_name = os.path.basename(snmp_file).replace('.txt', '').replace('_SNMP_', '_')[:31]
                        ws = wb.create_sheet(sheet_name)
                        with open(snmp_file, 'r') as f:
                            for r_idx, line in enumerate(f, 1):
                                ws.cell(row=r_idx, column=1, value=line.rstrip())
                    
                    excel_file = os.path.join(parent_dir, f"{test_group_name}_SNMP_Combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    wb.save(excel_file)
                    self.logger.info(f"✓ Combined SNMP Excel: {os.path.basename(excel_file)}")
                except Exception as e:
                    self.logger.warning(f"Excel creation skipped: {e}")
        except Exception as e:
            self.logger.error(f"Error consolidating SNMP: {e}")

def main():
    parser = argparse.ArgumentParser(description='ByteBlower, PacketStorm, iPerf3, and SpeedTest CLI Tool')
    parser.add_argument('-byteblower', action='store_true', help='Enable ByteBlower mode')
    parser.add_argument('--bbp', required=False, help='ByteBlower .bbp file path (e.g., bb_flows/US_Classic.bbp)')
    parser.add_argument('--scenario', required=False, help='Scenario name (e.g., US_Classic_Only)')
    parser.add_argument('-test-group-name', '--test-group-name', help='Test group name (e.g., HSI029_RTT_0)')
    parser.add_argument('-packetstorm', action='store_true', help='Enable PacketStorm mode')
    parser.add_argument('--rtt', help='RTT configuration file (e.g., vcmts10ms.json)')
    parser.add_argument('-iperf3', action='store_true', help='Enable iPerf3 Linux mode')
    parser.add_argument('-iperf3-darwin', action='store_true', help='Enable iPerf3 macOS mode with iperf3-darwin')
    parser.add_argument('--clientIP', help='Linux client IP address for iPerf3 (required with -iperf3)')
    parser.add_argument('--output', choices=['json', 'txt'], default='json', help='iPerf3 output format: json or txt (default: json)')
    parser.add_argument('-speedtest', action='store_true', help='Enable SpeedTest mode')
    parser.add_argument('--client', default='linux,macos,nvidia', help='SpeedTest clients: linux,macos,nvidia (default: all)')
    parser.add_argument('--report-formats', default='html pdf csv xls xlsx json docx', help='ByteBlower report formats (default: all formats)')
    parser.add_argument('-iteration', type=int, default=1, help='Number of iterations (default: 1)')
    
    args = parser.parse_args()
    
    if not args.byteblower and not args.packetstorm and not args.iperf3 and not getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False):
        parser.error("At least one of -byteblower, -packetstorm, -iperf3, -iperf3-darwin, or -speedtest is required")
    
    if (args.iperf3 or getattr(args, 'iperf3_darwin', False)) and (not args.scenario or not args.clientIP):
        parser.error("Both --scenario and --clientIP are required when using -iperf3 or -iperf3-darwin")
    
    if args.packetstorm and not args.rtt:
        parser.error("--rtt is required when using -packetstorm")
    
    bb_file = args.bbp or 'default.bbp'
    speedtest_clients = [c.strip() for c in args.client.split(',')] if getattr(args, 'speedtest', False) else None
    
    tool = ByteBlowerCLI()
    return tool.execute(
        bb_file, 
        args.rtt or 'default.json', 
        args.iteration,
        args.scenario or 'default',
        getattr(args, 'test_group_name', None),
        getattr(args, 'clientIP', None),
        getattr(args, 'output', 'json'),
        byteblower_only=args.byteblower and not args.packetstorm and not args.iperf3 and not getattr(args, 'speedtest', False),
        packetstorm_only=args.packetstorm and not args.byteblower and not args.iperf3 and not getattr(args, 'speedtest', False),
        iperf3_only=args.iperf3 and not args.byteblower and not args.packetstorm and not getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False),
        iperf3_darwin=getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False),
        speedtest_only=getattr(args, 'speedtest', False),
        speedtest_clients=speedtest_clients,
        report_formats=getattr(args, 'report_formats', 'html pdf csv xls xlsx json docx')
    )

if __name__ == '__main__':
    sys.exit(main())
