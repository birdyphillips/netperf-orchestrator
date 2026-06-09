"""
CMTS Kafka Metrics Collector for downstream latency analysis.

Collects dp_flow_* metrics from the Harmonic vCMTS Kafka telemetry stream
during a test window and generates a latency bin report (Excel) matching
the format produced by latency_calculator.py for SNMP-based upstream data.

Usage as a module (from netperf_orchestrator.py):
    collector = CmtsCollector(broker, topic, mac)
    collector.start()
    # ... run traffic test ...
    collector.stop()
    collector.generate_report(output_dir, test_name)

Usage standalone:
    python cmts_collector.py --mac 206a949223b8 --duration 120 --output results/
"""
import os
import re
import time
import threading
from datetime import datetime
from collections import defaultdict

try:
    from kafka import KafkaConsumer
    import logging as _logging
    for _name in ["kafka", "kafka.conn", "kafka.client", "kafka.consumer",
                  "kafka.coordinator", "kafka.cluster", "kafka.protocol",
                  "kafka.metrics", "kafka.producer"]:
        _logging.getLogger(_name).setLevel(_logging.CRITICAL)
    KAFKA_AVAILABLE = True
except ImportError:
    KAFKA_AVAILABLE = False

from config_loader import config
from logger import Logger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Bin edges (16 bins) — upstream and downstream differ in bins 13-16
US_BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 100.00, 150.00, 200.00, 500.00,
]
DS_BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 60.00, 70.00, 80.00, 500.00,
]
BIN_EDGES_MS = US_BIN_EDGES_MS  # default; overridden by direction
NUM_BINS = len(BIN_EDGES_MS) - 1

# Metrics we collect per service flow
METRIC_NAMES = [
    "dp_flow_QueueLatencyAvgUsec",
    "dp_flow_QueueLatencyMaxUsec",
    "dp_flow_QueueLatencyBinPktCount",
    "dp_flow_AqmDroppedPackets",
    "dp_flow_AqmMarkedCongestedPackets",
    "dp_flow_SanctionedPackets",
    "K_Samis1_DeltaPacketsDropped",
    "K_Samis1_DeltaPacketsPassed",
    "K_Samis1_DeltaOctetsPassed",
    "K_Samis1_ServiceTimeCreated",
    "snmp_docsQosServiceFlowPackets",
    "snmp_docsQosServiceFlowOctets",
]

# Prometheus line pattern: metric_name{labels} value timestamp
_PROM_RE = re.compile(
    r'^(\w+)\{(.+?)\}\s+([\d.eE+-]+)\s+(\d+)$'
)


def _parse_labels(label_str):
    """Parse Prometheus label string into dict."""
    labels = {}
    for m in re.finditer(r'(\w+)="([^"]*)"', label_str):
        labels[m.group(1)] = m.group(2)
    return labels


class CmtsCollector:
    """Collects CMTS Kafka metrics in a background thread during a test."""

    def __init__(self, broker=None, topic=None, mac=None, direction="downstream"):
        self.logger = Logger("CmtsCollector")
        self.broker = broker or config.get('kafka', 'broker', default='65.185.232.139:11203')
        self.topic = topic or config.get('kafka', 'topic', default='cmts_metrics_apc01k1dccc')
        self.direction = direction
        self.bin_edges = DS_BIN_EDGES_MS if direction == "downstream" else US_BIN_EDGES_MS
        self.enabled = KAFKA_AVAILABLE

        if not KAFKA_AVAILABLE:
            self.logger.warning("kafka-python not installed — CMTS collection disabled")

        # Normalize MAC
        raw = (mac or config.get('vcmts', 'cm_mac', default='')).replace(':', '').replace('.', '').lower()
        self.mac_colon = ':'.join(raw[i:i+2] for i in range(0, 12, 2))

        self._thread = None
        self._stop_event = threading.Event()
        self._poll_event = threading.Event()   # signalled on each new polling timestamp
        self._seen_timestamps = set()          # track unique polling timestamps
        self._poll_count = 0

        # Collected samples: {(sfIndex, metric_name): [(timestamp_ms, value, labels), ...]}
        self.samples = defaultdict(list)
        # Bin counts per sfIndex per timestamp: {(sfIndex, timestamp): {bin_num: count}}
        self.bin_snapshots = defaultdict(lambda: defaultdict(int))
        # Raw Kafka messages matching our MAC for debugging
        self._raw_messages = []
        self._started_at = None
        self._stopped_at = None

    # ------------------------------------------------------------------
    # Start / Stop
    # ------------------------------------------------------------------

    def start(self):
        if not self.enabled:
            self.logger.warning("CMTS collection skipped (kafka-python not installed)")
            return
        if self._thread and self._thread.is_alive():
            return
        self._stop_event.clear()
        self._poll_event.clear()
        self._seen_timestamps.clear()
        self._poll_count = 0
        self.samples.clear()
        self.bin_snapshots.clear()
        self._raw_messages.clear()
        self._started_at = time.time()
        self._thread = threading.Thread(target=self._consume_loop, daemon=True)
        self._thread.start()
        self.logger.info(f"Started — broker={self.broker} topic={self.topic} mac={self.mac_colon} dir={self.direction}")

    def stop(self):
        if not self.enabled or not self._thread:
            return
        self._stop_event.set()
        self._stopped_at = time.time()
        self._thread.join(timeout=10)
        total = sum(len(v) for v in self.samples.values())
        bin_total = sum(sum(b.values()) for b in self.bin_snapshots.values())
        duration = self._stopped_at - self._started_at
        if total == 0 and bin_total == 0:
            self.logger.warning(f"No data received for MAC {self.mac_colon} — modem may not be on this vCMTS ({self.topic})")
        else:
            self.logger.info(f"Stopped — {total} metric samples, {int(bin_total)} bin packets collected in {duration:.0f}s")

    def wait_for_poll(self, timeout=45):
        """Block until the next new polling timestamp arrives from Kafka.

        Returns True if a poll was received, False on timeout.
        """
        if not self.enabled:
            return False
        count_before = self._poll_count
        self._poll_event.clear()
        deadline = time.time() + timeout
        while self._poll_count == count_before:
            remaining = deadline - time.time()
            if remaining <= 0:
                self.logger.warning(f"wait_for_poll timed out after {timeout}s")
                return False
            self._poll_event.wait(timeout=min(remaining, 1.0))
            self._poll_event.clear()
        self.logger.info(f"Poll received (count={self._poll_count})")
        return True

    # ------------------------------------------------------------------
    # Background consumer
    # ------------------------------------------------------------------

    def _consume_loop(self):
        try:
            consumer = KafkaConsumer(
                self.topic,
                bootstrap_servers=self.broker,
                group_id=f'cmts-collector-{int(time.time())}',
                auto_offset_reset='latest',
                enable_auto_commit=True,
                consumer_timeout_ms=5000,
            )
        except Exception as e:
            self.logger.error(f"Failed to connect to Kafka: {e}")
            return
        try:
            while not self._stop_event.is_set():
                for message in consumer:
                    if self._stop_event.is_set():
                        break
                    try:
                        self._process_message(message.value.decode('utf-8'))
                    except Exception:
                        pass
        finally:
            consumer.close()

    def _process_message(self, line):
        m = _PROM_RE.match(line)
        if not m:
            return
        metric_name, label_str, value_str, ts_str = m.groups()

        # Quick filter: must be a metric we care about
        if metric_name not in METRIC_NAMES:
            return

        labels = _parse_labels(label_str)

        # MAC filter
        if labels.get('cmMacAddr', '').lower() != self.mac_colon:
            return

        # Direction filter
        if labels.get('dir', '') != self.direction:
            return

        # Store raw message for debugging
        self._raw_messages.append(line)

        sf_index = labels.get('sfIndex', '0')
        value = float(value_str)
        ts = int(ts_str)

        # Bin packet counts have an extra 'bin' label
        if metric_name == "dp_flow_QueueLatencyBinPktCount":
            bin_num = int(labels.get('bin', '0'))
            self.bin_snapshots[(sf_index, ts)][bin_num] = value
        else:
            self.samples[(sf_index, metric_name)].append((ts, value, labels))

        # Signal when a new polling timestamp appears
        if ts not in self._seen_timestamps:
            self._seen_timestamps.add(ts)
            self._poll_count += 1
            self._poll_event.set()

    # ------------------------------------------------------------------
    # Peak window detection
    # ------------------------------------------------------------------

    def _find_peak_window(self, sf_index):
        """Find the interval with the highest octet delta for a given sfIndex
        and return the before/peak/after entries.

        Uses _get_all_intervals (which merges zero-delta gaps) so that rates
        reflect the true counter accumulation window.
        """
        rows = self._get_all_intervals(sf_index)
        if not rows:
            return []

        # Find peak index
        peak_idx = max(range(len(rows)), key=lambda i: rows[i]["delta"])

        # Return before + peak + after (when available)
        window = []
        for i in [peak_idx - 1, peak_idx, peak_idx + 1]:
            if 0 <= i < len(rows):
                tag = "before" if i < peak_idx else ("peak" if i == peak_idx else "after")
                window.append({**rows[i], "tag": tag})
        return window

    # ------------------------------------------------------------------
    # Report generation
    # ------------------------------------------------------------------

    def generate_report(self, output_dir, test_name="CMTS"):
        if not EXCEL_AVAILABLE:
            self.logger.warning("openpyxl not available — skipping Excel report")
            return None

        os.makedirs(output_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = os.path.join(output_dir, f"Kafka_DS_Latency_Report_{test_name}_{timestamp}.xlsx")

        # Group data by sfIndex
        sf_indices = set()
        for key in list(self.samples.keys()) + list(self.bin_snapshots.keys()):
            sf_indices.add(key[0] if isinstance(key, tuple) else key)

        if not sf_indices:
            self.logger.warning(f"No data collected for {self.mac_colon} — modem not on this vCMTS. Skipping Kafka DS report.")
            return None

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Write editable Bin_Edges sheet first (other sheets reference it)
        self._write_bin_edges_sheet(wb)

        # Raw time-series sheet (tab 1)
        self._write_timeseries(wb)

        # Summary sheet (tab 2)
        self._write_summary(wb, sorted(sf_indices))

        # Throughput / peak window sheet (tab 3)
        self._write_peak_window(wb, sorted(sf_indices))

        # Per-SF sheets
        for sf in sorted(sf_indices):
            self._write_sf_sheet(wb, sf)

        wb.save(filename)
        self.logger.info(f"Report saved: {filename}")

        # Write raw Kafka messages to txt file for verification
        self._write_raw_messages(output_dir, test_name, timestamp)

        return filename

    def _write_raw_messages(self, output_dir, test_name, timestamp):
        """Write raw Kafka messages to a txt file for parsing verification."""
        if not self._raw_messages:
            self.logger.info("No raw Kafka messages to write")
            return
        raw_file = os.path.join(output_dir, f"Kafka_Raw_Messages_{test_name}_{timestamp}.txt")
        with open(raw_file, 'w') as f:
            f.write(f"# Raw Kafka messages for MAC: {self.mac_colon} direction: {self.direction}\n")
            f.write(f"# Collected: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"# Total messages: {len(self._raw_messages)}\n")
            f.write(f"# {'='*80}\n")
            for msg in self._raw_messages:
                f.write(msg + '\n')
        self.logger.info(f"Raw Kafka messages saved: {os.path.basename(raw_file)} ({len(self._raw_messages)} lines)")

    # ------------------------------------------------------------------
    # Excel helpers
    # ------------------------------------------------------------------

    _HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
    _HEADER_FILL = PatternFill("solid", fgColor="4472C4")
    _CALC_FILL = PatternFill("solid", fgColor="D9E2F3")
    _RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
    _THIN_BORDER = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    _CENTER = Alignment(horizontal="center", vertical="center")
    _BOLD = Font(bold=True, size=11)

    _INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

    def _cell(self, ws, row, col, value, font=None, fill=None, fmt=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = self._CENTER
        cell.border = self._THIN_BORDER
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if fmt:
            cell.number_format = fmt
        return cell

    def _write_bin_edges_sheet(self, wb):
        """Write editable Bin_Edges sheet. Other sheets reference it via formulas."""
        ws = wb.create_sheet(title="Bin_Edges")
        ws.sheet_properties.tabColor = "FFC000"

        ws.merge_cells("A1:C1")
        ws["A1"] = "EDITABLE BIN EDGES (ms) \u2014 Change values below to recalculate all sheets"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A1"].alignment = self._CENTER

        self._cell(ws, 3, 1, "Bin #", font=self._HEADER_FONT, fill=self._HEADER_FILL)
        self._cell(ws, 3, 2, "Lower Edge (ms)", font=self._HEADER_FONT, fill=self._HEADER_FILL)
        self._cell(ws, 3, 3, "Upper Edge (ms)", font=self._HEADER_FONT, fill=self._HEADER_FILL)

        for i in range(NUM_BINS):
            row = 4 + i
            self._cell(ws, row, 1, i + 1)
            self._cell(ws, row, 2, self.bin_edges[i], fill=self._INPUT_FILL, fmt="0.00")
            self._cell(ws, row, 3, self.bin_edges[i + 1], fill=self._INPUT_FILL, fmt="0.00")

        for i, w in enumerate([8, 18, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_summary(self, wb, sf_indices):
        ws = wb.create_sheet(title="Summary")
        ws.sheet_properties.tabColor = "4472C4"

        ws.merge_cells("A1:L1")
        ws["A1"] = f"CMTS {self.direction.upper()} LATENCY SUMMARY — {self.mac_colon}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = self._CENTER

        headers = [
            "SF Index", "Total Bins Pkts",
            "Weighted Avg (ms)", "Avg Latency (µs)", "Max Latency (µs)",
            "P50 (ms)", "P99 (ms)", "P99.9 (ms)",
            "P50 AVG (ms)", "P99 AVG (ms)", "P99.9 AVG (ms)",
            "AQM Drops", "Congestion Marked", "Sanctioned Pkts",
            "Peak Throughput (Mbps)", "Pkt Loss %",
            "Total Pkt Delta", "Total Octet Delta",
        ]
        for col, h in enumerate(headers, 1):
            self._cell(ws, 3, col, h, font=self._HEADER_FONT, fill=self._HEADER_FILL)

        row = 4
        for sf in sf_indices:
            bin_deltas = self._get_bin_deltas(sf)
            total_pkts = sum(bin_deltas)
            avg_samples = [v for _, v, _ in self.samples.get((sf, "dp_flow_QueueLatencyAvgUsec"), [])]
            max_samples = [v for _, v, _ in self.samples.get((sf, "dp_flow_QueueLatencyMaxUsec"), [])]
            pkts_passed = [v for _, v, _ in self.samples.get((sf, "K_Samis1_DeltaPacketsPassed"), [])]
            pkts_dropped = [v for _, v, _ in self.samples.get((sf, "K_Samis1_DeltaPacketsDropped"), [])]

            avg_lat = sum(avg_samples) / len(avg_samples) if avg_samples else 0
            max_lat = max(max_samples) if max_samples else 0
            total_aqm = self._counter_delta(sf, "dp_flow_AqmDroppedPackets")
            total_cong = self._counter_delta(sf, "dp_flow_AqmMarkedCongestedPackets")
            total_sanc = self._counter_delta(sf, "dp_flow_SanctionedPackets")
            total_passed = sum(pkts_passed)
            total_dropped = sum(pkts_dropped)
            loss_pct = (total_dropped / (total_passed + total_dropped) * 100) if (total_passed + total_dropped) > 0 else 0

            # Total octets/packets across all intervals from snmp counters
            all_intervals = self._get_all_intervals(sf)
            total_octet_delta = sum(int(r["delta"]) for r in all_intervals)
            total_pkt_delta = self._get_total_pkt_delta(sf)

            # Peak throughput from highest single interval
            peak_throughput = max((r["rate_mbps"] for r in all_intervals), default=0)

            p50 = self._calc_percentile(bin_deltas, 0.50)
            p99 = self._calc_percentile(bin_deltas, 0.99)
            p999 = self._calc_percentile(bin_deltas, 0.999)
            p50a = self._calc_percentile_avg(bin_deltas, 0.50)
            p99a = self._calc_percentile_avg(bin_deltas, 0.99)
            p999a = self._calc_percentile_avg(bin_deltas, 0.999)
            weighted_avg = self._calc_weighted_avg(bin_deltas)

            self._cell(ws, row, 1, f"sfIndex {sf}")
            self._cell(ws, row, 2, total_pkts, fill=self._CALC_FILL)
            self._cell(ws, row, 3, round(weighted_avg, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 4, round(avg_lat, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 5, round(max_lat, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 6, round(p50, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 7, round(p99, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 8, round(p999, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 9, round(p50a, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 10, round(p99a, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 11, round(p999a, 4), fill=self._RESULT_FILL, fmt="0.0000")
            self._cell(ws, row, 12, int(total_aqm), fill=self._CALC_FILL)
            self._cell(ws, row, 13, int(total_cong), fill=self._CALC_FILL)
            self._cell(ws, row, 14, int(total_sanc), fill=self._CALC_FILL)
            self._cell(ws, row, 15, round(peak_throughput, 4), fill=self._CALC_FILL, fmt="0.0000")
            self._cell(ws, row, 16, round(loss_pct, 4), fill=self._CALC_FILL, fmt="0.0000")
            self._cell(ws, row, 17, total_pkt_delta, fill=self._CALC_FILL)
            self._cell(ws, row, 18, total_octet_delta, fill=self._CALC_FILL)
            row += 1

        # TOTAL row — combined across all sfIndices
        total_row = row
        sum_throughput = 0
        sum_pkt_delta = 0
        sum_octet_delta = 0
        for sf in sf_indices:
            ai = self._get_all_intervals(sf)
            if ai:
                peak_r = max(ai, key=lambda r: r["delta"])
                sum_throughput += peak_r["rate_mbps"]
            sum_octet_delta += sum(int(r["delta"]) for r in ai)
            sum_pkt_delta += self._get_total_pkt_delta(sf)
        self._cell(ws, total_row, 1, "TOTAL", font=self._BOLD, fill=self._RESULT_FILL)
        self._cell(ws, total_row, 15, round(sum_throughput, 4), font=self._BOLD, fill=self._RESULT_FILL, fmt="0.0000")
        self._cell(ws, total_row, 17, sum_pkt_delta, font=self._BOLD, fill=self._RESULT_FILL)
        self._cell(ws, total_row, 18, sum_octet_delta, font=self._BOLD, fill=self._RESULT_FILL)

        for i, w in enumerate([14, 16, 18, 18, 18, 14, 14, 14, 16, 16, 16, 14, 18, 16, 18, 14, 16, 18], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_sf_sheet(self, wb, sf_index):
        """Write per-SF sheet with Excel formulas referencing Bin_Edges sheet."""
        ws = wb.create_sheet(title=f"sfIndex_{sf_index}")
        bin_deltas = self._get_bin_deltas(sf_index)
        total = sum(bin_deltas)

        ws.merge_cells("A1:H1")
        ws["A1"] = f"CMTS {self.direction.upper()} LATENCY \u2014 sfIndex {sf_index}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = self._CENTER

        headers = [
            "BIN", "LOWER (ms)", "UPPER (ms)", "AVG (ms)",
            "DELTA", "CUMULATIVE", "CUMULATIVE %", "BIN %",
        ]
        for col, h in enumerate(headers, 1):
            self._cell(ws, 3, col, h, font=self._HEADER_FONT, fill=self._HEADER_FILL)

        first_data = 4
        last_data = 4 + NUM_BINS - 1

        for i in range(NUM_BINS):
            row = 4 + i
            be_row = 4 + i  # corresponding row in Bin_Edges sheet

            self._cell(ws, row, 1, i + 1)
            # Col B: LOWER from Bin_Edges
            c = self._cell(ws, row, 2, None, fmt="0.00")
            c.value = f"=Bin_Edges!B{be_row}"
            # Col C: UPPER from Bin_Edges
            c = self._cell(ws, row, 3, None, fmt="0.00")
            c.value = f"=Bin_Edges!C{be_row}"
            # Col D: AVG = (LOWER + UPPER) / 2
            c = self._cell(ws, row, 4, None, fill=self._CALC_FILL, fmt="0.0000")
            c.value = f"=(B{row}+C{row})/2"
            # Col E: DELTA (raw data)
            self._cell(ws, row, 5, int(bin_deltas[i]), fill=self._CALC_FILL)
            # Col F: CUMULATIVE
            if i == 0:
                c = self._cell(ws, row, 6, None, fill=self._CALC_FILL)
                c.value = f"=E{row}"
            else:
                c = self._cell(ws, row, 6, None, fill=self._CALC_FILL)
                c.value = f"=F{row-1}+E{row}"
            # Col G: CUMULATIVE %
            total_cell = f"E{4 + NUM_BINS + 1}"
            c = self._cell(ws, row, 7, None, fill=self._CALC_FILL, fmt="0.00")
            c.value = f"=IF({total_cell}=0,0,F{row}/{total_cell}*100)"
            # Col H: BIN %
            c = self._cell(ws, row, 8, None, fill=self._CALC_FILL, fmt="0.00")
            c.value = f"=IF({total_cell}=0,0,E{row}/{total_cell}*100)"

        total_row = 4 + NUM_BINS + 1
        self._cell(ws, total_row, 1, "TOTAL", font=self._BOLD)
        c = self._cell(ws, total_row, 5, None, font=self._BOLD, fill=self._CALC_FILL)
        c.value = f"=SUM(E{first_data}:E{last_data})"
        c = self._cell(ws, total_row, 7, None, font=self._BOLD, fill=self._CALC_FILL, fmt="0.00")
        c.value = f"=IF(E{total_row}=0,0,100)"

        # Percentiles (Linear Interpolation) with formulas
        pct_row = total_row + 2
        ws.merge_cells(f"A{pct_row}:H{pct_row}")
        ws.cell(row=pct_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

        for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
            pct_row += 1
            self._cell(ws, pct_row, 1, f"{label} TARGET", font=self._BOLD)
            c = self._cell(ws, pct_row, 2, None, fill=self._RESULT_FILL, fmt="0.00")
            c.value = f"=E{total_row}*{pct}"
            self._cell(ws, pct_row, 3, f"{label} (ms)", font=self._BOLD)
            tgt = f"B{pct_row}"
            fd = first_data
            ld = last_data
            idx_expr = f"MATCH({tgt},F{fd}:F{ld},1)+1"
            c = self._cell(ws, pct_row, 4, None, fill=self._RESULT_FILL, fmt="0.0000")
            c.value = (
                f'=IF(E{total_row}=0,0,'
                f'INDEX(B{fd}:B{ld},{idx_expr})'
                f'+IF(INDEX(E{fd}:E{ld},{idx_expr})=0,0,'
                f'({tgt}-IF({idx_expr}=1,0,INDEX(F{fd}:F{ld},{idx_expr}-1)))'
                f'/INDEX(E{fd}:E{ld},{idx_expr})'
                f'*(INDEX(C{fd}:C{ld},{idx_expr})-INDEX(B{fd}:B{ld},{idx_expr}))))'
            )

        legend_row = pct_row + 2
        ws.cell(row=legend_row, column=1, value="FORMULA:").font = self._BOLD
        ws.merge_cells(f"B{legend_row}:H{legend_row}")
        ws.cell(row=legend_row, column=2, value="P = BIN_LOW + ((TARGET \u2212 PREV_CUMULATIVE) / BIN_COUNT) \u00d7 (BIN_HIGH \u2212 BIN_LOW)").font = Font(italic=True, size=10)

        # Percentiles (AVG Method) with formulas
        avg_row = legend_row + 2
        ws.merge_cells(f"A{avg_row}:H{avg_row}")
        ws.cell(row=avg_row, column=1, value="PERCENTILE RESULTS (AVG METHOD)").font = Font(bold=True, size=12)

        for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
            avg_row += 1
            self._cell(ws, avg_row, 1, f"{label} TARGET", font=self._BOLD)
            c = self._cell(ws, avg_row, 2, None, fill=self._RESULT_FILL, fmt="0.00")
            c.value = f"=E{total_row}*{pct}"
            self._cell(ws, avg_row, 3, f"{label} AVG (ms)", font=self._BOLD)
            tgt = f"B{avg_row}"
            fd = first_data
            ld = last_data
            idx_expr = f"MATCH({tgt},F{fd}:F{ld},1)+1"
            c = self._cell(ws, avg_row, 4, None, fill=self._RESULT_FILL, fmt="0.0000")
            c.value = (
                f'=IF(E{total_row}=0,0,'
                f'INDEX(D{fd}:D{ld},{idx_expr}))'
            )

        legend_row2 = avg_row + 2
        ws.cell(row=legend_row2, column=1, value="FORMULA:").font = self._BOLD
        ws.merge_cells(f"B{legend_row2}:H{legend_row2}")
        ws.cell(row=legend_row2, column=2, value="P = AVG (col D) of first bin where cumulative count >= percentile target").font = Font(italic=True, size=10)

        # Weighted Average formula
        wavg_row = legend_row2 + 2
        self._cell(ws, wavg_row, 1, "Weighted Avg (ms)", font=self._BOLD)
        c = self._cell(ws, wavg_row, 2, None, fill=self._RESULT_FILL, fmt="0.0000")
        c.value = f"=IF(E{total_row}=0,0,SUMPRODUCT(E{first_data}:E{last_data},D{first_data}:D{last_data})/E{total_row})"
        ws.cell(row=wavg_row + 1, column=1, value="FORMULA:").font = self._BOLD
        ws.merge_cells(f"B{wavg_row + 1}:H{wavg_row + 1}")
        ws.cell(row=wavg_row + 1, column=2, value="Weighted Avg = SUM(delta_i \u00d7 bin_avg_i) / total_packets").font = Font(italic=True, size=10)

        for i, w in enumerate([8, 14, 14, 14, 14, 16, 16, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_peak_window(self, wb, sf_indices):
        """Write all octet-delta intervals per sfIndex with peak highlighted."""
        ws = wb.create_sheet(title="Throughput")
        ws.sheet_properties.tabColor = "00B050"

        ws.merge_cells("A1:I1")
        ws["A1"] = f"QOS SERVICE FLOW OCTETS — {self.mac_colon}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = self._CENTER

        headers = ["SF Index", "Poll Before", "Poll After",
                   "Before Octets", "After Octets", "Delta Octets",
                   "Interval (s)", "Rate (Mbps)", "Tag"]
        for col, h in enumerate(headers, 1):
            self._cell(ws, 3, col, h, font=self._HEADER_FONT, fill=self._HEADER_FILL)

        row = 4
        grand_total_delta = 0
        grand_total_rate = 0

        for sf in sf_indices:
            all_rows = self._get_all_intervals(sf)
            if not all_rows:
                continue

            # Find peak index
            peak_idx = max(range(len(all_rows)), key=lambda i: all_rows[i]["delta"])
            sf_total_delta = 0

            for i, entry in enumerate(all_rows):
                is_peak = (i == peak_idx)
                fill = self._RESULT_FILL if is_peak else None
                tag = "PEAK" if is_peak else ""
                ts_before_dt = datetime.utcfromtimestamp(entry["ts_before"] / 1000)
                ts_after_dt = datetime.utcfromtimestamp(entry["timestamp_ms"] / 1000)
                self._cell(ws, row, 1, f"sfIndex {sf}")
                self._cell(ws, row, 2, ts_before_dt.strftime("%Y-%m-%d %H:%M:%S.%f"), fill=fill)
                self._cell(ws, row, 3, ts_after_dt.strftime("%Y-%m-%d %H:%M:%S.%f"), fill=fill)
                self._cell(ws, row, 4, int(entry["before"]), fill=fill)
                self._cell(ws, row, 5, int(entry["after"]), fill=fill)
                self._cell(ws, row, 6, int(entry["delta"]), fill=fill)
                self._cell(ws, row, 7, round(entry["interval"], 6), fill=fill, fmt="0.000000")
                self._cell(ws, row, 8, round(entry["rate_mbps"], 4), fill=fill, fmt="0.0000")
                self._cell(ws, row, 9, tag, font=self._BOLD if is_peak else None, fill=fill)
                sf_total_delta += int(entry["delta"])
                row += 1

            # SF total row
            total_duration = sum(r["interval"] for r in all_rows) if all_rows else 0
            sf_total_rate = (sf_total_delta * 8) / (total_duration * 1_000_000) if total_duration > 0 else 0
            self._cell(ws, row, 1, f"sfIndex {sf}", font=self._BOLD)
            self._cell(ws, row, 5, "TOTAL", font=self._BOLD)
            self._cell(ws, row, 6, sf_total_delta, font=self._BOLD, fill=self._CALC_FILL)
            row += 1

            grand_total_delta += sf_total_delta
            row += 1  # blank separator

        # Combined total across all sfIndices
        self._cell(ws, row, 1, "ALL SF COMBINED", font=self._BOLD, fill=self._RESULT_FILL)
        self._cell(ws, row, 5, "TOTAL", font=self._BOLD, fill=self._RESULT_FILL)
        self._cell(ws, row, 6, grand_total_delta, font=self._BOLD, fill=self._RESULT_FILL)

        for i, w in enumerate([12, 28, 28, 18, 18, 18, 14, 14, 8], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_timeseries(self, wb):
        """Write raw time-series data for charting."""
        ws = wb.create_sheet(title="TimeSeries")
        ws.sheet_properties.tabColor = "FF6600"

        headers = ["Timestamp", "Time (s)", "SF Index", "Metric", "Value"]
        for col, h in enumerate(headers, 1):
            self._cell(ws, 1, col, h, font=self._HEADER_FONT, fill=self._HEADER_FILL)

        row = 2
        base_ts = self._started_at * 1000 if self._started_at else 0
        for (sf, metric), entries in sorted(self.samples.items()):
            for ts, val, _ in entries:
                self._cell(ws, row, 1, datetime.fromtimestamp(ts / 1000).strftime("%H:%M:%S"))
                self._cell(ws, row, 2, round((ts - base_ts) / 1000, 1))
                self._cell(ws, row, 3, sf)
                self._cell(ws, row, 4, metric)
                self._cell(ws, row, 5, round(val, 6) if val != int(val) else int(val))
                row += 1

        # Write bin snapshot data (all 16 bins per poll, including zero-traffic polls)
        # Collect all sf indices that have bin data
        bin_sf_indices = set(sf for (sf, ts) in self.bin_snapshots.keys())
        all_timestamps = sorted(self._seen_timestamps)
        for sf in sorted(bin_sf_indices):
            for ts in all_timestamps:
                bins = self.bin_snapshots.get((sf, ts), {})
                for bin_num in range(1, NUM_BINS + 1):
                    self._cell(ws, row, 1, datetime.fromtimestamp(ts / 1000).strftime("%H:%M:%S"))
                    self._cell(ws, row, 2, round((ts - base_ts) / 1000, 1))
                    self._cell(ws, row, 3, sf)
                    self._cell(ws, row, 4, f"dp_flow_QueueLatencyBinPktCount_bin{bin_num}")
                    self._cell(ws, row, 5, int(bins.get(bin_num, 0)))
                    row += 1

        for i, w in enumerate([14, 12, 12, 40, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _get_all_intervals(self, sf_index):
        """Build all consecutive interval deltas for snmp_docsQosServiceFlowOctets.

        Skips zero-delta intervals (no traffic) so that throughput calculations
        only reflect intervals where data actually flowed.
        """
        samples = self.samples.get((sf_index, "snmp_docsQosServiceFlowOctets"), [])
        if len(samples) < 2:
            return []
        pts = sorted(samples, key=lambda x: x[0])
        rows = []
        for i in range(1, len(pts)):
            ts_before, val_before, _ = pts[i - 1]
            ts_after, val_after, _ = pts[i]
            delta = val_after - val_before
            if delta == 0:
                continue
            interval_s = (ts_after - ts_before) / 1000.0
            rate = (delta * 8) / (interval_s * 1_000_000) if interval_s > 0 else 0
            rows.append({
                "timestamp_ms": ts_after,
                "ts_before": ts_before,
                "before": val_before,
                "after": val_after,
                "delta": delta,
                "interval": interval_s,
                "rate_mbps": rate,
            })
        return rows

    def _get_total_pkt_delta(self, sf_index):
        """Return total packet delta (last - first) from snmp_docsQosServiceFlowPackets."""
        samples = self.samples.get((sf_index, "snmp_docsQosServiceFlowPackets"), [])
        if len(samples) < 2:
            return 0
        pts = sorted(samples, key=lambda x: x[0])
        return int(pts[-1][1] - pts[0][1])

    def _get_peak_pkt_delta(self, sf_index, peak_entry):
        """Get the packet delta for the same interval as the peak octet entry."""
        if not peak_entry:
            return 0
        samples = self.samples.get((sf_index, "snmp_docsQosServiceFlowPackets"), [])
        if len(samples) < 2:
            return 0
        pts = sorted(samples, key=lambda x: x[0])
        for i in range(1, len(pts)):
            if pts[i][0] == peak_entry["timestamp_ms"]:
                return int(pts[i][1] - pts[i - 1][1])
        return 0

    def _counter_delta(self, sf_index, metric_name):
        """Return last - first value for a cumulative counter metric."""
        samples = self.samples.get((sf_index, metric_name), [])
        if len(samples) < 2:
            return 0
        pts = sorted(samples, key=lambda x: x[0])
        return int(pts[-1][1] - pts[0][1])

    # ------------------------------------------------------------------
    # Bin delta & percentile helpers
    # ------------------------------------------------------------------

    def _get_bin_deltas(self, sf_index):
        """Sum bin packet counts across all timestamps for a given sfIndex.

        Each 30s snapshot gives the delta bin counts for that interval,
        so we sum them to get total counts over the test window.
        """
        totals = [0] * NUM_BINS
        for (sf, ts), bins in self.bin_snapshots.items():
            if sf != sf_index:
                continue
            for bin_num, count in bins.items():
                idx = bin_num - 1  # bin labels are 1-based
                if 0 <= idx < NUM_BINS:
                    totals[idx] += int(count)
        return totals

    def _calc_percentile(self, deltas, percentile):
        total = sum(deltas)
        if total == 0:
            return 0.0
        target = total * percentile
        cumulative = 0
        for i, count in enumerate(deltas):
            cumulative += count
            if cumulative >= target:
                prev_cum = cumulative - count
                low = self.bin_edges[i]
                high = self.bin_edges[i + 1]
                denom = count if count > 0 else 1
                return low + ((target - prev_cum) / denom) * (high - low)
        return self.bin_edges[-1]

    def _calc_percentile_avg(self, deltas, percentile):
        """AVG method: return the bin midpoint of the first bin where
        cumulative count >= percentile target."""
        total = sum(deltas)
        if total == 0:
            return 0.0
        target = total * percentile
        cumulative = 0
        for i, count in enumerate(deltas):
            cumulative += count
            if cumulative >= target:
                return (self.bin_edges[i] + self.bin_edges[i + 1]) / 2
        return (self.bin_edges[-2] + self.bin_edges[-1]) / 2

    def _calc_weighted_avg(self, deltas):
        """Weighted average latency: sum(delta_i × bin_avg_i) / total."""
        total = sum(deltas)
        if total == 0:
            return 0.0
        weighted = sum(
            deltas[i] * (self.bin_edges[i] + self.bin_edges[i + 1]) / 2
            for i in range(NUM_BINS)
        )
        return weighted / total


# ---------------------------------------------------------------------------
# Standalone CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse as _ap

    p = _ap.ArgumentParser(description="CMTS Kafka metrics collector")
    p.add_argument("--broker", default=None, help="Kafka broker")
    p.add_argument("--topic", default=None, help="Kafka topic")
    p.add_argument("--mac", required=True, help="CM MAC address")
    p.add_argument("--direction", default="downstream", choices=["downstream", "upstream"])
    p.add_argument("--duration", type=int, default=120, help="Collection duration in seconds")
    p.add_argument("--output", default=".", help="Output directory")
    p.add_argument("--test-name", default="manual", help="Test name for report filename")
    args = p.parse_args()

    collector = CmtsCollector(
        broker=args.broker,
        topic=args.topic,
        mac=args.mac,
        direction=args.direction,
    )
    collector.start()
    try:
        print(f"Collecting for {args.duration}s... (Ctrl+C to stop early)")
        for i in range(args.duration):
            if collector._stop_event.is_set():
                break
            time.sleep(1)
    except KeyboardInterrupt:
        pass
    collector.stop()
    collector.generate_report(args.output, args.test_name)
