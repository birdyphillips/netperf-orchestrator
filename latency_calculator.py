"""
CMTS Downstream Latency Bin Calculator.

Parses SNMP before/after result files, computes per-service-flow latency bin
deltas, and generates an Excel workbook with P50/P99/P99.9 percentiles.

Also retains the original blank template generation via --template flag.
"""
import os
import re
import sys
import glob
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 17 edges → 16 bins.  Bin 1: 0–0.05 ms … Bin 16: 200–500 ms
BIN_EDGES_MS = [
    0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
    20.00, 30.00, 40.00, 50.00, 100.00, 150.00, 200.00, 500.00,
]
NUM_BINS = len(BIN_EDGES_MS) - 1  # 16

# SNMP OID prefix for latency stats bins
# Full OID: enterprises.4491.2.1.21.1.29.2.1.{sub_oid}.2.{sfid}
# sub_oid 1 = Gauge32 (active bin count)
# sub_oid 2..18 = Counter64 bin packet counts (bins 1-17, we use 2-18 → 17 values, first 16 are our bins)
LATENCY_OID_PREFIX = "enterprises.4491.2.1.21.1.29.2.1"


# ---------------------------------------------------------------------------
# Parsing
# ---------------------------------------------------------------------------

def parse_latency_bins(filepath):
    """Parse Latency Stats Table from an SNMP text file.

    Returns dict: {sfid: {bin_sub_oid: count, ...}, ...}
    bin_sub_oid ranges 2..18 (maps to bins 1-17).
    """
    with open(filepath, "r") as f:
        content = f.read()

    # Extract only the Latency Stats Table section
    match = re.search(
        r"Latency Stats Table\n=+\n(.*?)(?:\n\n|\Z)", content, re.DOTALL
    )
    if not match:
        return {}

    section = match.group(1)
    bins = {}  # {sfid: {sub_oid: value}}

    for line in section.splitlines():
        # Match: ...29.2.1.{sub}.2.{sfid} = Counter64: {val}
        m = re.search(
            r"\.29\.2\.1\.(\d+)\.2\.(\d+)\s*=\s*(?:Counter64|Gauge32):\s*(\d+)",
            line,
        )
        if m:
            sub_oid = int(m.group(1))
            sfid = int(m.group(2))
            value = int(m.group(3))
            bins.setdefault(sfid, {})[sub_oid] = value

    return bins


def compute_deltas(before_bins, after_bins):
    """Compute per-service-flow bin deltas (after − before).

    Returns dict: {sfid: {"before": [...], "after": [...], "deltas": [...]}}
    Only includes SFIDs present in both files with non-zero total delta.
    """
    results = {}
    all_sfids = set(before_bins) & set(after_bins)

    for sfid in sorted(all_sfids):
        before_vals = []
        after_vals = []
        deltas = []
        for sub in range(2, 2 + NUM_BINS):  # sub_oid 2..17
            bv = before_bins[sfid].get(sub, 0)
            av = after_bins[sfid].get(sub, 0)
            before_vals.append(bv)
            after_vals.append(av)
            deltas.append(max(av - bv, 0))
        if sum(deltas) > 0:
            results[sfid] = {"before": before_vals, "after": after_vals, "deltas": deltas}

    return results


# ---------------------------------------------------------------------------
# Percentile calculation (linear interpolation within bin)
# ---------------------------------------------------------------------------

def calc_percentile(deltas, percentile):
    """Linear interpolation percentile from bin deltas.

    P = bin_low + ((target − prev_cumulative) / bin_count) × (bin_high − bin_low)
    """
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            prev_cum = cumulative - count
            bin_low = BIN_EDGES_MS[i]
            bin_high = BIN_EDGES_MS[i + 1]
            denom = count if count > 0 else 1
            return bin_low + ((target - prev_cum) / denom) * (bin_high - bin_low)
    return BIN_EDGES_MS[-1]


def calc_weighted_avg(deltas):
    """Weighted average latency: sum(delta_i × bin_avg_i) / total."""
    total = sum(deltas)
    if total == 0:
        return 0.0
    weighted = sum(
        deltas[i] * (BIN_EDGES_MS[i] + BIN_EDGES_MS[i + 1]) / 2
        for i in range(NUM_BINS)
    )
    return weighted / total


def calc_percentile_avg(deltas, percentile):
    """AVG method: return the bin midpoint (AVG) of the first bin where
    cumulative count >= percentile target."""
    total = sum(deltas)
    if total == 0:
        return 0.0
    target = total * percentile
    cumulative = 0
    for i, count in enumerate(deltas):
        cumulative += count
        if cumulative >= target:
            return (BIN_EDGES_MS[i] + BIN_EDGES_MS[i + 1]) / 2
    return (BIN_EDGES_MS[-2] + BIN_EDGES_MS[-1]) / 2


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

# Styles (module-level so they're reusable)
_HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_CALC_FILL = PatternFill("solid", fgColor="D9E2F3")
_RESULT_FILL = PatternFill("solid", fgColor="C6EFCE")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_BOLD = Font(bold=True, size=11)


def _styled_cell(ws, row, col, value, font=None, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = _CENTER
    cell.border = _THIN_BORDER
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")


def write_sf_sheet(wb, sheet_name, sf_data):
    """Write one worksheet for a service flow's latency bin data."""
    deltas = sf_data["deltas"]
    before = sf_data["before"]
    after = sf_data["after"]
    ws = wb.create_sheet(title=sheet_name)

    # Title
    ws.merge_cells("A1:J1")
    ws["A1"] = f"CMTS DS LATENCY — {sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    # Headers (row 3)
    headers = [
        "BIN", "LOWER (ms)", "UPPER (ms)", "AVG (ms)",
        "START COUNT", "END COUNT", "DELTA",
        "CUMULATIVE", "CUMULATIVE %", "BIN %",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    total = sum(deltas)
    cumulative = 0

    for i in range(NUM_BINS):
        row = 4 + i
        low = BIN_EDGES_MS[i]
        high = BIN_EDGES_MS[i + 1]
        avg = (low + high) / 2
        delta = deltas[i]
        cumulative += delta
        cum_pct = (cumulative / total * 100) if total else 0
        bin_pct = (delta / total * 100) if total else 0

        _styled_cell(ws, row, 1, i + 1)
        _styled_cell(ws, row, 2, low, fmt="0.00")
        _styled_cell(ws, row, 3, high if i < 15 else "200.00+", fmt="0.00")
        _styled_cell(ws, row, 4, avg, fill=_CALC_FILL, fmt="0.0000")
        _styled_cell(ws, row, 5, before[i], fill=_INPUT_FILL)
        _styled_cell(ws, row, 6, after[i], fill=_INPUT_FILL)
        _styled_cell(ws, row, 7, delta, fill=_CALC_FILL)
        _styled_cell(ws, row, 8, cumulative, fill=_CALC_FILL)
        _styled_cell(ws, row, 9, cum_pct, fill=_CALC_FILL, fmt="0.00")
        _styled_cell(ws, row, 10, bin_pct, fill=_CALC_FILL, fmt="0.00")

    # Totals row
    total_row = 4 + NUM_BINS + 1
    _styled_cell(ws, total_row, 1, "TOTAL", font=_BOLD)
    _styled_cell(ws, total_row, 5, sum(before), font=_BOLD, fill=_INPUT_FILL)
    _styled_cell(ws, total_row, 6, sum(after), font=_BOLD, fill=_INPUT_FILL)
    _styled_cell(ws, total_row, 7, total, font=_BOLD, fill=_CALC_FILL)
    _styled_cell(ws, total_row, 9, 100.00 if total else 0, font=_BOLD, fill=_CALC_FILL, fmt="0.00")

    # Percentile results (Linear Interpolation)
    pct_row = total_row + 2
    ws.merge_cells(f"A{pct_row}:J{pct_row}")
    ws.cell(row=pct_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

    for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
        pct_row += 1
        target_count = total * pct
        result_ms = calc_percentile(deltas, pct)
        _styled_cell(ws, pct_row, 1, f"{label} TARGET", font=_BOLD)
        _styled_cell(ws, pct_row, 2, round(target_count, 2), fill=_RESULT_FILL, fmt="0.00")
        _styled_cell(ws, pct_row, 3, f"{label} (ms)", font=_BOLD)
        _styled_cell(ws, pct_row, 4, round(result_ms, 4), fill=_RESULT_FILL, fmt="0.0000")

    # Linear interpolation formula legend
    legend_row = pct_row + 2
    ws.cell(row=legend_row, column=1, value="FORMULA:").font = _BOLD
    ws.merge_cells(f"B{legend_row}:J{legend_row}")
    ws.cell(
        row=legend_row, column=2,
        value="P = BIN_LOW + ((TARGET − PREV_CUMULATIVE) / BIN_COUNT) × (BIN_HIGH − BIN_LOW)",
    ).font = Font(italic=True, size=10)

    # Percentile results (AVG Method)
    avg_row = legend_row + 2
    ws.merge_cells(f"A{avg_row}:J{avg_row}")
    ws.cell(row=avg_row, column=1, value="PERCENTILE RESULTS (AVG METHOD)").font = Font(bold=True, size=12)

    for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
        avg_row += 1
        target_count = total * pct
        result_ms = calc_percentile_avg(deltas, pct)
        _styled_cell(ws, avg_row, 1, f"{label} TARGET", font=_BOLD)
        _styled_cell(ws, avg_row, 2, round(target_count, 2), fill=_RESULT_FILL, fmt="0.00")
        _styled_cell(ws, avg_row, 3, f"{label} AVG (ms)", font=_BOLD)
        _styled_cell(ws, avg_row, 4, round(result_ms, 4), fill=_RESULT_FILL, fmt="0.0000")

    # AVG method formula legend
    legend_row2 = avg_row + 2
    ws.cell(row=legend_row2, column=1, value="FORMULA:").font = _BOLD
    ws.merge_cells(f"B{legend_row2}:J{legend_row2}")
    ws.cell(
        row=legend_row2, column=2,
        value="P = AVG (col D) of first bin where cumulative count >= percentile target",
    ).font = Font(italic=True, size=10)

    # Column widths
    for i, w in enumerate([8, 14, 14, 14, 16, 16, 14, 16, 16, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_summary_sheet(wb, all_results):
    """Write a summary sheet comparing all service flows."""
    ws = wb.create_sheet(title="Summary")
    ws.sheet_properties.tabColor = "4472C4"

    ws.merge_cells("A1:I1")
    ws["A1"] = "LATENCY PERCENTILE SUMMARY"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    headers = [
        "Service Flow", "Total Pkts",
        "P50 (ms)", "P99 (ms)", "P99.9 (ms)",
        "P50 AVG (ms)", "P99 AVG (ms)", "P99.9 AVG (ms)",
        "Peak Bin",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 3, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    row = 4
    for sfid, sf_data in sorted(all_results.items()):
        deltas = sf_data["deltas"]
        total = sum(deltas)
        p50 = calc_percentile(deltas, 0.50)
        p99 = calc_percentile(deltas, 0.99)
        p999 = calc_percentile(deltas, 0.999)
        p50a = calc_percentile_avg(deltas, 0.50)
        p99a = calc_percentile_avg(deltas, 0.99)
        p999a = calc_percentile_avg(deltas, 0.999)
        peak_idx = deltas.index(max(deltas))
        peak_label = f"{BIN_EDGES_MS[peak_idx]}-{BIN_EDGES_MS[peak_idx+1]} ms"

        _styled_cell(ws, row, 1, f"SFID {sfid}")
        _styled_cell(ws, row, 2, total, fill=_CALC_FILL)
        _styled_cell(ws, row, 3, round(p50, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 4, round(p99, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 5, round(p999, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 6, round(p50a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 7, round(p99a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 8, round(p999a, 4), fill=_RESULT_FILL, fmt="0.0000")
        _styled_cell(ws, row, 9, peak_label, fill=_CALC_FILL)
        row += 1

    for i, w in enumerate([16, 14, 14, 14, 14, 16, 16, 16, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def generate_latency_report(before_file, after_file, output_file=None):
    """Main entry: parse SNMP files, compute deltas, write Excel report.

    Returns the output file path.
    """
    before_bins = parse_latency_bins(before_file)
    after_bins = parse_latency_bins(after_file)

    if not before_bins or not after_bins:
        print("WARNING: No latency stats found in one or both SNMP files.")
        print(f"  Before file SFIDs: {sorted(before_bins.keys()) if before_bins else 'none'}")
        print(f"  After file SFIDs:  {sorted(after_bins.keys()) if after_bins else 'none'}")
        return None

    all_deltas = compute_deltas(before_bins, after_bins)

    if not all_deltas:
        print("WARNING: All latency bin deltas are zero — no traffic detected.")
        return None

    if output_file is None:
        output_dir = os.path.dirname(after_file) or "."
        dir_name = os.path.basename(os.path.abspath(output_dir))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"Latency_Bin_Report_{dir_name}_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Summary first
    write_summary_sheet(wb, all_deltas)

    # One sheet per service flow
    for sfid, sf_data in sorted(all_deltas.items()):
        write_sf_sheet(wb, f"SFID_{sfid}", sf_data)

    wb.save(output_file)
    print(f"Latency report saved: {output_file}")

    # Print summary to console
    print("\n--- Latency Summary ---")
    for sfid, sf_data in sorted(all_deltas.items()):
        deltas = sf_data["deltas"]
        total = sum(deltas)
        avg = calc_weighted_avg(deltas)
        p50 = calc_percentile(deltas, 0.50)
        p99 = calc_percentile(deltas, 0.99)
        p999 = calc_percentile(deltas, 0.999)
        print(f"  SFID {sfid}: {total} pkts | AVG={avg:.4f}ms  P50={p50:.4f}ms  P99={p99:.4f}ms  P99.9={p999:.4f}ms")

    return output_file


def find_snmp_files(results_dir):
    """Auto-discover before/after SNMP .txt files in a results directory."""
    pattern = os.path.join(results_dir, "**", "*SNMP_before_*.txt")
    before_files = sorted(glob.glob(pattern, recursive=True))

    pattern = os.path.join(results_dir, "**", "*SNMP_after_*.txt")
    after_files = sorted(glob.glob(pattern, recursive=True))

    if not before_files or not after_files:
        return None, None

    return before_files[0], after_files[0]


# ---------------------------------------------------------------------------
# Blank template generation (original functionality)
# ---------------------------------------------------------------------------

def create_template(output_file="CMTS_Latency_Bin_Template.xlsx"):
    """Generate a fillable Excel template with formulas (no data)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Latency Bins"

    input_fill = PatternFill("solid", fgColor="FFF2CC")

    ws.merge_cells("A1:I1")
    ws["A1"] = "CMTS DOWNSTREAM LATENCY BIN ANALYSIS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    ws.merge_cells("A2:I2")
    ws["A2"] = "Yellow = input  |  Blue = calculated  |  Green = results"
    ws["A2"].font = Font(italic=True, size=10)

    headers = [
        "BIN", "LOWER (MS)", "UPPER (MS)", "AVG (MS)",
        "START COUNT", "END COUNT", "DELTA", "CUMULATIVE", "CUMULATIVE %",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 4, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    start_row = 5
    for i in range(NUM_BINS):
        row = start_row + i
        low = BIN_EDGES_MS[i]
        high = BIN_EDGES_MS[i + 1]

        _styled_cell(ws, row, 1, i + 1)
        _styled_cell(ws, row, 2, low, fmt="0.00")
        _styled_cell(ws, row, 3, "200.00+" if i == 15 else high, fmt="0.00")

        # AVG formula
        avg_formula = f"=(B{row}+200)/2" if i == 15 else f"=(B{row}+C{row})/2"
        _styled_cell(ws, row, 4, avg_formula, fill=_CALC_FILL, fmt="0.0000")

        # Fillable start/end
        _styled_cell(ws, row, 5, 0, fill=input_fill)
        _styled_cell(ws, row, 6, 0, fill=input_fill)

        # Delta
        _styled_cell(ws, row, 7, f"=F{row}-E{row}", fill=_CALC_FILL)

        # Cumulative
        cum = f"=G{row}" if i == 0 else f"=H{row-1}+G{row}"
        _styled_cell(ws, row, 8, cum, fill=_CALC_FILL)

        # Cumulative %
        end_row = start_row + 15
        _styled_cell(ws, row, 9, f"=IF(H${end_row}=0,0,H{row}/H${end_row}*100)", fill=_CALC_FILL, fmt="0.00")

    end_row = start_row + 15

    # Totals
    total_row = end_row + 2
    _styled_cell(ws, total_row, 1, "TOTAL", font=_BOLD)
    _styled_cell(ws, total_row, 7, f"=SUM(G{start_row}:G{end_row})", font=_BOLD, fill=_CALC_FILL)
    _styled_cell(ws, total_row, 9, "100.00", font=_BOLD, fill=_CALC_FILL)

    # Percentile formulas
    result_row = total_row + 2
    ws.merge_cells(f"A{result_row}:I{result_row}")
    ws.cell(row=result_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

    for label, pct in [("P50", 0.50), ("P99", 0.99), ("P99.9", 0.999)]:
        result_row += 1
        ws.cell(row=result_row, column=1, value=f"{label} TARGET").font = _BOLD
        _styled_cell(ws, result_row, 2, f"=H{end_row}*{pct}", fill=_RESULT_FILL, fmt="0.00")
        ws.cell(row=result_row, column=3, value=f"{label} (MS)").font = _BOLD
        _styled_cell(ws, result_row, 4, _build_interpolation_formula(start_row, end_row, pct), fill=_RESULT_FILL, fmt="0.0000")

    # Column widths
    for i, w in enumerate([8, 14, 14, 14, 16, 16, 14, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_file)
    print(f"Template saved: {output_file}")


def _build_interpolation_formula(start_row, end_row, percentile):
    target = f"H{end_row}*{percentile}"
    parts = []
    for i in range(NUM_BINS):
        row = start_row + i
        prev_cum = f"H{row-1}" if i > 0 else "0"
        interp = f"B{row}+(({target}-{prev_cum})/IF(G{row}=0,1,G{row}))*(C{row}-B{row})"
        parts.append(f"IF(H{row}>={target},{interp},")
    return "=" + "".join(parts) + f"C{end_row}" + ")" * NUM_BINS


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def print_usage():
    print("Usage:")
    print("  python latency_calculator.py <results_dir>")
    print("  python latency_calculator.py <before.txt> <after.txt> [output.xlsx]")
    print("  python latency_calculator.py --template [output.xlsx]")
    print()
    print("Examples:")
    print("  python latency_calculator.py Results/Config_4_iPerf3_Linux_20260325_173114")
    print("  python latency_calculator.py snmp_before.txt snmp_after.txt report.xlsx")
    print("  python latency_calculator.py --template")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()
        sys.exit(1)

    if sys.argv[1] == "--template":
        out = sys.argv[2] if len(sys.argv) > 2 else "CMTS_Latency_Bin_Template.xlsx"
        create_template(out)
        sys.exit(0)

    arg1 = sys.argv[1]

    # Single arg = results directory (auto-discover files)
    if os.path.isdir(arg1):
        before_file, after_file = find_snmp_files(arg1)
        if not before_file or not after_file:
            print(f"ERROR: Could not find SNMP before/after .txt files in {arg1}")
            sys.exit(1)
        print(f"Before: {before_file}")
        print(f"After:  {after_file}")
        output = os.path.join(arg1, f"Latency_Bin_Report_{os.path.basename(os.path.abspath(arg1))}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        generate_latency_report(before_file, after_file, output)

    # Two args = explicit before/after files
    elif len(sys.argv) >= 3:
        before_file = arg1
        after_file = sys.argv[2]
        output = sys.argv[3] if len(sys.argv) > 3 else None
        generate_latency_report(before_file, after_file, output)

    else:
        print_usage()
        sys.exit(1)
