#!/home/aphillips/Projects/LLD_TEST_CLT_Dev_linux_compatible/venv/bin/python3

import argparse
import sys
import subprocess
import os
import glob
import re
import time
from datetime import datetime

import csv

try:
    import readline  # enables arrow-key editing in input() prompts
except ImportError:
    pass

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

from packetstorm_logic import PacketStormLogic
from byteblower_logic import ByteBlowerLogic
from iperf3_logic import IPerf3Logic
from speedtest_logic import SpeedTestLogic
from thousandeyes_logic import ThousandEyesLogic
from logger import Logger
from snmp_collector import collect_snmp_data, generate_latency_report, find_snmp_files, parse_modem_info
from cmts_modem_info import normalize_mac, collect_cmts_data

try:
    from kafka_collector import CmtsCollector
    CMTS_AVAILABLE = True
except ImportError:
    CMTS_AVAILABLE = False

try:
    from cm_data_collector import (
        snmp_collector_thread, kafka_collector_thread,
        _make_session_dir, _write_csv_comment,
        SNMP_CSV_FIELDS, KAFKA_CSV_FIELDS,
        DEFAULT_SNMP_POLL_INTERVAL, DEFAULT_KAFKA_BROKER, DEFAULT_KAFKA_TOPIC,
        DEFAULT_SNMP_JUMPSERVER, DEFAULT_SNMP_USERNAME,
        DEFAULT_MODEM_COMMUNITY, DEFAULT_ICMTS_COMMUNITY, DEFAULT_ICMTS_TARGET_IP,
        DEFAULT_SNMP_TIMEOUT, DEFAULT_SNMP_RETRIES,
    )
    CM_COLLECTOR_AVAILABLE = True
except ImportError:
    CM_COLLECTOR_AVAILABLE = False

class NetperfCLI:
    def __init__(self, cmts_type="vcmts"):
        self.logger = Logger("NetperfCLI")
        self.cmts_type = cmts_type
        self.logger.info(f"CMTS type: {cmts_type} ({'Kafka for DS latency' if cmts_type == 'vcmts' else 'SNMP for DS latency'})")

        # Prompt for CM MAC
        while True:
            mac_input = input("Enter CM MAC address (or press Enter to skip CMTS collection): ").strip()
            if not mac_input:
                self.cm_mac = None
                self.logger.warning("No CM MAC — CMTS collection will be skipped")
                break
            try:
                self.cm_mac = normalize_mac(mac_input)
                break
            except ValueError:
                print(f"Invalid MAC address: '{mac_input}'. Expected format: aabbccddeeff or aa:bb:cc:dd:ee:ff")

        # Auto-lookup IPv6 from CMTS using cmts_modem_info
        self.target_ip = None
        self._cmts_lookup_file = None
        if self.cm_mac:
            self.target_ip = self._lookup_ipv6_from_cmts(self.cm_mac, output_dir='Results/.cmts_lookup_tmp')
            # find the file written so we can move it later
            import glob as _g
            _files = _g.glob('Results/.cmts_lookup_tmp/*.json')
            self._cmts_lookup_file = _files[-1] if _files else None
            if not self.target_ip:
                answer = input("Cable modem not found on CMTS. Continue without SNMP? [y/N]: ").strip().lower()
                if answer != 'y':
                    print("Test stopped.")
                    sys.exit(1)
        if self.target_ip:
            print(f"Modem IPv6: {self.target_ip}")
        else:
            self.logger.warning("No modem IPv6 — upstream SNMP collection will be skipped")

        self.output_dir   = None
        self.cmts_collector = None
        # cm_collector continuous polling state
        self._cm_stop_event      = None
        self._cm_threads         = []
        self._cm_session_dir     = None
        self._cm_csv_paths       = {}
        self._cm_kafka_csv       = None
        self._cm_session_dir_ref = None
        self._cm_scenario_ref    = None
        self._cm_kafka_ready     = None
    
    def _lookup_ipv6_from_cmts(self, mac, output_dir=None):
        """Use cmts_modem_info.collect_cmts_data to SSH to CMTS and resolve modem IPv6."""
        try:
            from config_loader import config
            hosts = config.get('cmts_hosts', default=[]) or []
            matched = next((h for h in hosts if h.get('type', 'vcmts') == self.cmts_type), None)
            if not matched:
                matched = hosts[0] if hosts else {}
            cmts_host = matched.get('name', '')
            if not cmts_host:
                cmts_host = 'apc01k1dccc' if self.cmts_type == 'vcmts' else 'cts01k1dccc'
            self.logger.info(f"Looking up IPv6 for {mac} via {cmts_host} ({self.cmts_type})...")
            ipv6 = collect_cmts_data(cmts_host, mac, self.cmts_type,
                                     output_dir=output_dir or 'Results')
            if ipv6:
                print(f"Modem IPv6 (auto-detected): {ipv6}")
            else:
                self.logger.warning("cmts_modem_info ran but no IPv6 found")
            return ipv6
        except Exception as e:
            self.logger.warning(f"CMTS IPv6 lookup failed: {e}")
        return None

    def start_cm_collector(self, test_name, output_dir=None):
        """Prepare cm_collector config. Threads start on first set_cm_collector_dir call."""
        if not CM_COLLECTOR_AVAILABLE:
            self.logger.warning("cm_collector not available — skipping continuous polling")
            return
        if not self.cm_mac or not self.target_ip:
            return
        if self._cm_stop_event and not self._cm_stop_event.is_set():
            self.logger.debug("cm_collector already running — skipping duplicate start")
            return
        mac_norm  = self.cm_mac.replace('.', '').replace(':', '').lower()
        mac_colon = ':'.join(mac_norm[i:i+2] for i in range(0, 12, 2))
        self._cm_session_dir     = output_dir if output_dir else _make_session_dir(mac_norm, self.cmts_type, datetime.now().strftime('%Y%m%d_%H%M%S'))
        self._cm_mac_norm        = mac_norm
        self._cm_mac_colon       = mac_colon
        self._cm_session_dir_ref = [self._cm_session_dir]
        self._cm_scenario_ref    = ['']
        self._cm_stop_event      = __import__('threading').Event()
        self._cm_kafka_ready     = __import__('threading').Event()
        self._cm_poll_ref        = [1]
        self._cm_cfg_base = {
            'mac_norm': mac_norm, 'mac_colon': mac_colon,
            'cmts_type': self.cmts_type,
            'target_ip': self.target_ip,
            'session_dir': self._cm_session_dir,
            'session_dir_ref': self._cm_session_dir_ref,
            'scenario_ref':    self._cm_scenario_ref,
            'kafka_ready_event': self._cm_kafka_ready,
            'modem_community': DEFAULT_MODEM_COMMUNITY,
            'icmts_community': DEFAULT_ICMTS_COMMUNITY,
            'icmts_target': DEFAULT_ICMTS_TARGET_IP if self.cmts_type == 'icmts' else '',
            'kafka_broker': DEFAULT_KAFKA_BROKER,
            'kafka_topic':  DEFAULT_KAFKA_TOPIC,
            'snmp_jumpserver':   DEFAULT_SNMP_JUMPSERVER,
            'snmp_username':     DEFAULT_SNMP_USERNAME,
            'snmp_timeout':      DEFAULT_SNMP_TIMEOUT,
            'snmp_retries':      DEFAULT_SNMP_RETRIES,
            'snmp_poll_interval': DEFAULT_SNMP_POLL_INTERVAL,
        }
        # threads are started by set_cm_collector_dir once the scenario dir is known

    def set_cm_collector_dir(self, scenario_dir, scenario_name):
        """Point all cm_collector output into scenario_dir.
        First call starts the threads; subsequent calls just redirect poll .txt output."""
        if self._cm_session_dir_ref is None:
            return
        os.makedirs(scenario_dir, exist_ok=True)
        self._cm_session_dir_ref[0] = scenario_dir
        self._cm_scenario_ref[0]    = scenario_name

        if self._cm_threads:
            # threads already running — poll .txt redirect is enough
            return

        # First call — build CSV paths directly inside scenario_dir
        mac_norm = self._cm_mac_norm
        dir_ts_m = re.search(r'(\d{8}_\d{6})', os.path.basename(self._cm_session_dir))
        ts_str   = dir_ts_m.group(1) if dir_ts_m else datetime.now().strftime('%Y%m%d_%H%M%S')

        self._cm_csv_paths = {'us': os.path.join(scenario_dir, f'snmp_us_{mac_norm}_{ts_str}.csv')}
        if self.cmts_type == 'icmts':
            self._cm_csv_paths['ds'] = os.path.join(scenario_dir, f'snmp_ds_{mac_norm}_{ts_str}.csv')
        self._cm_kafka_csv = os.path.join(scenario_dir, f'kafka_{mac_norm}_{ts_str}.csv') \
                             if self.cmts_type == 'vcmts' else None

        import threading
        cfg = self._cm_cfg_base
        snmp_t = threading.Thread(
            target=snmp_collector_thread,
            args=(cfg, self._cm_stop_event, self._cm_csv_paths, self._cm_poll_ref),
            daemon=True,
        )
        self._cm_threads.append(snmp_t)
        if self.cmts_type == 'vcmts' and self._cm_kafka_csv:
            kafka_t = threading.Thread(
                target=kafka_collector_thread,
                args=(cfg, self._cm_stop_event, self._cm_kafka_csv),
                daemon=True,
            )
            self._cm_threads.append(kafka_t)
        for t in self._cm_threads:
            t.start()
        self.logger.info(f"cm_collector started → {scenario_dir}")

    def stop_cm_collector(self):
        """Stop cm_collector threads, write Excel timeseries, return session dir."""
        if not self._cm_stop_event:
            return None
        self._cm_stop_event.set()
        for t in self._cm_threads:
            t.join(timeout=15)
        self._cm_threads = []
        self._cm_stop_event = None
        session_dir      = self._cm_session_dir_ref[0] if self._cm_session_dir_ref else self._cm_session_dir
        cm_csv_paths     = self._cm_csv_paths
        cm_kafka_csv     = self._cm_kafka_csv
        self._cm_session_dir     = None
        self._cm_csv_paths       = {}
        self._cm_kafka_csv       = None
        self._cm_session_dir_ref = None
        self._cm_scenario_ref    = None
        self.logger.info(f"cm_collector stopped — session: {session_dir}")
        return session_dir

    def generate_pdf_report(self, session_dir, session_name):
        """Generate PDF report from cm_collector session CSVs."""
        if not session_dir or not os.path.isdir(session_dir):
            return
        try:
            import subprocess
            script = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'metrics_pdf_report.py')
            if not os.path.exists(script):
                self.logger.warning("metrics_pdf_report.py not found — skipping PDF")
                return
            # Build a meaningful name: "TestGroup — TrafficType_Scenario"
            # e.g. "HSI021_Thousandeyes — ThousandEyes"
            #      "HSI029_RTT_0 — ByteBlower_DS_Classic"
            traffic_type = getattr(self, 'traffic_type', '')
            if traffic_type and not session_name.startswith(traffic_type):
                pdf_name = f"{session_name} — {traffic_type}"
            else:
                pdf_name = session_name
            result = subprocess.run(
                [sys.executable, script, session_dir, '--name', pdf_name],
                capture_output=True, text=True, timeout=120,
            )
            if result.returncode == 0:
                for line in result.stdout.splitlines():
                    if 'PDF saved' in line:
                        self.logger.info(line.strip())
            else:
                self.logger.warning(f"PDF report failed:\n{result.stderr.strip()}\n{result.stdout.strip()}")
        except Exception as e:
            self.logger.warning(f"PDF report generation failed: {e}")

    def start_cmts_collection(self, direction="downstream"):
        """Start CMTS Kafka metrics collection in background.
        Only used for vCMTS downstream. For iCMTS downstream, SNMP handles it."""
        if self.cmts_type == "icmts" and direction == "downstream":
            self.logger.info("iCMTS mode — DS latency collected via SNMP (skipping Kafka)")
            return False
        if not self.cm_mac:
            return False
        if not CMTS_AVAILABLE:
            self.logger.warning("CMTS collector not available (kafka-python not installed) — skipping")
            return False
        try:
            self.cmts_collector = CmtsCollector(mac=self.cm_mac, direction=direction, cm_ipv6=self.target_ip)
            self.cmts_collector.start()
            return True
        except Exception as e:
            self.logger.warning(f"CMTS collection unavailable: {e} — test will continue without it")
            self.cmts_collector = None
            return False

    def wait_for_cmts_poll(self, timeout=None):
        """Wait for next Kafka polling interval. Safe to call if collector not started.

        Uses detected polling interval + 20% headroom. Falls back to 150s on first
        call before any interval is known (covers up to a 120s polling environment).
        """
        if not self.cmts_collector:
            return False
        try:
            if timeout is None:
                interval = self.cmts_collector._get_polling_interval_s()
                # If interval is still the 60s default (not yet detected), use 150s
                # to safely cover environments with up to 120s polling
                timeout = interval * 1.2 if self.cmts_collector._detected_interval_s else 150
            return self.cmts_collector.wait_for_poll(timeout=timeout)
        except Exception as e:
            self.logger.warning(f"wait_for_cmts_poll failed: {e}")
            return False

    def stop_cmts_collection(self, output_dir, test_name, post_test_polls=2, test_duration_s=60):
        """Stop CMTS Kafka collection and generate report.

        Detects the actual Kafka polling interval and waits for
        post_test_polls additional poll cycles after the test ends
        to capture delayed bin reporting data.

        test_duration_s: actual test duration for throughput calculation.
        """
        if not self.cmts_collector:
            return None
        try:
            poll_interval_s = self.cmts_collector._get_polling_interval_s()
            poll_timeout = poll_interval_s * 1.2

            for i in range(post_test_polls):
                if not self.cmts_collector.wait_for_poll(timeout=poll_timeout):
                    self.logger.warning(f"Post-test poll {i+1} timed out")
                    break
            self.cmts_collector.stop()
            report = self.cmts_collector.generate_report(output_dir, test_name, test_duration_s=test_duration_s)
            self.cmts_collector = None
            if report:
                self.logger.info(f"CMTS latency report: {os.path.basename(report)}")
            return report
        except Exception as e:
            self.logger.warning(f"CMTS report generation failed: {e}")
            self.cmts_collector = None
            return None

    def run_snmp_collection(self, test_name, phase, output_dir=None, rtt_suffix=""):
        if not self.target_ip:
            return False
        try:
            if output_dir is None:
                output_dir = "Results"
            collect_snmp_data(self.target_ip, f"{test_name}{rtt_suffix}", phase, output_dir)
            return True
        except Exception as e:
            self.logger.error(f"SNMP collection failed: {e}")
            return False
    
    def execute(self, bbp_file, rtt_files, iterations, scenarios, test_group_name=None, client_ip=None, output_format="json", byteblower_only=False, packetstorm_only=False, iperf3_only=False, iperf3_darwin=False, speedtest_only=False, speedtest_clients=None, thousandeyes_only=False, thousandeyes_unit_id=None, report_formats="html pdf csv xls xlsx json docx", granular=False):
        """Execute workflow based on selected modes"""
        try:
            scenario_list = [s.strip() for s in scenarios.split(',')] if scenarios else ['default']
            rtt_list = [r.strip() for r in rtt_files.split(',')] if rtt_files else ['default.json']
            multi_scenario = len(scenario_list) > 1

            # Determine traffic type
            if thousandeyes_only:
                traffic_type = "ThousandEyes"
            elif byteblower_only or (not iperf3_only and not iperf3_darwin and not speedtest_only and not thousandeyes_only):
                traffic_type = "ByteBlower"
            elif iperf3_darwin:
                traffic_type = "iPerf3_macOS"
            elif iperf3_only:
                traffic_type = "iPerf3_Linux"
            elif speedtest_only:
                traffic_type = "SpeedTest"
            else:
                traffic_type = "PacketStorm"

            # Extract RTT suffix from first RTT file
            rtt_suffix = ""
            if rtt_list[0] and rtt_list[0] != 'default.json':
                rtt_match = re.search(r'(\d+)ms', rtt_list[0])
                if rtt_match:
                    rtt_suffix = f"_RTT_{rtt_match.group(1)}ms"

            # Create parent output directory
            parent_output_dir = f"Results/{test_group_name}_{traffic_type}{rtt_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            os.makedirs(parent_output_dir, exist_ok=True)

            # Move modem summary into this test's results directory
            if self._cmts_lookup_file and os.path.exists(self._cmts_lookup_file):
                import shutil
                shutil.move(self._cmts_lookup_file, os.path.join(parent_output_dir, 'modem_summary.json'))
                try:
                    os.rmdir('Results/.cmts_lookup_tmp')
                except OSError:
                    pass
                self._cmts_lookup_file = None

            self.traffic_type = traffic_type
            self.rtt_suffix = rtt_suffix

            if multi_scenario or len(rtt_list) > 1:
                self.logger.info(f"Multi-scenario run: {len(scenario_list)} scenarios x {len(rtt_list)} RTT configs → {parent_output_dir}")

            all_success = True
            all_snmp_files = []

            # --- Multi-scenario: single polling session spans all scenarios (default) ---
            # --- --granular: poll each scenario independently ---
            if multi_scenario and not granular:
                combined_name = '_'.join(scenario_list)
                # Start polling once into parent dir; first set_cm_collector_dir call starts threads
                self.start_cm_collector(combined_name, output_dir=parent_output_dir)
                self.set_cm_collector_dir(parent_output_dir, combined_name)
                self.logger.info(f"Pre-test baseline collection — 30s")
                time.sleep(30)

                for rtt_file in rtt_list:
                    for scenario in scenario_list:
                        success, snmp_files = self._run_single_test(
                            bbp_file, rtt_file, iterations, scenario,
                            test_group_name, client_ip, output_format,
                            byteblower_only, packetstorm_only, iperf3_only,
                            iperf3_darwin, speedtest_only, speedtest_clients,
                            thousandeyes_only, thousandeyes_unit_id, report_formats,
                            parent_output_dir, rtt_list,
                            polling_managed=True,  # skip per-scenario start/stop/baseline
                        )
                        all_success = all_success and success
                        all_snmp_files.extend(snmp_files)

                self.logger.info("Post-test tail collection — 30s")
                time.sleep(30)
                cm_session = self.stop_cm_collector()
                self.generate_pdf_report(cm_session, combined_name)
            elif multi_scenario and granular:
                # Each scenario gets its own polling session
                for rtt_file in rtt_list:
                    for scenario in scenario_list:
                        success, snmp_files = self._run_single_test(
                            bbp_file, rtt_file, iterations, scenario,
                            test_group_name, client_ip, output_format,
                            byteblower_only, packetstorm_only, iperf3_only,
                            iperf3_darwin, speedtest_only, speedtest_clients,
                            thousandeyes_only, thousandeyes_unit_id, report_formats,
                            parent_output_dir, rtt_list,
                        )
                        all_success = all_success and success
                        all_snmp_files.extend(snmp_files)
            else:
                # Single scenario — polling runs per scenario by default
                for scenario in scenario_list:
                    for rtt_file in rtt_list:
                        success, snmp_files = self._run_single_test(
                            bbp_file, rtt_file, iterations, scenario,
                            test_group_name, client_ip, output_format,
                            byteblower_only, packetstorm_only, iperf3_only,
                            iperf3_darwin, speedtest_only, speedtest_clients,
                            thousandeyes_only, thousandeyes_unit_id, report_formats,
                            parent_output_dir, rtt_list,
                        )
                        all_success = all_success and success
                        all_snmp_files.extend(snmp_files)

            self._zip_results(parent_output_dir)

            if all_success:
                self.logger.info("All workflows completed successfully")
                return 0
            else:
                self.logger.error("Some workflows completed with errors")
                return 1

        except Exception as e:
            self.logger.error(f"Workflow failed: {e}")
            return 1
    
    def _run_single_test(self, bbp_file, rtt_file, iterations, scenario_name, test_group_name, client_ip, output_format, byteblower_only, packetstorm_only, iperf3_only, iperf3_darwin, speedtest_only, speedtest_clients, thousandeyes_only, thousandeyes_unit_id, report_formats, parent_output_dir, rtt_list, polling_managed=False):
        """Run a single test scenario.
        polling_managed=True: caller owns start/stop/baseline/tail — skip them here.
        """
        try:
            success = True
            snmp_files = []
            
            # Extract RTT value from filename for naming
            rtt_suffix = ""
            rtt_dir_suffix = ""
            if rtt_file and rtt_file != 'default.json':
                import re
                rtt_match = re.search(r'(\d+)ms', rtt_file)
                if rtt_match:
                    rtt_suffix = f"_RTT_{rtt_match.group(1)}ms"
                    rtt_dir_suffix = f"RTT_{rtt_match.group(1)}ms"
            
            # Create subdirectory only if multiple RTT values
            test_output_dir = parent_output_dir
            if len(rtt_list) > 1 and rtt_dir_suffix:
                traffic_subdir = f"{self.traffic_type}_{rtt_dir_suffix}"
                test_output_dir = os.path.join(parent_output_dir, traffic_subdir)
                os.makedirs(test_output_dir, exist_ok=True)
            
            # Determine test name for SNMP collection
            test_name = scenario_name
            if byteblower_only:
                test_name = f"ByteBlower_{scenario_name}"
            elif iperf3_only:
                test_name = f"iPerf3_Linux_{scenario_name}"
            elif iperf3_darwin:
                test_name = f"iPerf3_macOS_{scenario_name}"
            elif packetstorm_only:
                test_name = f"PacketStorm_{scenario_name}"
            else:
                if client_ip:
                    test_name = f"iPerf3_Linux_{scenario_name}"
                else:
                    test_name = f"ByteBlower_{scenario_name}"
            
            if thousandeyes_only:
                self.logger.info(f"ThousandEyes mode - all tests (http_get_mt, http_post_mt, udp_jitter)")
                sk = ThousandEyesLogic(scenario_name, test_group_name, rtt_suffix, thousandeyes_unit_id)
                self.output_dir = test_output_dir
                success = True
                snmp_dir = test_output_dir
                if not polling_managed:
                    self.start_cm_collector(test_name, output_dir=parent_output_dir)
                    self.set_cm_collector_dir(snmp_dir, 'ThousandEyes')
                    # Wait for Kafka first poll before starting tests
                    if self._cm_kafka_ready is not None:
                        self.logger.info("Waiting for first Kafka poll...")
                        if not self._cm_kafka_ready.wait(timeout=180):
                            self.logger.warning("Kafka ready timeout — proceeding anyway")
                    else:
                        self.logger.info("Pre-test baseline collection — 30s")
                        time.sleep(30)
                for i in range(iterations):
                    self.logger.info(f"Starting test: ThousandEyes iteration {i+1}/{iterations}")
                    ds_ok, ds_elapsed = sk.run_downstream(i, iterations, test_output_dir)
                    if not ds_ok:
                        self.logger.error("ThousandEyes downstream failed — stopping test")
                        if not polling_managed:
                            self.logger.info("Post-test collection — 30s")
                            time.sleep(30)
                            cm_session = self.stop_cm_collector()
                            self.generate_pdf_report(cm_session, test_name)
                        sys.exit(1)
                    us_ok, us_elapsed = sk.run_upstream(i, iterations, test_output_dir)
                    if not us_ok:
                        self.logger.error("ThousandEyes upstream failed — stopping test")
                        if not polling_managed:
                            self.logger.info("Post-test collection — 30s")
                            time.sleep(30)
                            cm_session = self.stop_cm_collector()
                            self.generate_pdf_report(cm_session, test_name)
                        sys.exit(1)
                    sk.run_jitter(i, iterations, test_output_dir)
                if not polling_managed:
                    self.logger.info("Post-test tail collection — 30s")
                    time.sleep(30)
                    cm_session = self.stop_cm_collector()
                    self.generate_pdf_report(cm_session, test_name)
            elif speedtest_only:
                self.logger.info(f"SpeedTest mode - clients: {speedtest_clients}")
                st = SpeedTestLogic(speedtest_clients, test_group_name)
                self.output_dir = test_output_dir
                snmp_dir = test_output_dir
                if not polling_managed:
                    self.start_cm_collector(test_name, output_dir=parent_output_dir)
                    self.set_cm_collector_dir(snmp_dir, 'SpeedTest')
                    self.logger.info("Pre-test baseline collection — 30s")
                    time.sleep(30)
                self.logger.info(f"Starting test: SpeedTest")
                if not st.run_iterations(iterations, output_dir=snmp_dir):
                    self.logger.error("SpeedTest failed — stopping test")
                    if not polling_managed:
                        self.logger.info("Post-test collection — 30s")
                        time.sleep(30)
                        cm_session = self.stop_cm_collector()
                        self.generate_pdf_report(cm_session, test_name)
                    return False, []
                success = True
                if not polling_managed:
                    self.logger.info("Post-test tail collection — 30s")
                    time.sleep(30)
                    cm_session = self.stop_cm_collector()
                    self.generate_pdf_report(cm_session, test_name)
            elif packetstorm_only:
                self.logger.info(f"PacketStorm only mode - config: {rtt_file}")
                ps = PacketStormLogic(rtt_file)
                success = ps.start_config() and ps.stop_config()
            elif byteblower_only:
                bb = ByteBlowerLogic(bbp_file, scenario_name, scenario_name, test_group_name, rtt_suffix, report_formats, cm_mac=self.cm_mac, cm_ipv6=self.target_ip)
                self.output_dir = test_output_dir
                success = True
                snmp_dir = test_output_dir
                cmts_dir = "upstream" if scenario_name.lower().startswith("us") else "downstream"
                if not polling_managed:
                    self.start_cm_collector(test_name, output_dir=parent_output_dir)
                    self.set_cm_collector_dir(snmp_dir, scenario_name)
                for i in range(iterations):
                    iter_dir = snmp_dir
                    if i == 0:
                        snmp_files = glob.glob(os.path.join(iter_dir, "*_SNMP_before_*.txt"))
                        if snmp_files:
                            info = parse_modem_info(snmp_files[-1])
                            if info:
                                print(f"\n  Modem Info:")
                                print(f"    Model:   {info.get('model', 'N/A')}")
                                print(f"    Vendor:  {info.get('vendor', 'N/A')}")
                                print(f"    SW Rev:  {info.get('sw_rev', 'N/A')}")
                                print(f"    HW Rev:  {info.get('hw_rev', 'N/A')}")
                                if self.cm_mac:
                                    print(f"    CM MAC:  {self.cm_mac}")
                                if self.target_ip:
                                    print(f"    CM IPv6: {self.target_ip}")
                    if not polling_managed:
                        self.logger.info("Pre-test baseline collection — 30s")
                        time.sleep(30)
                    self.logger.info(f"Starting test: {scenario_name} iteration {i+1}/{iterations}")
                    test_start = time.time()
                    bb_ok, bb_duration = bb.run_scenario(i, iterations, test_output_dir)
                    if not bb_ok:
                        self.logger.error("ByteBlower failed — stopping test")
                        if not polling_managed:
                            self.logger.info("Post-test collection — 30s")
                            time.sleep(30)
                            cm_session = self.stop_cm_collector()
                            self.generate_pdf_report(cm_session, test_name)
                        return False, []
                    test_elapsed = bb_duration or (time.time() - test_start)
                    if not polling_managed:
                        self.logger.info("Post-test tail collection — 30s")
                        time.sleep(30)
                if not polling_managed:
                    cm_session = self.stop_cm_collector()
                    self.generate_pdf_report(cm_session, test_name)
            elif iperf3_only or iperf3_darwin:
                platform_override = 'macos' if iperf3_darwin else None
                platform_suffix = "_macOS" if iperf3_darwin else "_Linux"
                self.logger.info(f"iPerf3 only mode - client: {client_ip}, scenario: {scenario_name}")
                iperf3 = IPerf3Logic(client_ip, scenario_name, test_group_name, rtt_suffix, output_format, platform_override, test_output_dir)
                self.output_dir = test_output_dir
                if not iperf3.setup_ssh_keys():
                    self.logger.error("SSH key setup failed")
                    return False, []
                if not iperf3.setup_iperf3_servers():
                    self.logger.error("iPerf3 server setup failed")
                    return False, []
                success = True
                snmp_dir = test_output_dir
                cmts_dir = "upstream" if scenario_name.lower().startswith("us") else "downstream"
                if not polling_managed:
                    self.start_cm_collector(test_name, output_dir=parent_output_dir)
                    self.set_cm_collector_dir(snmp_dir, scenario_name)
                for i in range(iterations):
                    if not polling_managed:
                        self.logger.info("Pre-test baseline collection — 30s")
                        time.sleep(30)
                    self.logger.info(f"Starting test: {scenario_name} iteration {i+1}/{iterations}")
                    test_start = time.time()
                    ip3_ok, ip3_duration = iperf3.run_scenario(i, iterations)
                    if not ip3_ok:
                        self.logger.error("iPerf3 failed — stopping test")
                        iperf3.stop_iperf3_servers()
                        if not polling_managed:
                            self.logger.info("Post-test collection — 30s")
                            time.sleep(30)
                            cm_session = self.stop_cm_collector()
                            self.generate_pdf_report(cm_session, test_name)
                        return False, []
                    test_elapsed = ip3_duration or (time.time() - test_start)
                    if not polling_managed:
                        self.logger.info("Post-test tail collection — 30s")
                        time.sleep(30)
                iperf3.stop_iperf3_servers()
                if not polling_managed:
                    cm_session = self.stop_cm_collector()
                    self.generate_pdf_report(cm_session, test_name)
            else:
                ps = PacketStormLogic(rtt_file)
                
                if client_ip:
                    iperf3 = IPerf3Logic(client_ip, scenario_name, test_group_name, rtt_suffix, output_format, None, test_output_dir)
                    self.output_dir = test_output_dir
                    
                    success = ps.start_config()
                    if success:
                        # Setup SSH and servers once
                        if not iperf3.setup_ssh_keys():
                            self.logger.error("SSH key setup failed")
                            ps.stop_config()
                            return False, []
                        if not iperf3.setup_iperf3_servers():
                            self.logger.error("iPerf3 server setup failed")
                            ps.stop_config()
                            return False, []
                        
                        snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
                        os.makedirs(snmp_dir, exist_ok=True)
                        
                        cmts_dir = "upstream" if scenario_name.lower().startswith("us") else "downstream"
                        
                        for i in range(iterations):
                            test_start = time.time()
                            ip3_ok, ip3_duration = iperf3.run_scenario(i, iterations)
                            if not ip3_ok:
                                self.logger.error("iPerf3 failed — stopping test")
                                iperf3.stop_iperf3_servers()
                                ps.stop_config()
                                return False, []
                            test_elapsed = ip3_duration or (time.time() - test_start)
                        iperf3.stop_iperf3_servers()
                        success = success and ps.stop_config()
                else:
                    bb = ByteBlowerLogic(bbp_file, scenario_name, scenario_name, test_group_name, rtt_suffix, report_formats)
                    self.output_dir = test_output_dir
                    
                    success = ps.start_config()
                    if success:
                        snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
                        os.makedirs(snmp_dir, exist_ok=True)
                        
                        for i in range(iterations):
                            test_start = time.time()
                            bb_ok, bb_duration = bb.run_scenario(i, iterations, test_output_dir)
                            if not bb_ok:
                                self.logger.error("ByteBlower failed — stopping test")
                                ps.stop_config()
                                return False, []
                            test_elapsed = bb_duration or (time.time() - test_start)
                        success = success and ps.stop_config()
            
            # Collect SNMP files from scenario directory
            if iperf3_only or iperf3_darwin:
                platform_suffix = "_macOS" if iperf3_darwin else "_Linux"
                scenario_snmp_dir = os.path.join(test_output_dir, scenario_name + platform_suffix + rtt_suffix)
            else:
                scenario_snmp_dir = os.path.join(test_output_dir, scenario_name + rtt_suffix)
            scenario_snmp_files = glob.glob(os.path.join(scenario_snmp_dir, "*_SNMP_*.txt"))
            
            snmp_files = scenario_snmp_files if scenario_snmp_files else glob.glob(os.path.join(test_output_dir, "*_SNMP_*.txt"))
            
            return success, snmp_files
                
        except Exception as e:
            self.logger.error(f"Single test failed: {e}")
            return False, []
    
    def _run_latency_report(self, snmp_dir, iteration=None, prefix=None, duration_s=None):
        """Generate latency bin report from SNMP before/after files"""
        try:
            if iteration is not None:
                # Find before/after files for this specific iteration
                if prefix:
                    before_files = sorted(glob.glob(os.path.join(snmp_dir, f"*{prefix}_iteration_{iteration}_SNMP_before_*.txt")))
                    after_files = sorted(glob.glob(os.path.join(snmp_dir, f"*{prefix}_iteration_{iteration}_SNMP_after_*.txt")))
                else:
                    before_files = sorted(glob.glob(os.path.join(snmp_dir, f"*_iteration_{iteration}_SNMP_before_*.txt")))
                    after_files = sorted(glob.glob(os.path.join(snmp_dir, f"*_iteration_{iteration}_SNMP_after_*.txt")))
                before_file = before_files[-1] if before_files else None
                after_file = after_files[-1] if after_files else None
            else:
                before_file, after_file = find_snmp_files(snmp_dir)
            if before_file and after_file:
                # Build output filename with iteration
                output_dir = os.path.dirname(after_file) or "."
                dir_name = os.path.basename(os.path.abspath(output_dir))
                iter_tag = f"_iteration_{iteration}" if iteration else ""
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                # US report (from modem) — always generated
                us_output = os.path.join(output_dir, f"SNMP_US_Latency_Report_{dir_name}{iter_tag}_{timestamp}.xlsx")
                result = generate_latency_report(before_file, after_file, output_file=us_output, direction="US", cm_mac=self.cm_mac, cm_ipv6=self.target_ip, duration_s=duration_s)
                if result:
                    self.logger.info(f"US Latency report: {os.path.basename(result)}")
                else:
                    self.logger.info("US Latency report skipped (no US latency OIDs)")
                # DS report (from iCMTS SNMP) — only for iCMTS mode
                if self.cmts_type == "icmts":
                    ds_output = os.path.join(output_dir, f"SNMP_DS_Latency_Report_{dir_name}{iter_tag}_{timestamp}.xlsx")
                    result_ds = generate_latency_report(before_file, after_file, output_file=ds_output, direction="DS", cm_mac=self.cm_mac, cm_ipv6=self.target_ip, duration_s=duration_s)
                    if result_ds:
                        self.logger.info(f"DS Latency report (iCMTS SNMP): {os.path.basename(result_ds)}")
                    else:
                        self.logger.info("DS Latency report skipped (no DS latency OIDs found)")
                else:
                    self.logger.info("DS latency via Kafka (vCMTS) — see CMTS_Latency_Report_*.xlsx")
            else:
                self.logger.warning("Latency report skipped (SNMP files not found)")
        except Exception as e:
            self.logger.error(f"Latency report failed: {e}")

    def _zip_results(self, folder):
        """Compress results folder to a zip file alongside it."""
        import zipfile
        zip_path = folder.rstrip('/') + '.zip'
        try:
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
                for root, _, files in os.walk(folder):
                    for file in files:
                        abs_path = os.path.join(root, file)
                        zf.write(abs_path, os.path.relpath(abs_path, os.path.dirname(folder)))
            self.logger.info(f"Results compressed: {os.path.basename(zip_path)}")
        except Exception as e:
            self.logger.warning(f"Zip failed: {e}")

    def _consolidate_snmp_to_excel(self, snmp_files, parent_dir, test_group_name):
        """Consolidate all SNMP files into Excel with separate sheets"""
        try:
            if not snmp_files:
                return
            
            # Create Excel with separate sheets if pandas available
            if HAS_PANDAS:
                try:
                    import openpyxl
                    wb = openpyxl.Workbook()
                    wb.remove(wb.active)
                    
                    for snmp_file in sorted(set(snmp_files)):
                        sheet_name = os.path.basename(snmp_file).replace('.txt', '').replace('_SNMP_', '_')[:31]
                        ws = wb.create_sheet(sheet_name)
                        with open(snmp_file, 'r') as f:
                            for r_idx, line in enumerate(f, 1):
                                ws.cell(row=r_idx, column=1, value=line.rstrip())
                    
                    excel_file = os.path.join(parent_dir, f"{test_group_name}_SNMP_Combined_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    wb.save(excel_file)
                    self.logger.info(f"Combined SNMP Excel: {os.path.basename(excel_file)}")
                except Exception as e:
                    self.logger.warning(f"Excel creation skipped: {e}")
        except Exception as e:
            self.logger.error(f"Error consolidating SNMP: {e}")

def main():
    parser = argparse.ArgumentParser(description='ByteBlower, PacketStorm, iPerf3, and SpeedTest CLI Tool')
    # Required: CMTS type must be specified first
    parser.add_argument('--cmts-type', choices=['vcmts', 'icmts'], required=True, help='CMTS type: vcmts (Kafka for DS latency) or icmts (SNMP for DS latency)')
    # Traffic generation mode
    parser.add_argument('-byteblower', action='store_true', help='Enable ByteBlower mode')
    parser.add_argument('--bbp', required=False, help='ByteBlower .bbp file path (e.g., bb_flows/US_Classic.bbp)')
    parser.add_argument('--scenario', required=False, help='Scenario name (e.g., US_Classic_Only)')
    parser.add_argument('-test-group-name', '--test-group-name', help='Test group name (e.g., HSI029_RTT_0)')
    parser.add_argument('-packetstorm', action='store_true', help='Enable PacketStorm mode')
    parser.add_argument('--rtt', help='RTT configuration file (e.g., vcmts10ms.json)')
    parser.add_argument('-iperf3', action='store_true', help='Enable iPerf3 Linux mode')
    parser.add_argument('-iperf3-darwin', action='store_true', help='Enable iPerf3 macOS mode with iperf3-darwin')
    parser.add_argument('--clientIP', help='Linux client IP address for iPerf3 (required with -iperf3)')
    parser.add_argument('--output', choices=['json', 'txt'], default='json', help='iPerf3 output format: json or txt (default: json)')
    parser.add_argument('-speedtest', action='store_true', help='Enable SpeedTest mode')
    parser.add_argument('--client', default='linux,macos,nvidia', help='SpeedTest clients: linux,macos,nvidia (default: all)')
    parser.add_argument('-thousandeyes', action='store_true', help='Enable ThousandEyes instant-test mode')
    parser.add_argument('--unit-id', default=None, help='ThousandEyes unit ID (overrides config.yaml)')
    parser.add_argument('--report-formats', default='html pdf csv xls xlsx json docx', help='ByteBlower report formats (default: all formats)')
    parser.add_argument('-iteration', type=int, default=1, help='Number of iterations (default: 1)')
    parser.add_argument('--granular', action='store_true', help='Multi-scenario: poll each scenario independently instead of one shared session')
    
    args = parser.parse_args()
    
    if not args.byteblower and not args.packetstorm and not args.iperf3 and not getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False):
        parser.error("At least one of -byteblower, -packetstorm, -iperf3, -iperf3-darwin, -speedtest, or -thousandeyes is required")
    
    if (args.iperf3 or getattr(args, 'iperf3_darwin', False)) and (not args.scenario or not args.clientIP):
        parser.error("Both --scenario and --clientIP are required when using -iperf3 or -iperf3-darwin")
    
    if getattr(args, 'thousandeyes', False) and not getattr(args, 'unit_id', None):
        parser.error("--unit-id is required when using -thousandeyes (e.g., --unit-id 82670821)")

    if args.packetstorm and not args.rtt:
        parser.error("--rtt is required when using -packetstorm")
    
    bb_file = args.bbp or 'default.bbp'
    speedtest_clients = [c.strip() for c in args.client.split(',')] if getattr(args, 'speedtest', False) else None
    
    tool = NetperfCLI(cmts_type=args.cmts_type)
    return tool.execute(
        bb_file, 
        args.rtt or 'default.json', 
        args.iteration,
        args.scenario or 'default',
        getattr(args, 'test_group_name', None),
        getattr(args, 'clientIP', None),
        getattr(args, 'output', 'json'),
        byteblower_only=args.byteblower and not args.packetstorm and not args.iperf3 and not getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False),
        packetstorm_only=args.packetstorm and not args.byteblower and not args.iperf3 and not getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False),
        iperf3_only=args.iperf3 and not args.byteblower and not args.packetstorm and not getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False),
        iperf3_darwin=getattr(args, 'iperf3_darwin', False) and not getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False),
        speedtest_only=getattr(args, 'speedtest', False) and not getattr(args, 'thousandeyes', False),
        speedtest_clients=speedtest_clients,
        thousandeyes_only=getattr(args, 'thousandeyes', False),
        thousandeyes_unit_id=getattr(args, 'unit_id', None),
        report_formats=getattr(args, 'report_formats', 'html pdf csv xls xlsx json docx'),
        granular=args.granular,
    )

if __name__ == '__main__':
    sys.exit(main())
