"""
CMTS Downstream Latency Bin Calculator.

DEPRECATED: All functionality has been moved to snmp_collector.py.
This file is kept for backward compatibility and CLI usage.

Usage:
  python latency_calculator.py <results_dir>
  python latency_calculator.py <before.txt> <after.txt> [output.xlsx]
  python latency_calculator.py --template [output.xlsx]
"""
import os
import sys
from datetime import datetime

# Re-export everything from snmp_collector for backward compatibility
from snmp_collector import (
    BIN_EDGES_MS,
    NUM_BINS,
    parse_snmp_timestamp,
    parse_flow_stats,
    compute_throughput_and_loss,
    parse_latency_bins,
    compute_deltas,
    calc_percentile,
    calc_weighted_avg,
    calc_percentile_avg,
    write_sf_sheet,
    write_summary_sheet,
    generate_latency_report,
    find_snmp_files,
)


def create_template(output_file="CMTS_Latency_Bin_Template.xlsx"):
    """Generate a fillable Excel template with formulas (no data)."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl required for template generation")
        return

    from snmp_collector import (
        _styled_cell, _HEADER_FONT, _HEADER_FILL, _CALC_FILL,
        _RESULT_FILL, _INPUT_FILL, _BOLD, _CENTER,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Latency Bins"

    ws.merge_cells("A1:I1")
    ws["A1"] = "CMTS DOWNSTREAM LATENCY BIN ANALYSIS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = _CENTER

    ws.merge_cells("A2:I2")
    ws["A2"] = "Yellow = input  |  Blue = calculated  |  Green = results"
    ws["A2"].font = Font(italic=True, size=10)

    headers = [
        "BIN", "LOWER (MS)", "UPPER (MS)", "AVG (MS)",
        "START COUNT", "END COUNT", "DELTA", "CUMULATIVE", "CUMULATIVE %",
    ]
    for col, h in enumerate(headers, 1):
        _styled_cell(ws, 4, col, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    start_row = 5
    for i in range(NUM_BINS):
        row = start_row + i
        low = BIN_EDGES_MS[i]
        high = BIN_EDGES_MS[i + 1]

        _styled_cell(ws, row, 1, i + 1)
        _styled_cell(ws, row, 2, low, fmt="0.00")
        _styled_cell(ws, row, 3, "200.00+" if i == 15 else high, fmt="0.00")
        avg_formula = f"=(B{row}+200)/2" if i == 15 else f"=(B{row}+C{row})/2"
        _styled_cell(ws, row, 4, avg_formula, fill=_CALC_FILL, fmt="0.0000")
        _styled_cell(ws, row, 5, 0, fill=_INPUT_FILL)
        _styled_cell(ws, row, 6, 0, fill=_INPUT_FILL)
        _styled_cell(ws, row, 7, f"=F{row}-E{row}", fill=_CALC_FILL)
        cum = f"=G{row}" if i == 0 else f"=H{row-1}+G{row}"
        _styled_cell(ws, row, 8, cum, fill=_CALC_FILL)
        end_row = start_row + 15
        _styled_cell(ws, row, 9, f"=IF(H${end_row}=0,0,H{row}/H${end_row}*100)", fill=_CALC_FILL, fmt="0.00")

    end_row = start_row + 15
    total_row = end_row + 2
    _styled_cell(ws, total_row, 1, "TOTAL", font=_BOLD)
    _styled_cell(ws, total_row, 7, f"=SUM(G{start_row}:G{end_row})", font=_BOLD, fill=_CALC_FILL)
    _styled_cell(ws, total_row, 9, "100.00", font=_BOLD, fill=_CALC_FILL)

    for i, w in enumerate([8, 14, 14, 14, 16, 16, 14, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_file)
    print(f"Template saved: {output_file}")


def print_usage():
    print("Usage:")
    print("  python latency_calculator.py <results_dir>")
    print("  python latency_calculator.py <before.txt> <after.txt> [output.xlsx]")
    print("  python latency_calculator.py --template [output.xlsx]")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print_usage()
        sys.exit(1)

    if sys.argv[1] == "--template":
        out = sys.argv[2] if len(sys.argv) > 2 else "CMTS_Latency_Bin_Template.xlsx"
        create_template(out)
        sys.exit(0)

    arg1 = sys.argv[1]

    if os.path.isdir(arg1):
        before_file, after_file = find_snmp_files(arg1)
        if not before_file or not after_file:
            print(f"ERROR: Could not find SNMP before/after .txt files in {arg1}")
            sys.exit(1)
        print(f"Before: {before_file}")
        print(f"After:  {after_file}")
        output = os.path.join(arg1, f"Latency_Bin_Report_{os.path.basename(os.path.abspath(arg1))}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        generate_latency_report(before_file, after_file, output)

    elif len(sys.argv) >= 3:
        before_file = arg1
        after_file = sys.argv[2]
        output = sys.argv[3] if len(sys.argv) > 3 else None
        generate_latency_report(before_file, after_file, output)

    else:
        print_usage()
        sys.exit(1)
