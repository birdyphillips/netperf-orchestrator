"""
Generates a fillable Excel template for CMTS downstream latency bin analysis.
16 bins with Start/End packet counts, Delta, Cumulative, Cumulative %, P50 & P99.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 17 edges = 16 bins. Bin 1: 0–0.05, Bin 2: 0.05–0.10, ... Bin 15: 150–200, Bin 16: 200+
BIN_EDGES_MS = [0, 0.05, 0.10, 0.25, 0.50, 1.00, 2.00, 5.00, 10.00,
                20.00, 30.00, 40.00, 50.00, 100.00, 150.00, 200.00, 500.00]

OUTPUT_FILE = "CMTS_Latency_Bin_Template_With_Avg_v2.xlsx"


def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Latency Bins"

    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF2CC")  # yellow fillable
    calc_fill = PatternFill("solid", fgColor="D9E2F3")   # blue calculated
    result_fill = PatternFill("solid", fgColor="C6EFCE")  # green results
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    center = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "CMTS DOWNSTREAM LATENCY BIN ANALYSIS"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    # Instructions
    ws.merge_cells("A2:I2")
    ws["A2"] = "Yellow cells = fillable input  |  Blue cells = auto-calculated  |  Green cells = results"
    ws["A2"].font = Font(italic=True, size=10)

    # Headers (row 4) — A:BIN, B:LOWER, C:UPPER, D:AVG, E:START, F:END, G:DELTA, H:CUMULATIVE, I:CUM%
    headers = ["BIN", "LOWER (MS)", "UPPER (MS)", "AVG (MS)", "START COUNT", "END COUNT",
               "DELTA", "CUMULATIVE", "CUMULATIVE %"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Bin rows (rows 5–20)
    start_row = 5
    for i in range(16):
        row = start_row + i
        low = BIN_EDGES_MS[i]
        high = BIN_EDGES_MS[i + 1]

        # A: Bin number
        cell = ws.cell(row=row, column=1, value=i + 1)
        cell.alignment = center
        cell.border = thin_border

        # B: Lower edge
        cell = ws.cell(row=row, column=2, value=low)
        cell.number_format = "0.00"
        cell.alignment = center
        cell.border = thin_border

        # C: Upper edge
        if i == 15:
            cell = ws.cell(row=row, column=3, value="200.00+")
        else:
            cell = ws.cell(row=row, column=3, value=high)
        cell.number_format = "0.00"
        cell.alignment = center
        cell.border = thin_border

        # D: AVG (MS) = average of lower and upper
        cell = ws.cell(row=row, column=4)
        if i == 15:
            cell.value = f"=(B{row}+200)/2"
        else:
            cell.value = f"=(B{row}+C{row})/2"
        cell.number_format = "0.0000"
        cell.fill = calc_fill
        cell.alignment = center
        cell.border = thin_border

        # E: Start Count (fillable)
        cell = ws.cell(row=row, column=5, value=0)
        cell.fill = input_fill
        cell.alignment = center
        cell.border = thin_border

        # F: End Count (fillable)
        cell = ws.cell(row=row, column=6, value=0)
        cell.fill = input_fill
        cell.alignment = center
        cell.border = thin_border

        # G: Delta = End - Start
        cell = ws.cell(row=row, column=7)
        cell.value = f"=F{row}-E{row}"
        cell.fill = calc_fill
        cell.alignment = center
        cell.border = thin_border

        # H: Cumulative
        cell = ws.cell(row=row, column=8)
        if i == 0:
            cell.value = f"=G{row}"
        else:
            cell.value = f"=H{row-1}+G{row}"
        cell.fill = calc_fill
        cell.alignment = center
        cell.border = thin_border

        # I: Cumulative %
        cell = ws.cell(row=row, column=9)
        cell.value = f'=IF(H$20=0,0,H{row}/H$20*100)'
        cell.number_format = "0.00"
        cell.fill = calc_fill
        cell.alignment = center
        cell.border = thin_border

    end_row = start_row + 15  # row 20

    # Totals row (row 22)
    total_row = end_row + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = header_font
    ws.cell(row=total_row, column=1).border = thin_border

    cell = ws.cell(row=total_row, column=7)
    cell.value = f"=SUM(G{start_row}:G{end_row})"
    cell.font = header_font
    cell.fill = calc_fill
    cell.alignment = center
    cell.border = thin_border

    cell = ws.cell(row=total_row, column=9, value="100.00")
    cell.font = header_font
    cell.fill = calc_fill
    cell.alignment = center
    cell.border = thin_border

    # Results section (row 24+)
    result_row = total_row + 2
    ws.merge_cells(f"A{result_row}:I{result_row}")
    ws.cell(row=result_row, column=1, value="PERCENTILE RESULTS (LINEAR INTERPOLATION)").font = Font(bold=True, size=12)

    # P50
    r = result_row + 1
    ws.cell(row=r, column=1, value="P50 TARGET").font = header_font
    cell = ws.cell(row=r, column=2)
    cell.value = f"=H{end_row}*0.50"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r, column=3, value="P50 (MS)").font = header_font
    cell = ws.cell(row=r, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_interpolation_formula(start_row, end_row, 0.50)

    # P99
    r = result_row + 2
    ws.cell(row=r, column=1, value="P99 TARGET").font = header_font
    cell = ws.cell(row=r, column=2)
    cell.value = f"=H{end_row}*0.99"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r, column=3, value="P99 (MS)").font = header_font
    cell = ws.cell(row=r, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_interpolation_formula(start_row, end_row, 0.99)

    # P99.9
    r = result_row + 3
    ws.cell(row=r, column=1, value="P99.9 TARGET").font = header_font
    cell = ws.cell(row=r, column=2)
    cell.value = f"=H{end_row}*0.999"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r, column=3, value="P99.9 (MS)").font = header_font
    cell = ws.cell(row=r, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_interpolation_formula(start_row, end_row, 0.999)

    # Formula legend
    r = result_row + 5
    ws.cell(row=r, column=1, value="FORMULA:").font = header_font
    ws.merge_cells(f"B{r}:I{r}")
    ws.cell(row=r, column=2, value="P = BIN_LOW + ((TARGET - PREV_CUMULATIVE) / BIN_COUNT) × (BIN_HIGH - BIN_LOW)")
    ws.cell(row=r, column=2).font = Font(italic=True, size=10)

    # --- AVG METHOD section ---
    avg_section_row = r + 2
    ws.merge_cells(f"A{avg_section_row}:I{avg_section_row}")
    ws.cell(row=avg_section_row, column=1, value="PERCENTILE RESULTS (AVG METHOD)").font = Font(bold=True, size=12)

    # P50 AVG
    r2 = avg_section_row + 1
    ws.cell(row=r2, column=1, value="P50 TARGET").font = header_font
    cell = ws.cell(row=r2, column=2)
    cell.value = f"=H{end_row}*0.50"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r2, column=3, value="P50 AVG (MS)").font = header_font
    cell = ws.cell(row=r2, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_avg_interpolation_formula(start_row, end_row, 0.50)

    # P99 AVG
    r2 = avg_section_row + 2
    ws.cell(row=r2, column=1, value="P99 TARGET").font = header_font
    cell = ws.cell(row=r2, column=2)
    cell.value = f"=H{end_row}*0.99"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r2, column=3, value="P99 AVG (MS)").font = header_font
    cell = ws.cell(row=r2, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_avg_interpolation_formula(start_row, end_row, 0.99)

    # P99.9 AVG
    r2 = avg_section_row + 3
    ws.cell(row=r2, column=1, value="P99.9 TARGET").font = header_font
    cell = ws.cell(row=r2, column=2)
    cell.value = f"=H{end_row}*0.999"
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.00"

    ws.cell(row=r2, column=3, value="P99.9 AVG (MS)").font = header_font
    cell = ws.cell(row=r2, column=4)
    cell.fill = result_fill
    cell.alignment = center
    cell.border = thin_border
    cell.number_format = "0.0000"
    cell.value = _build_avg_interpolation_formula(start_row, end_row, 0.999)

    # AVG method formula legend
    r3 = r2 + 2
    ws.cell(row=r3, column=1, value="FORMULA:").font = header_font
    ws.merge_cells(f"B{r3}:I{r3}")
    ws.cell(row=r3, column=2, value="P = AVG (col D) of first bin where cumulative count >= percentile target")
    ws.cell(row=r3, column=2).font = Font(italic=True, size=10)

    # Column widths
    widths = [8, 14, 14, 14, 16, 16, 14, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(OUTPUT_FILE)
    print(f"Template saved: {OUTPUT_FILE}")


def _build_interpolation_formula(start_row, end_row, percentile):
    """Build nested IF formula to interpolate percentile from cumulative bins."""
    target = f"H{end_row}*{percentile}"
    parts = []
    for i in range(16):
        row = start_row + i
        prev_cum = f"H{row-1}" if i > 0 else "0"
        interp = f"B{row}+(({target}-{prev_cum})/IF(G{row}=0,1,G{row}))*(C{row}-B{row})"
        parts.append(f"IF(H{row}>={target},{interp},")

    formula = "".join(parts) + f"C{end_row}" + ")" * 16
    return f"={formula}"


def _build_avg_interpolation_formula(start_row, end_row, percentile):
    """Build nested IF formula that returns the AVG (col D) of the bin where cumulative first reaches the target."""
    target = f"H{end_row}*{percentile}"
    parts = []
    for i in range(16):
        row = start_row + i
        parts.append(f"IF(H{row}>={target},D{row},")

    formula = "".join(parts) + f"D{end_row}" + ")" * 16
    return f"={formula}"


if __name__ == "__main__":
    create_template()
