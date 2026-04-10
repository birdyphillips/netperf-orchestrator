#!/usr/bin/env python3
"""
modem_metrics.py - CLI tool for querying the modem_metrics PostgreSQL database.

Usage:
    ./modem_metrics.py tables                          List all tables
    ./modem_metrics.py describe <table>                Show table schema
    ./modem_metrics.py query <table> [options]         Query a table
    ./modem_metrics.py sql "<SQL>"                     Run arbitrary SQL
    ./modem_metrics.py count <table>                   Row count for a table
    ./modem_metrics.py vcmts-snapshot [options]         Snapshot all vCMTS tables

Query options:
    -n, --limit N          Number of rows (default: 10)
    -c, --columns COL,...  Comma-separated columns to select
    -w, --where CLAUSE     WHERE clause (e.g. "value > 100")
    -o, --order COL        ORDER BY column (default: timestamp/time DESC)
    --cm-mac MAC           Filter by CM MAC address
    --direction DIR        Filter by direction (us/ds)
    --sfid SFID            Filter by SFID
    --since TIMESTAMP      Filter rows after this timestamp (e.g. "2024-01-01")
    --csv                  Output as CSV

vCMTS Snapshot options:
    --since TIMESTAMP      Filter rows after this timestamp (required)
    --until TIMESTAMP      Filter rows before this timestamp
    --cm-mac MAC           Filter by CM MAC address
    -o, --outdir DIR       Output directory (default: ./vcmts_snapshot)
"""

import argparse
import csv
import io
import os
import sys
from collections import defaultdict
from datetime import datetime

import psycopg2
from psycopg2.extras import RealDictCursor

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None


def normalize_mac(mac):
    """Convert any MAC format (0cb9.3764.3ab0, 0c-b9-37-64-3a-b0, etc.) to colon-separated lowercase."""
    raw = mac.lower().replace(".", "").replace(":", "").replace("-", "")
    return ":".join(raw[i:i+2] for i in range(0, 12, 2))

DB_CONFIG = {
    "host": "172.30.80.33",
    "port": 5433,
    "user": "postgres",
    "password": "postgres",
    "dbname": "modem_metrics",
}

CM_METRICS_TABLES = [
    "flow_stats", "histogram_stats", "vcmts_sfid_info",
    "agg_flow_stats", "congestion_stats",
]

VCMTS_METRICS_TABLES = [
    "channel_utilization", "cm_reg_status_config",
    "dp_flow_aqm_dropped_packets", "dp_flow_aqm_marked_congested_packets",
    "dp_flow_queue_latency_bin_pkt_count", "dp_flow_queue_latency_max_usec",
    "dp_flow_sanctioned_packets", "ksamis1_delta_packets_dropped",
    "ksamis1_service_time_created", "qos_service_flow_packets",
    "qos_service_flow_octets",
]

UNUSED_TABLES = ["cable_modem_metrics", "cable_modems", "cmts_metrics"]


def get_connection():
    return psycopg2.connect(**DB_CONFIG)


def get_time_column(table):
    """Return the timestamp/time column name for a given table."""
    if table in CM_METRICS_TABLES:
        return "time"
    return "timestamp"


def list_tables(conn):
    cur = conn.cursor()
    cur.execute(
        "SELECT table_name FROM information_schema.tables "
        "WHERE table_schema='public' ORDER BY table_name"
    )
    tables = [r[0] for r in cur.fetchall()]

    print("CM Metrics Tables:")
    for t in tables:
        if t in CM_METRICS_TABLES:
            print(f"  {t}")

    print("\nvCMTS Metrics Tables:")
    for t in tables:
        if t in VCMTS_METRICS_TABLES:
            print(f"  {t}")

    print("\nUnused Tables:")
    for t in tables:
        if t in UNUSED_TABLES:
            print(f"  {t}")

    other = [t for t in tables if t not in CM_METRICS_TABLES + VCMTS_METRICS_TABLES + UNUSED_TABLES]
    if other:
        print("\nOther Tables:")
        for t in other:
            print(f"  {t}")


def describe_table(conn, table):
    cur = conn.cursor()
    cur.execute(
        "SELECT column_name, data_type, is_nullable "
        "FROM information_schema.columns "
        "WHERE table_schema='public' AND table_name=%s "
        "ORDER BY ordinal_position",
        (table,),
    )
    rows = cur.fetchall()
    if not rows:
        print(f"Table '{table}' not found.")
        return
    print(f"\n  Table: {table}")
    print(f"  {'Column':<30} {'Type':<35} {'Nullable'}")
    print(f"  {'-'*30} {'-'*35} {'-'*8}")
    for col, dtype, nullable in rows:
        print(f"  {col:<30} {dtype:<35} {nullable}")


def count_table(conn, table):
    cur = conn.cursor()
    cur.execute(f'SELECT COUNT(*) FROM "{table}"')
    print(f"{table}: {cur.fetchone()[0]} rows")


def format_output(headers, rows, as_csv=False):
    if as_csv:
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        print(buf.getvalue(), end="")
        return

    if not rows:
        print("(0 rows)")
        return

    str_rows = [[str(v) if v is not None else "NULL" for v in row] for row in rows]
    widths = [max(len(h), *(len(r[i]) for r in str_rows)) for i, h in enumerate(headers)]

    hdr = " | ".join(h.ljust(w) for h, w in zip(headers, widths))
    sep = "-+-".join("-" * w for w in widths)
    print(hdr)
    print(sep)
    for row in str_rows:
        print(" | ".join(v.ljust(w) for v, w in zip(row, widths)))
    print(f"({len(rows)} rows)")


def query_table(conn, table, args):
    columns = args.columns if args.columns else "*"
    time_col = get_time_column(table)
    order = args.order if args.order else f"{time_col} DESC NULLS LAST"
    limit = args.limit

    conditions = []
    params = []

    if args.where:
        conditions.append(f"({args.where})")
    if args.cm_mac:
        mac_col = "mac" if table == "vcmts_sfid_info" else "cm_mac" if table == "cm_reg_status_config" else "cm_mac_addr"
        conditions.append(f"{mac_col} = %s")
        params.append(args.cm_mac)
    if args.direction:
        conditions.append("direction = %s")
        params.append(args.direction)
    if args.sfid:
        sfid_col = "sfid" if table in CM_METRICS_TABLES else "sf_index"
        conditions.append(f"{sfid_col} = %s")
        params.append(args.sfid)
    if args.since:
        conditions.append(f'"{time_col}" >= %s')
        params.append(args.since)

    where = (" WHERE " + " AND ".join(conditions)) if conditions else ""
    sql = f'SELECT {columns} FROM "{table}"{where} ORDER BY {order} LIMIT {limit}'

    cur = conn.cursor()
    cur.execute(sql, params)
    headers = [desc[0] for desc in cur.description]
    rows = cur.fetchall()
    format_output(headers, rows, as_csv=args.csv)


# Table -> MAC column (None = no MAC column)
MAC_COL_MAP = {t: "cm_mac_addr" for t in VCMTS_METRICS_TABLES}
MAC_COL_MAP["cm_reg_status_config"] = "cm_mac"
MAC_COL_MAP["channel_utilization"] = None
MAC_COL_MAP["qos_service_flow_octets"] = "cm_mac_addr"

# vCMTS table -> grouping key columns (excluding timestamp and value)
VCMTS_GROUP_KEYS = {
    "channel_utilization": ["cluster", "node", "pod", "namespace", "fn_name", "if_name", "md_name", "rpd_name"],
    "cm_reg_status_config": None,  # no value column, skip delta
    "dp_flow_queue_latency_bin_pkt_count": ["cm_mac_addr", "direction", "sf_index", "bin_num"],
}
VCMTS_DEFAULT_GROUP_KEYS = ["cm_mac_addr", "direction", "sf_index"]

# CM metrics table -> value columns for delta calculation
CM_VALUE_COLS = {
    "flow_stats": ["pkts", "octets", "timecreated", "timeactive", "phsunkowns", "policeddroppkts", "policeddelaypkts", "aqmdroppedpkts"],
    "histogram_stats": ["maxlatency", "numhistupdates", "bin1pkts", "bin2pkts", "bin3pkts", "bin4pkts", "bin5pkts", "bin6pkts", "bin7pkts", "bin8pkts", "bin9pkts", "bin10pkts", "bin11pkts", "bin12pkts", "bin13pkts", "bin14pkts", "bin15pkts", "bin16pkts"],
    "agg_flow_stats": ["pkts", "octets", "timecreated", "timeactive"],
    "congestion_stats": ["sanctioned", "totalect0pkts", "totalect1pkts", "cemarkpkts", "cearrivepkts"],
}


POLL_INTERVAL_SEC = 30  # 2x vCMTS polling interval (15s) to safely capture bracketing samples


def _get_vcmts_edge_values(conn, table, timestamp, mac_filter, edge="start"):
    """Get the value closest to a timestamp per group key for vCMTS tables.
    Widens by POLL_INTERVAL_SEC to capture the bracketing sample."""
    group_keys = VCMTS_GROUP_KEYS.get(table, VCMTS_DEFAULT_GROUP_KEYS)
    key_cols = ", ".join(f'"{k}"' for k in group_keys)
    mac_col = MAC_COL_MAP[table]

    if edge == "start":
        order = '"timestamp" DESC'
        sql = (
            f'SELECT DISTINCT ON ({key_cols}) {key_cols}, "timestamp", value '
            f'FROM "{table}" WHERE "timestamp" <= %s AND "timestamp" >= %s::timestamp - interval \'{POLL_INTERVAL_SEC} seconds\''
        )
    else:
        order = '"timestamp" ASC'
        sql = (
            f'SELECT DISTINCT ON ({key_cols}) {key_cols}, "timestamp", value '
            f'FROM "{table}" WHERE "timestamp" >= %s AND "timestamp" <= %s::timestamp + interval \'{POLL_INTERVAL_SEC} seconds\''
        )
    params = [timestamp, timestamp]

    if mac_filter and mac_col:
        sql += f' AND "{mac_col}" = %s'
        params.append(mac_filter)

    sql += f" ORDER BY {key_cols}, {order}"

    cur = conn.cursor()
    cur.execute(sql, params)
    return cur.fetchall(), [desc[0] for desc in cur.description]


def _get_cm_edge_values(conn, table, timestamp, sfids, edge="start"):
    """Get the row closest to a timestamp per SFID for CM metrics tables.
    Widens by POLL_INTERVAL_SEC to capture the bracketing sample."""
    val_cols = CM_VALUE_COLS[table]
    col_list = ", ".join(f'"{c}"' for c in val_cols)

    if edge == "start":
        order = '"time" DESC'
        sql = (
            f'SELECT DISTINCT ON (sfid) sfid, "time", {col_list} '
            f'FROM "{table}" WHERE "time" <= %s AND "time" >= %s::timestamp - interval \'{POLL_INTERVAL_SEC} seconds\''
        )
    else:
        order = '"time" ASC'
        sql = (
            f'SELECT DISTINCT ON (sfid) sfid, "time", {col_list} '
            f'FROM "{table}" WHERE "time" >= %s AND "time" <= %s::timestamp + interval \'{POLL_INTERVAL_SEC} seconds\''
        )
    params = [timestamp, timestamp]

    if sfids:
        sql += f" AND sfid IN ({','.join(['%s'] * len(sfids))})"
        params.extend(sfids)

    sql += f" ORDER BY sfid, {order}"

    cur = conn.cursor()
    cur.execute(sql, params)
    return cur.fetchall()


def _get_sfids_for_mac(conn, mac_filter):
    """Look up SFIDs and names for a MAC address from vcmts_sfid_info."""
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT sfid, name FROM vcmts_sfid_info WHERE mac = %s ORDER BY sfid",
        (mac_filter,),
    )
    return cur.fetchall()  # [(sfid, name), ...]


def _write_sheet(wb, outdir, table, headers, rows):
    """Write delta rows to CSV and Excel sheet. Returns row count."""
    csv_path = os.path.join(outdir, f"{table}.csv")
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)

    ws = wb.create_sheet(title=table[:31])
    ws.append(headers)
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])

    print(f"  {table}: {len(rows)} rows -> {csv_path}")
    return len(rows)


def vcmts_snapshot(conn, args):
    """Compute deltas between start/end timestamps for all vCMTS + CM metrics tables."""
    if Workbook is None:
        print("Error: openpyxl is required for Excel export. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    outdir = args.outdir
    os.makedirs(outdir, exist_ok=True)

    mac_filter = normalize_mac(args.cm_mac) if args.cm_mac else None
    timestamp_label = args.since.replace(" ", "_").replace(":", "-")
    wb = Workbook()
    wb.remove(wb.active)

    total_rows = 0

    # --- vCMTS Metrics ---
    print("vCMTS Metrics:")
    for table in VCMTS_METRICS_TABLES:
        group_keys = VCMTS_GROUP_KEYS.get(table, VCMTS_DEFAULT_GROUP_KEYS)

        if group_keys is None:
            mac_col = MAC_COL_MAP[table]
            sql = f'SELECT * FROM "{table}" WHERE "timestamp" >= %s AND "timestamp" <= %s'
            params = [args.since, args.until]
            if mac_filter and mac_col:
                sql += f' AND "{mac_col}" = %s'
                params.append(mac_filter)
            sql += ' ORDER BY "timestamp" ASC'
            cur = conn.cursor()
            cur.execute(sql, params)
            delta_headers = [desc[0] for desc in cur.description]
            delta_rows = cur.fetchall()
        else:
            start_rows, _ = _get_vcmts_edge_values(conn, table, args.since, mac_filter, "start")
            end_rows, _ = _get_vcmts_edge_values(conn, table, args.until, mac_filter, "end")

            n_keys = len(group_keys)
            start_map = {row[:n_keys]: row[-1] for row in start_rows}
            end_map = {row[:n_keys]: row[-1] for row in end_rows}

            all_keys = sorted(set(start_map) | set(end_map))
            delta_headers = group_keys + ["start_value", "end_value", "delta"]
            delta_rows = []
            for key in all_keys:
                sv = start_map.get(key)
                ev = end_map.get(key)
                delta = (ev or 0) - (sv or 0)
                delta_rows.append(list(key) + [sv, ev, delta])

        total_rows += _write_sheet(wb, outdir, table, delta_headers, delta_rows)

    # --- CM Metrics ---
    print("\nCM Metrics:")
    sfid_map = {}  # sfid -> name
    sfids = []
    if mac_filter:
        sfid_rows = _get_sfids_for_mac(conn, mac_filter)
        sfids = [r[0] for r in sfid_rows]
        sfid_map = {r[0]: r[1] for r in sfid_rows}
        if not sfids:
            print("  (no SFIDs found for this MAC in vcmts_sfid_info)")

    # vcmts_sfid_info — export as reference
    if sfid_map:
        sfid_headers = ["sfid", "name", "direction"]
        sfid_rows_out = []
        for sfid, name in sorted(sfid_map.items()):
            direction = "US" if name.startswith("us") else "DS" if name.startswith("ds") else "?"
            sfid_rows_out.append([sfid, name, direction])
        total_rows += _write_sheet(wb, outdir, "vcmts_sfid_info", sfid_headers, sfid_rows_out)

    for table in CM_VALUE_COLS:
        val_cols = CM_VALUE_COLS[table]
        start_rows = _get_cm_edge_values(conn, table, args.since, sfids, "start")
        end_rows = _get_cm_edge_values(conn, table, args.until, sfids, "end")

        # row format: (sfid, time, val1, val2, ...)
        start_map = {row[0]: row[2:] for row in start_rows}
        end_map = {row[0]: row[2:] for row in end_rows}

        all_sfids = sorted(set(start_map) | set(end_map))
        delta_headers = ["sfid", "name", "direction"]
        for col in val_cols:
            delta_headers.extend([f"{col}_start", f"{col}_end", f"{col}_delta"])

        delta_rows = []
        for sfid in all_sfids:
            name = sfid_map.get(sfid, "")
            direction = "US" if name.startswith("us") else "DS" if name.startswith("ds") else "?"
            row = [sfid, name, direction]
            sv_vals = start_map.get(sfid)
            ev_vals = end_map.get(sfid)
            for i in range(len(val_cols)):
                sv = sv_vals[i] if sv_vals else None
                ev = ev_vals[i] if ev_vals else None
                delta = (ev or 0) - (sv or 0)
                row.extend([sv, ev, delta])
            delta_rows.append(row)

        total_rows += _write_sheet(wb, outdir, table, delta_headers, delta_rows)

    # --- Summary Sheet ---
    print("\nGenerating summary...")
    summary_headers = [
        "direction", "sf_index",
        "total_packets", "total_octets", "total_bits",
        "throughput_bps", "throughput_kbps", "throughput_mbps",
        "avg_latency_ms", "p99_latency_ms", "max_latency_usec",
        "aqm_dropped_pkts", "aqm_marked_congested_pkts",
        "sanctioned_pkts", "ksamis1_dropped_pkts",
    ]
    summary_rows = []

    # Compute time window in seconds
    try:
        t_start = datetime.fromisoformat(args.since.replace("T", " "))
        t_end = datetime.fromisoformat(args.until.replace("T", " "))
        duration_sec = (t_end - t_start).total_seconds()
    except Exception:
        duration_sec = None

    # Collect deltas from vCMTS tables keyed by (direction, sf_index)
    def _get_delta_map(conn, table, mac_filter, since, until):
        """Return {(direction, sf_index): delta} for a single-value vCMTS table."""
        start_rows, _ = _get_vcmts_edge_values(conn, table, since, mac_filter, "start")
        end_rows, _ = _get_vcmts_edge_values(conn, table, until, mac_filter, "end")
        # keys: (cm_mac_addr, direction, sf_index) -> value
        start_map = {(r[1], r[2]): r[-1] for r in start_rows}  # skip cm_mac_addr
        end_map = {(r[1], r[2]): r[-1] for r in end_rows}
        return {k: (end_map.get(k, 0) or 0) - (start_map.get(k, 0) or 0)
                for k in set(start_map) | set(end_map)}

    packets_delta = _get_delta_map(conn, "qos_service_flow_packets", mac_filter, args.since, args.until)
    octets_delta = _get_delta_map(conn, "qos_service_flow_octets", mac_filter, args.since, args.until)
    aqm_dropped = _get_delta_map(conn, "dp_flow_aqm_dropped_packets", mac_filter, args.since, args.until)
    aqm_marked = _get_delta_map(conn, "dp_flow_aqm_marked_congested_packets", mac_filter, args.since, args.until)
    sanctioned = _get_delta_map(conn, "dp_flow_sanctioned_packets", mac_filter, args.since, args.until)
    ksamis_dropped = _get_delta_map(conn, "ksamis1_delta_packets_dropped", mac_filter, args.since, args.until)

    # Max latency per (direction, sf_index)
    max_lat_start, _ = _get_vcmts_edge_values(conn, "dp_flow_queue_latency_max_usec", args.since, mac_filter, "start")
    max_lat_end, _ = _get_vcmts_edge_values(conn, "dp_flow_queue_latency_max_usec", args.until, mac_filter, "end")
    max_lat_map = {(r[1], r[2]): r[-1] for r in max_lat_end}  # use end value as peak
    # also check start in case it was higher
    for r in max_lat_start:
        k = (r[1], r[2])
        if k not in max_lat_map or (r[-1] or 0) > (max_lat_map[k] or 0):
            max_lat_map[k] = r[-1]

    # Latency histogram: compute avg and P99 from bin deltas
    # Get bin edges per (direction, bin_num)
    cur = conn.cursor()
    cur.execute(
        "SELECT DISTINCT direction, bin_num, lower_edge_msec, upper_edge_msec "
        "FROM dp_flow_queue_latency_bin_pkt_count ORDER BY direction, bin_num"
    )
    bin_edges = {}  # (direction, bin_num) -> (lower, upper)
    for d, bn, lo, hi in cur.fetchall():
        bin_edges[(d, bn)] = (lo, hi)

    # Get bin pkt count deltas
    grp = VCMTS_GROUP_KEYS["dp_flow_queue_latency_bin_pkt_count"]
    start_bins, _ = _get_vcmts_edge_values(conn, "dp_flow_queue_latency_bin_pkt_count", args.since, mac_filter, "start")
    end_bins, _ = _get_vcmts_edge_values(conn, "dp_flow_queue_latency_bin_pkt_count", args.until, mac_filter, "end")
    # key: (cm_mac_addr, direction, sf_index, bin_num) -> value
    sb_map = {r[:4]: r[-1] for r in start_bins}
    eb_map = {r[:4]: r[-1] for r in end_bins}
    # Aggregate: (direction, sf_index) -> [(bin_num, delta_count)]
    bin_deltas = defaultdict(list)
    for k in set(sb_map) | set(eb_map):
        delta = (eb_map.get(k, 0) or 0) - (sb_map.get(k, 0) or 0)
        direction, sf_index, bin_num = k[1], k[2], k[3]
        bin_deltas[(direction, sf_index)].append((bin_num, delta))

    def _calc_latency_stats(direction, sf_index):
        """Compute avg and P99 latency from histogram bin deltas."""
        bins = bin_deltas.get((direction, sf_index), [])
        if not bins:
            return None, None
        bins.sort(key=lambda x: x[0])
        total_pkts = sum(c for _, c in bins)
        if total_pkts <= 0:
            return 0.0, 0.0
        # Weighted average using bin midpoints
        weighted_sum = 0.0
        for bn, count in bins:
            lo, hi = bin_edges.get((direction, bn), (0, 0))
            if hi == float('inf'):
                mid = lo * 1.5  # estimate for open-ended bin
            else:
                mid = (lo + hi) / 2.0
            weighted_sum += mid * count
        avg = weighted_sum / total_pkts
        # P99: find bin where cumulative count >= 99% of total
        p99_threshold = total_pkts * 0.99
        cumulative = 0
        p99 = 0.0
        for bn, count in bins:
            cumulative += count
            lo, hi = bin_edges.get((direction, bn), (0, 0))
            if cumulative >= p99_threshold:
                p99 = hi if hi != float('inf') else lo * 1.5
                break
        return round(avg, 4), round(p99, 4)

    all_flows = sorted(set(packets_delta) | set(octets_delta) | set(aqm_dropped))
    for direction, sf_index in all_flows:
        k = (direction, sf_index)
        pkts = packets_delta.get(k, 0)
        octets = octets_delta.get(k, 0)
        total_bits = octets * 8 if octets else 0
        if duration_sec and duration_sec > 0 and octets:
            throughput_bps = round(total_bits / duration_sec, 2)
            throughput_kbps = round(throughput_bps / 1_000, 4)
            throughput_mbps = round(throughput_bps / 1_000_000, 4)
        else:
            throughput_bps = 0
            throughput_kbps = 0.0
            throughput_mbps = 0.0
        avg_lat, p99_lat = _calc_latency_stats(direction, sf_index)
        max_lat = max_lat_map.get(k)


        summary_rows.append([
            direction, sf_index,
            pkts, octets, total_bits,
            throughput_bps, throughput_kbps, throughput_mbps,
            avg_lat, p99_lat, max_lat,
            aqm_dropped.get(k, 0), aqm_marked.get(k, 0),
            sanctioned.get(k, 0), ksamis_dropped.get(k, 0),
        ])

    total_rows += _write_sheet(wb, outdir, "summary", summary_headers, summary_rows)

    # --- Style the summary sheet and move to front ---
    ws_summary = wb["summary"]
    wb.move_sheet(ws_summary, offset=-len(wb.sheetnames) + 1)

    # Colors
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    us_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")  # light blue
    ds_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
    zero_font = Font(name="Calibri", color="AAAAAA", size=10)
    data_font = Font(name="Calibri", size=10)
    number_font = Font(name="Calibri", size=10, bold=True, color="C00000")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    # Title row
    ws_summary.insert_rows(1, 3)
    title_text = f"vCMTS Metrics Summary"
    ws_summary.cell(row=1, column=1, value=title_text).font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
    ws_summary.cell(row=2, column=1, value=f"Window: {args.since} → {args.until}  ({duration_sec:.1f}s)").font = Font(name="Calibri", size=10, italic=True)
    mac_label = args.cm_mac if args.cm_mac else "all"
    ws_summary.cell(row=2, column=5, value=f"CM MAC: {mac_label}").font = Font(name="Calibri", size=10, italic=True)

    # Style header row (row 4 after inserting 3 title rows)
    header_row = 4
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws_summary.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Style data rows
    for row_idx in range(header_row + 1, header_row + 1 + len(summary_rows)):
        direction_val = ws_summary.cell(row=row_idx, column=1).value
        row_fill = us_fill if direction_val and "upstream" in str(direction_val) else ds_fill
        for col_idx in range(1, len(summary_headers) + 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            val = cell.value
            # Highlight non-zero drop/congestion columns (cols 12-15)
            if col_idx >= 12 and val and str(val) not in ("0", "0.0", "", "None"):
                cell.font = number_font
            elif val in (None, "", "0", "0.0", "NULL"):
                cell.font = zero_font
            else:
                cell.font = data_font

    # Auto-fit column widths
    for col_idx in range(1, len(summary_headers) + 1):
        max_len = len(str(summary_headers[col_idx - 1]))
        for row_idx in range(header_row + 1, header_row + 1 + len(summary_rows)):
            val = ws_summary.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(val)) if val else 0)
        ws_summary.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Freeze panes below header
    ws_summary.freeze_panes = f"A{header_row + 1}"

    # Print summary to terminal
    print(f"\n{'='*80}")
    print(f"  SUMMARY  (window: {args.since} -> {args.until}, {duration_sec:.1f}s)")
    print(f"{'='*80}")
    format_output(summary_headers, summary_rows)

    excel_path = os.path.join(outdir, f"vcmts_snapshot_{timestamp_label}.xlsx")
    wb.save(excel_path)
    print(f"\nTotal: {total_rows} rows across all tables")
    print(f"Excel workbook: {excel_path}")


def run_sql(conn, sql, as_csv=False):
    cur = conn.cursor()
    cur.execute(sql)
    if cur.description:
        headers = [desc[0] for desc in cur.description]
        rows = cur.fetchall()
        format_output(headers, rows, as_csv=as_csv)
    else:
        print(f"OK ({cur.rowcount} rows affected)")


def main():
    parser = argparse.ArgumentParser(description="Query the modem_metrics PostgreSQL database")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("tables", help="List all tables")

    p_desc = sub.add_parser("describe", help="Describe a table")
    p_desc.add_argument("table")

    p_count = sub.add_parser("count", help="Row count for a table")
    p_count.add_argument("table")

    p_query = sub.add_parser("query", help="Query a table")
    p_query.add_argument("table")
    p_query.add_argument("-n", "--limit", type=int, default=10)
    p_query.add_argument("-c", "--columns", help="Comma-separated columns")
    p_query.add_argument("-w", "--where", help="WHERE clause")
    p_query.add_argument("-o", "--order", help="ORDER BY clause")
    p_query.add_argument("--cm-mac", help="Filter by CM MAC address")
    p_query.add_argument("--direction", help="Filter by direction")
    p_query.add_argument("--sfid", type=int, help="Filter by SFID/sf_index")
    p_query.add_argument("--since", help="Filter rows after timestamp")
    p_query.add_argument("--csv", action="store_true", help="Output as CSV")

    p_sql = sub.add_parser("sql", help="Run arbitrary SQL")
    p_sql.add_argument("statement")
    p_sql.add_argument("--csv", action="store_true", help="Output as CSV")

    p_snap = sub.add_parser("vcmts-snapshot", help="Compute deltas for all vCMTS metrics between two timestamps")
    p_snap.add_argument("--since", required=True, help="Start timestamp")
    p_snap.add_argument("--until", required=True, help="End timestamp")
    p_snap.add_argument("--cm-mac", help="Filter by CM MAC address (any format)")
    p_snap.add_argument("-o", "--outdir", default="./vcmts_snapshot", help="Output directory")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return

    try:
        conn = get_connection()
    except psycopg2.OperationalError as e:
        print(f"Connection failed: {e}", file=sys.stderr)
        sys.exit(1)

    try:
        if args.command == "tables":
            list_tables(conn)
        elif args.command == "describe":
            describe_table(conn, args.table)
        elif args.command == "count":
            count_table(conn, args.table)
        elif args.command == "query":
            query_table(conn, args.table, args)
        elif args.command == "sql":
            run_sql(conn, args.statement, as_csv=args.csv)
        elif args.command == "vcmts-snapshot":
            vcmts_snapshot(conn, args)
    except psycopg2.Error as e:
        print(f"Database error: {e}", file=sys.stderr)
        sys.exit(1)
    finally:
        conn.close()


if __name__ == "__main__":
    main()
